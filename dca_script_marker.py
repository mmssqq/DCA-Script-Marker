# Copyright © 2026 马斯琪 Siqi Ma
# SPDX-License-Identifier: AGPL-3.0-or-later

import os
import copy
import fitz  # PyMuPDF
import re
import math
import unicodedata
import argparse
import html
import json
import platform
import sys
import tempfile
import warnings
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from datetime import date

TEMPLATE_FILE = "dca_template.xlsx"
PDF_FILE = "chinese_sample_script v3.pdf"
OUTPUT_FILE = "marked_script.pdf"
REPORT_FILE = "review_report.txt"

STATE_COLOUR = (0.0, 0.35, 0.75)
NUMBER_COLOUR = (0.85, 0.0, 0.35)
ANNOTATION_COLOURS = {
    "red": NUMBER_COLOUR,
    "blue": STATE_COLOUR,
    "black": (0.0, 0.0, 0.0),
    "green": (0.0, 0.45, 0.25),
    "orange": (0.78, 0.24, 0.0),
    "purple": (0.50, 0.20, 0.65),
    "grey": (0.35, 0.35, 0.35),
    "brown": (0.45, 0.25, 0.10),
}
ANNOTATION_COLOUR_CHOICES = tuple(ANNOTATION_COLOURS)
CHINESE_FONT_FILE = "/System/Library/Fonts/STHeiti Medium.ttc"
NUMBER_SCALE = 1.25
VISUAL_FRAGMENT_WORD_GAP = 4.0

# Some PDFs expose simplified CJK radicals instead of the ordinary character
# stored in the workbook. Keep this deliberately narrow so speaker matching is
# tolerant without making unrelated Chinese names compare as equal.
SPEAKER_CHARACTER_TRANSLATION = str.maketrans({
    "⻓": "长",
})

STAGE_DIRECTION_PREFIXES = ("(", "[", "{", "【", "〔")
CAST_TRACK_PREFIX = re.compile(
    r"^(?:"
    r"【\s*[A-Z](?:\s*[&/]\s*[A-Z])*\s*】"
    r"|"
    r"\[\s*[A-Z](?:\s*[&/]\s*[A-Z])*\s*\]"
    r")\s*"
)


def normalise(text):
    text = unicodedata.normalize("NFKC", str(text))
    return " ".join(text.strip().lower().split())


def cue_match_key(text):
    """Normalise a DCA State cue while allowing Chinese spacing variations."""
    cleaned = normalise(text)

    if any("\u4e00" <= character <= "\u9fff" for character in cleaned):
        return re.sub(r"\s+", "", cleaned)

    return cleaned


def cue_identifier(text):
    """Return a leading music/playback cue code such as M0 or PB3."""
    cleaned = unicodedata.normalize("NFKC", str(text)).strip().lower()
    match = re.match(r"^(pb|m)\s*(\d+)(?=$|[^0-9a-z_])", cleaned)

    if not match:
        return ""

    return f"{match.group(1)}{match.group(2)}"


def speaker_match_key(text):
    """Compare names while ignoring punctuation and European accents.

    This lets an Excel name such as ``ELODIE`` match a script cue printed as
    ``ÉLODIE``.  The original text remains unchanged in the marked PDF.
    """
    cleaned = normalise(text).translate(SPEAKER_CHARACTER_TRANSLATION)
    accent_free = "".join(
        character
        for character in unicodedata.normalize("NFD", cleaned)
        if unicodedata.category(character) != "Mn"
    )
    # The template may contain a label copied directly from a script, such as
    # ``林晓：`` or ``DR. Q.``. Colons are label punctuation, not part of
    # the character name, so ignore them during every speaker comparison.
    return re.sub(r"[\s·・.．。:：]", "", accent_free)


def speaker_base_key(text):
    """Match a speaker name while ignoring a cast-count note in brackets."""
    base_name = re.split(r"[（(]", str(text), maxsplit=1)[0]
    return speaker_match_key(base_name)


def contains_cjk(text):
    return any("\u4e00" <= character <= "\u9fff" for character in str(text))


def is_cjk_character(character):
    """Return true for a Han character used in a printed speaker name."""
    return (
        "\u3400" <= character <= "\u4dbf"
        or "\u4e00" <= character <= "\u9fff"
        or "\uf900" <= character <= "\ufaff"
    )


def has_spaced_cjk_speaker_prefix(text, speaker_name):
    """Recognise a Chinese name padded to a wider character column.

    Some scripts visually expand a two-character name to the width of a
    three-character name, for example ``顾  正 dialogue``. The first wide
    gap is part of the name, not the boundary before the dialogue. Require
    that wider internal padding plus a separate boundary after the complete
    known name so ordinary narration that merely begins with a character name
    stays rejected.
    """
    cleaned_text = unicodedata.normalize("NFKC", str(text)).translate(
        SPEAKER_CHARACTER_TRANSLATION
    )
    name_characters = [
        character
        for character in unicodedata.normalize(
            "NFKC",
            str(speaker_name),
        ).translate(SPEAKER_CHARACTER_TRANSLATION)
        if not character.isspace()
    ]

    if (
        len(name_characters) < 2
        or not all(is_cjk_character(character) for character in name_characters)
    ):
        return False

    spaced_name_pattern = r"\s+".join(
        re.escape(character) for character in name_characters
    )
    prefix_match = re.match(
        rf"^\s*({spaced_name_pattern})(?=\s+|[(:：])",
        cleaned_text,
    )
    return bool(
        prefix_match
        and re.search(r"\s{2,}", prefix_match.group(1))
    )


def css_colour(colour):
    """Convert a PyMuPDF RGB tuple into a CSS hexadecimal colour."""
    channels = (
        round(max(0.0, min(1.0, channel)) * 255)
        for channel in colour
    )
    return "#" + "".join(f"{channel:02x}" for channel in channels)


def rich_text_font_family(font_name, font_file):
    """Return a safe CSS family for a rich FreeText annotation."""
    if font_file:
        return "sans-serif"

    return {
        "helv": "Helvetica",
        "hebo": "Helvetica",
        "tiro": "Times New Roman",
        "tibo": "Times New Roman",
        "cour": "Courier New",
        "cobo": "Courier New",
    }.get(font_name, "Helvetica")


def starts_with_stage_direction(text):
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    return cleaned.startswith(STAGE_DIRECTION_PREFIXES)


def strip_cast_track_prefix(text):
    """Remove a bounded actor-track tag from an explicit speaker cue.

    Some translated rehearsal scripts prefix every dialogue label with the
    actor or doubling track, for example ``【A】林青：`` or
    ``【A/B】林青、周岚：``. Ordinary bracketed text remains a stage
    direction: only one-letter track codes joined by ``&`` or ``/`` are
    accepted, and the following text must contain an explicit colon.
    """
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    match = CAST_TRACK_PREFIX.match(cleaned)
    if match is None:
        return cleaned

    remainder = cleaned[match.end():].lstrip()
    if re.search(r"[：:]", remainder) is None:
        return cleaned

    return remainder


def is_untagged_punctuationless_speaker_label(text, speaker_names):
    """Return true for an ambiguous bare name outside a cast-tag cue."""
    if len(speaker_names) != 1:
        return False

    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    if strip_cast_track_prefix(cleaned) != cleaned:
        return False

    return (
        speaker_match_key(cleaned)
        == speaker_match_key(speaker_names[0])
    )


def looks_like_speaker_label(text, speaker_name):
    """Avoid marking a character name used inside dialogue or narration."""
    original_text = unicodedata.normalize("NFKC", str(text)).strip()
    text = strip_cast_track_prefix(original_text)

    # A cast-track tag is metadata only when the complete text before the
    # first colon is a known speaker label. This keeps a tagged action such as
    # ``【A】林青 走近门口：他停下。`` from becoming a 林青 dialogue cue.
    if text != original_text:
        separator = re.search(r"[：:]", text)
        tagged_speakers = (
            parse_explicit_speaker_prefix(
                text[:separator.start()].strip(),
                {speaker_name},
            )
            if separator is not None
            else []
        )
        if tagged_speakers != [speaker_name]:
            return False

    # A name inside a bracketed action is not a dialogue label. Check this
    # before accepting explicit colons so title-case cues such as ``Robin:`` can
    # be supported without turning ``[Robin: enters]`` into spoken dialogue.
    if starts_with_stage_direction(text):
        return False

    english_name = not contains_cjk(speaker_name)

    def has_english_cue_case(value):
        """English speaker labels are conventionally printed in capitals."""
        letters = "".join(letter for letter in value if letter.isalpha())
        return bool(letters) and letters.isupper()

    # A colon is a speaker separator only when the text *before* it is the
    # speaker name.  Dialogue can contain a later colon ("Now:",
    # "[MUSIC: ...]"), so it must not prevent the ordinary full-stop label
    # checks below from recognising ALEX., BLAIR., and similar cues.
    if "：" in text or ":" in text:
        label = re.split(r"[：:]", text, maxsplit=1)[0].strip()
        if speaker_match_key(label) == speaker_match_key(speaker_name):
            # An exact name followed by a colon is unambiguous enough to allow
            # title-case English labels such as Robin: and Avery：. Less explicit
            # English layouts below retain the uppercase safety requirement.
            return True

    if speaker_match_key(text) == speaker_match_key(speaker_name):
        return not english_name or has_english_cue_case(text)

    # Preserve abbreviations inside a name, such as ``DR. Q.``, while using
    # the final full stop / colon as the speaker-label boundary. This avoids
    # confusing an action such as ``DR. Q appears`` with spoken dialogue.
    escaped_speaker = re.escape(str(speaker_name))
    punctuated_label = re.match(
        rf"^\s*{escaped_speaker}(?=$|[.:：])",
        text,
        flags=re.IGNORECASE,
    )
    if punctuated_label:
        return (
            not english_name
            or has_english_cue_case(punctuated_label.group(0))
        )

    # English scripts commonly use a full stop after a multi-word speaker
    # label: ``FIRST GUARD. Rowan, over here.``  The old check below considered
    # only the first word (``FIRST``), which meant these valid labels were
    # skipped while one-word labels such as ``ROWAN.`` worked.
    speaker_prefix = re.split(r"[.:：]", text, maxsplit=1)[0].strip()
    if speaker_match_key(speaker_prefix) == speaker_match_key(speaker_name):
        return (
            not english_name
            or has_english_cue_case(speaker_prefix)
        )

    # A script may write a name with a different middle dot, then a single
    # space before the dialogue: 林.海 你是怎么发现的？
    # Compare only that first word using the punctuation-insensitive key.
    first_word = text.split(maxsplit=1)[0] if text else ""
    # A shared English label often begins ``ALEX, BLAIR, CASEY...``.
    # Treat the comma after the first name as a label separator, not as part
    # of that name.
    first_word = first_word.rstrip(",;；")
    if speaker_match_key(first_word) == speaker_match_key(speaker_name):
        return not english_name or has_english_cue_case(first_word)

    # Slash-separated and ampersand-separated group labels keep all names in
    # one visual word, for example ``ALEX/BLAIR/CASEY/DREW.``.  The first
    # component still establishes that this row is a speaker label.
    leading_component = re.split(
        r"[,/&＋+、/／;；]",
        first_word,
        maxsplit=1,
    )[0].rstrip(".")
    if speaker_match_key(leading_component) == speaker_match_key(speaker_name):
        return (
            not english_name
            or has_english_cue_case(leading_component)
        )

    # The first name in a shared label can itself contain spaces, such as
    # ``TWO GUARDS & CHORUS.``. In that layout `first_word` is only TWO, so
    # also inspect the complete text before the first group
    # separator. This remains safe for dialogue because an ordinary cue like
    # ``ALEX. (to BLAIR & CASEY)`` was already recognised by its full stop.
    leading_group_name = re.split(
        r"[,/&＋+、/／;；]",
        text,
        maxsplit=1,
    )[0].strip().rstrip(".")
    if speaker_match_key(leading_group_name) == speaker_match_key(speaker_name):
        return (
            not english_name
            or has_english_cue_case(leading_group_name)
        )

    # Group labels are often printed as, for example, 歌队（8位）.
    # The cast-count note is not part of the character name in the template.
    if speaker_base_key(first_word) == speaker_match_key(speaker_name):
        return not english_name or has_english_cue_case(first_word)

    # A lead speaker can be followed by a group count, for example
    # 林青+5群众. The count describes the group; it is not part of the name.
    group_label = re.match(r"^(.*?)\s*\+\s*\d+", first_word)
    if (
        group_label
        and speaker_match_key(group_label.group(1))
        == speaker_match_key(speaker_name)
    ):
        return not english_name or has_english_cue_case(first_word)

    # A writer may pad a two-character Chinese name to the same visual width
    # as a three-character name. In extracted PDF text, ``顾  正  dialogue``
    # therefore contains a wide gap inside the speaker name, while the gap
    # after it may be emitted as one or more PDF spaces. Validate the complete
    # known name before using the first wide gap as the speaker/dialogue
    # boundary.
    if has_spaced_cjk_speaker_prefix(text, speaker_name):
        return True

    # Chinese theatre scripts normally have a wide gap after a speaker label.
    label = re.split(r"\s{2,}", text, maxsplit=1)[0]
    if speaker_match_key(label) == speaker_match_key(speaker_name):
        return not english_name or has_english_cue_case(label)

    # Some English-style layouts use one ordinary space after the name.
    return (
        not english_name
        and normalise(text).startswith(normalise(speaker_name) + " ")
    )


def get_speaker_name(text, possible_characters):
    original_text = unicodedata.normalize("NFKC", str(text)).strip()
    text = strip_cast_track_prefix(original_text)
    if text != original_text:
        tagged_speakers = get_explicit_speaker_names(
            original_text,
            possible_characters,
        )
        return tagged_speakers[0] if tagged_speakers else None
    clean_text = normalise(text)

    for separator in ("：", ":"):
        if separator in text:
            possible_name = text.split(
                separator, 1
            )[0].strip()

            full_name = normalise(possible_name)
            if full_name in possible_characters:
                return full_name

            compact_name = speaker_match_key(possible_name)
            for character in possible_characters:
                if speaker_match_key(character) == compact_name:
                    return character

            base_name = normalise(
                re.split(r"[（(]", possible_name, maxsplit=1)[0]
            )
            base_key = speaker_match_key(base_name)
            for character in possible_characters:
                if speaker_match_key(character) == base_key:
                    return character
            # This colon belongs to the dialogue, not to a speaker label.
            # Fall through to the normal beginning-of-line name matching.
            break

    compact_text = speaker_match_key(clean_text)
    for character in sorted(possible_characters, key=len, reverse=True):
        if compact_text.startswith(speaker_match_key(character)):
            return character

    for character in sorted(possible_characters, key=len, reverse=True):
        if clean_text.startswith(character):
            return character

    for character in possible_characters:
        if clean_text.startswith(character + "("):
            return character

    for character in sorted(
        possible_characters, key=len, reverse=True
    ):
        if clean_text == character:
            return character

        if (
            any("\u4e00" <= letter <= "\u9fff" for letter in character)
            and clean_text.startswith(character + " ")
        ):
            return character

    spaced_name = re.match(r"^(\S+)\s{2,}", text)

    if spaced_name:
        return normalise(spaced_name.group(1))

    return clean_text


def match_known_speaker_label(text, possible_characters):
    """Return the workbook spelling for one complete speaker label."""
    label_key = speaker_match_key(text)
    if not label_key:
        return ""

    return next(
        (
            character
            for character in sorted(
                possible_characters,
                key=len,
                reverse=True,
            )
            if speaker_match_key(character) == label_key
        ),
        "",
    )


def strip_speaker_delivery_note(text):
    """Remove a trailing delivery note such as ``（唱）`` from a label."""
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()

    while True:
        without_note = re.sub(r"\s*\([^()]*\)\s*$", "", cleaned)
        if without_note == cleaned:
            return cleaned
        cleaned = without_note.strip()


def speaker_delivery_suffix_key(text):
    """Normalise a short delivery instruction after a speaker comma."""
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    cleaned = re.split(r"[：:]", cleaned, maxsplit=1)[0].strip()
    parenthesised = re.fullmatch(r"\(([^()]*)\)", cleaned)
    if parenthesised:
        cleaned = parenthesised.group(1).strip()
    return normalise(cleaned.rstrip(".。:："))


def is_speaker_delivery_suffix(text):
    """Return true for a compact singing or narration instruction."""
    return speaker_delivery_suffix_key(text) in {
        "唱",
        "合唱",
        "齐唱",
        "独唱",
        "重唱",
        "领唱",
        "念白",
        "旁白",
        "chorus",
        "ensemble",
    }


def split_known_speaker_group(text, possible_characters):
    """Parse a complete shared label only when every member is known."""
    # Preserve an exact workbook name before interpreting characters such as
    # 和, 与, or / as separators. They can legitimately occur inside a role
    # name copied verbatim from the script.
    exact_name = match_known_speaker_label(text, possible_characters)
    if exact_name:
        return [exact_name]

    cleaned = strip_speaker_delivery_note(text)
    exact_name = match_known_speaker_label(cleaned, possible_characters)
    if exact_name:
        return [exact_name]

    # WPS and older Word PDFs may insert visual tracking spaces inside a
    # name while keeping the ampersand or slash as a separate fragment, for
    # example ``DOLOKHOV & HÉ LÈ NE``. Individual-name matching already uses
    # ``speaker_match_key`` and therefore ignores those spaces and accents.
    # Apply that same strict, whole-part comparison to complete shared labels
    # before falling back to the character-by-character parser below.
    separated_parts = [
        part.strip()
        for part in re.split(r"[,，/&＋+、／;；]", cleaned)
    ]
    if len(separated_parts) >= 2 and all(separated_parts):
        separated_names = [
            match_known_speaker_label(part, possible_characters)
            for part in separated_parts
        ]
        if all(separated_names):
            return separated_names

    # GUIDE/NARRATOR/ (OFFSTAGE) is an existing supported form. Once the
    # delivery note is removed, the final slash is label punctuation.
    cleaned = normalise(cleaned)
    cleaned = re.sub(r"[,，/&＋+、／;；-]+$", "", cleaned).rstrip()
    candidates = sorted(
        (
            (normalise(character), character)
            for character in possible_characters
            if normalise(character)
        ),
        key=lambda item: len(item[0]),
        reverse=True,
    )
    names = []
    cursor = 0

    while cursor < len(cleaned):
        while cursor < len(cleaned) and cleaned[cursor].isspace():
            cursor += 1

        matched = next(
            (
                (candidate_text, character)
                for candidate_text, character in candidates
                if cleaned.startswith(candidate_text, cursor)
                and (
                    cursor + len(candidate_text) == len(cleaned)
                    or cleaned[cursor + len(candidate_text)].isspace()
                    or cleaned[cursor + len(candidate_text)]
                    in ",，/&＋+、／;；和与"
                )
            ),
            None,
        )
        if matched is None:
            return []

        candidate_text, character = matched
        names.append(character)
        cursor += len(candidate_text)

        while cursor < len(cleaned) and cleaned[cursor].isspace():
            cursor += 1
        if cursor >= len(cleaned):
            break

        connector_found = False
        if cleaned[cursor] in "和与":
            cursor += 1
            connector_found = True
        elif (
            cleaned.startswith("and", cursor)
            and (
                cursor + 3 == len(cleaned)
                or cleaned[cursor + 3].isspace()
            )
        ):
            cursor += 3
            connector_found = True
        elif cleaned[cursor] in ",，/&＋+、／;；":
            connector_found = True
            while (
                cursor < len(cleaned)
                and cleaned[cursor] in ",，/&＋+、／;；"
            ):
                cursor += 1
                while (
                    cursor < len(cleaned)
                    and cleaned[cursor].isspace()
                ):
                    cursor += 1

            if (
                cleaned.startswith("and", cursor)
                and (
                    cursor + 3 == len(cleaned)
                    or cleaned[cursor + 3].isspace()
                )
            ):
                cursor += 3
            elif cursor < len(cleaned) and cleaned[cursor] in "&＋+和与":
                cursor += 1

        if not connector_found:
            return []

    return names if len(names) >= 2 else []


def parse_explicit_speaker_prefix(prefix, possible_characters):
    """Resolve a complete colon-delimited speaker prefix."""
    speaker_names = split_known_speaker_group(
        prefix,
        possible_characters,
    )
    if speaker_names:
        return speaker_names

    delivery_label = re.fullmatch(r"(.*?)[,，](.+)", prefix)
    if delivery_label:
        speaker_name = match_known_speaker_label(
            delivery_label.group(1).strip(),
            possible_characters,
        )
        if (
            speaker_name
            and is_speaker_delivery_suffix(delivery_label.group(2))
        ):
            return [speaker_name]

    return []


def get_explicit_speaker_names(text, possible_characters):
    """Return a fully validated speaker prefix before a printed colon."""
    cleaned = strip_cast_track_prefix(text)
    if starts_with_stage_direction(cleaned):
        return []

    separator = re.search(r"[：:]", cleaned)
    if separator is None:
        return []

    prefix = cleaned[:separator.start()].strip()
    return parse_explicit_speaker_prefix(prefix, possible_characters)


def has_incomplete_explicit_speaker_group(text, possible_characters):
    """Detect a colon label that starts a group but contains an unknown name."""
    cleaned = strip_cast_track_prefix(text)
    separator = re.search(r"[：:]", cleaned)
    if separator is None:
        return False

    prefix = strip_speaker_delivery_note(
        cleaned[:separator.start()].strip()
    )
    if match_known_speaker_label(prefix, possible_characters):
        return False

    normalised_prefix = normalise(prefix)
    for character in sorted(possible_characters, key=len, reverse=True):
        character_text = normalise(character)
        if not normalised_prefix.startswith(character_text):
            continue
        remainder = normalised_prefix[len(character_text):].lstrip()
        return bool(
            remainder
            and (
                remainder[0] in ",，/&＋+、／;；和与"
                or re.match(r"and(?:\s|$)", remainder)
            )
        )

    return False


def get_speaker_names(text, possible_characters):
    """Return every character named in one shared dialogue cue.

    English scripts often use labels such as ``ROBIN & AVERY``. Both
    people are speaking, so both DCA assignments must appear together.
    """
    # Read names only from the start of a printed cue label.  Searching the
    # whole dialogue line causes false combinations such as
    # ``ALEX. ... (to BLAIR & CASEY ...)`` and makes CHORUS MEMBERS also match
    # the shorter template name CHORUS. A real shared label is a compact
    # prefix: ``ALEX/BLAIR/CASEY/DREW.``, ``ROBIN AND GUESTS.``, or
    # ``ALEX, BLAIR, CASEY, & DREW.``.
    original_text = unicodedata.normalize("NFKC", str(text)).strip()
    stripped_text = strip_cast_track_prefix(original_text)
    has_cast_track_prefix = stripped_text != original_text

    explicit_names = get_explicit_speaker_names(
        text,
        possible_characters,
    )
    if explicit_names:
        return explicit_names
    # A syntactically valid cast-track tag promises an explicit colon label.
    # If its complete prefix did not resolve above, do not fall back to the
    # looser name-at-start heuristics used by untagged script layouts.
    if has_cast_track_prefix:
        return []
    if has_incomplete_explicit_speaker_group(text, possible_characters):
        return []

    complete_shared_names = split_known_speaker_group(
        stripped_text,
        possible_characters,
    )
    if len(complete_shared_names) >= 2:
        return complete_shared_names

    # A complete shared label can end with a parenthetical delivery or group
    # note without using a colon, for example
    # ``亨利/麻省理工/斯坦福(员工合唱)``. Only accept this shortcut when the
    # text before the note resolves completely to at least two known names;
    # ordinary dialogue containing parentheses must remain rejected.
    without_delivery_note = strip_speaker_delivery_note(stripped_text)
    if without_delivery_note != stripped_text:
        delivery_group_names = split_known_speaker_group(
            stripped_text,
            possible_characters,
        )
        if len(delivery_group_names) >= 2:
            return delivery_group_names

    clean_text = normalise(stripped_text)
    candidates = sorted(possible_characters, key=len, reverse=True)
    combined_names = []
    cursor = 0

    while cursor < len(clean_text):
        while cursor < len(clean_text) and clean_text[cursor].isspace():
            cursor += 1

        matched_name = next(
            (
                character
                for character in candidates
                if clean_text.startswith(character, cursor)
                and (
                    cursor + len(character) == len(clean_text)
                    or not clean_text[cursor + len(character)].isalnum()
                    or clean_text[cursor + len(character)] in "和与"
                )
            ),
            None,
        )
        if not matched_name:
            # A connector means the printed group is incomplete unless the
            # following text resolves to another complete workbook name.
            # Never return only the known prefix of a longer narration or a
            # group containing an unknown member.
            if combined_names:
                remaining = clean_text[cursor:].strip()
                if (
                    len(combined_names) == 1
                    and is_speaker_delivery_suffix(remaining)
                ):
                    return combined_names
                if re.fullmatch(r"\([^()]*\)", remaining):
                    return combined_names
                return []
            break

        combined_names.append(matched_name)
        cursor += len(matched_name)

        # A dot, colon, or Chinese sentence stop ends the speaker label.
        # The punctuation may be part of abbreviations in a name (DR. Q),
        # so test longer template names first before reaching this point.
        while cursor < len(clean_text) and clean_text[cursor].isspace():
            cursor += 1
        if cursor >= len(clean_text) or clean_text[cursor] in ".:：。":
            return combined_names

        # One name can be joined to the next by punctuation, by the word
        # AND, or by a combination such as ``, & DREW``.
        if clean_text[cursor] in ",，/&＋+、/／;；":
            cursor += 1
            while cursor < len(clean_text) and clean_text[cursor].isspace():
                cursor += 1
            if clean_text.startswith("and ", cursor):
                cursor += 4
            elif cursor < len(clean_text) and clean_text[cursor] in "&＋+和与":
                cursor += 1
        elif clean_text[cursor] in "和与":
            cursor += 1
        elif (
            clean_text.startswith("and", cursor)
            and (cursor + 3 == len(clean_text)
                 or clean_text[cursor + 3].isspace())
        ):
            cursor += 3
        else:
            # Several known names at the beginning of a Chinese sentence do
            # not make that sentence a shared dialogue label. A complete
            # group must end or reach recognised label punctuation; prose
            # such as ``林青、周岚和方宁走进庭院`` is narration.
            if len(combined_names) >= 2:
                return []
            break

    if len(combined_names) >= 2:
        return combined_names

    speaker_name = get_speaker_name(text, possible_characters)
    return [speaker_name] if speaker_name else []


def leading_speaker_remainder(text, speaker_name):
    """Return the text following the shortest complete leading name."""
    cleaned = strip_cast_track_prefix(text)
    target_key = speaker_match_key(speaker_name)

    for index in range(1, len(cleaned) + 1):
        if speaker_match_key(cleaned[:index]) == target_key:
            return cleaned[index:].lstrip()

    return cleaned


def get_split_english_speaker_fragment_names(
    text,
    possible_characters,
    allow_bare=False,
):
    """Return names from an isolated title-case English cue fragment.

    Older Word PDFs can store ``Rowan.`` and its dialogue as separate
    physical lines on the same baseline.  Title case plus a full stop is too
    ambiguous to trust from text alone, so this helper only validates the
    isolated fragment.  The caller must still supply page-layout evidence
    before treating it as dialogue.  A trailing number is accepted because
    footnote references are sometimes emitted as another span in the label.
    """
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    if not cleaned or starts_with_stage_direction(cleaned):
        return []

    cleaned = re.sub(r"(?<=\.)\s*\d+\s*$", "", cleaned).strip()
    has_full_stop = cleaned.endswith(".")
    if has_full_stop:
        label = cleaned[:-1].rstrip()
    elif allow_bare:
        label = cleaned
    else:
        return []

    if not label:
        return []

    names = split_known_speaker_group(label, possible_characters)
    if not names:
        exact_name = match_known_speaker_label(
            label,
            possible_characters,
        )
        names = [exact_name] if exact_name else []

    if not names or any(contains_cjk(name) for name in names):
        return []

    return names


def get_inline_english_shared_speaker_names(
    text,
    possible_characters,
):
    """Return a complete shared label before its first valid full stop.

    A few older Word PDFs keep ``Robin & Avery. Are we ready?`` in one
    regular Times New Roman span.  A single title-case name followed by
    dialogue is too ambiguous to accept, but a complete group of two or more
    workbook names becomes safe when the caller also verifies the established
    speaker column.
    """
    cleaned = unicodedata.normalize("NFKC", str(text)).strip()
    if not cleaned or starts_with_stage_direction(cleaned):
        return []

    for full_stop in re.finditer(r"\.(?=\s|$)", cleaned):
        remainder = cleaned[full_stop.end():].strip()
        if not remainder:
            continue

        names = split_known_speaker_group(
            cleaned[:full_stop.start()].strip(),
            possible_characters,
        )
        if (
            len(names) >= 2
            and not any(contains_cjk(name) for name in names)
        ):
            return names

    return []


def looks_like_positioned_speaker_label(
    text,
    speaker_names,
    row_x,
    trusted_columns,
    tolerance=24,
    bold_prefix=False,
    layout_speaker_names=None,
):
    """Validate ambiguous labels with names, punctuation, and page layout."""
    if not speaker_names:
        return False

    layout_speaker_names = layout_speaker_names or []
    aligned_with_trusted_column = any(
        abs(float(row_x) - float(column)) <= tolerance
        for column in trusted_columns
    )
    has_split_layout_evidence = bool(
        layout_speaker_names == speaker_names
        and aligned_with_trusted_column
        and not starts_with_stage_direction(text)
    )
    inline_shared_names = get_inline_english_shared_speaker_names(
        text,
        set(speaker_names),
    )
    has_inline_shared_layout_evidence = bool(
        inline_shared_names == speaker_names
        and aligned_with_trusted_column
    )
    speaker_name = speaker_names[0]
    valid_shared_names = get_explicit_speaker_names(
        text,
        set(speaker_names),
    )
    is_valid_shared_label = (
        len(speaker_names) >= 2
        and valid_shared_names == speaker_names
    )
    if (
        not is_valid_shared_label
        and not looks_like_speaker_label(text, speaker_name)
        and not (
            bold_prefix
            and not starts_with_stage_direction(text)
        )
        and not has_split_layout_evidence
        and not has_inline_shared_layout_evidence
    ):
        return False

    if len(speaker_names) != 1:
        return True

    remainder = leading_speaker_remainder(text, speaker_name)
    if remainder.startswith((",", "，")):
        if not is_speaker_delivery_suffix(remainder[1:]):
            return False
    if remainder.startswith(("和", "与", "、")):
        return False

    # ``林青。`` can be either a genuine full-stop cue or one indented line
    # of dialogue. Colons elsewhere on the page establish the script's real
    # speaker column; use that evidence when available, and retain the old
    # behaviour for scripts that use only full-stop labels.
    if (
        contains_cjk(speaker_name)
        and remainder.startswith("。")
        and trusted_columns
        and not aligned_with_trusted_column
    ):
        return False

    return True


def is_standalone_speaker_label(text, possible_characters):
    """Return true when a PDF fragment contains only a speaker label.

    This is intentionally stricter than ``looks_like_speaker_label``. A
    fragment such as ``ROWAN, hello.`` begins with a valid name, but it is
    dialogue text rather than a second speaker column. Parenthetical role
    notes such as ``GUIDE/NARRATOR/ (OFFSTAGE)`` remain valid labels.
    """
    original_text = unicodedata.normalize("NFKC", str(text)).strip()
    cleaned = strip_cast_track_prefix(original_text)
    if cleaned == original_text and starts_with_stage_direction(cleaned):
        return False

    speaker_names = get_speaker_names(original_text, possible_characters)
    split_period_names = get_split_english_speaker_fragment_names(
        cleaned,
        possible_characters,
    )
    validated_shared_names = get_explicit_speaker_names(
        cleaned,
        set(speaker_names),
    ) if len(speaker_names) >= 2 else []
    if (
        not speaker_names
        or not all(name in possible_characters for name in speaker_names)
        or not (
            validated_shared_names == speaker_names
            or looks_like_speaker_label(cleaned, speaker_names[0])
            or split_period_names == speaker_names
        )
    ):
        return False

    remainder = normalise(cleaned)
    for speaker_name in speaker_names:
        remainder = re.sub(
            re.escape(normalise(speaker_name)),
            "",
            remainder,
            count=1,
            flags=re.IGNORECASE,
        )

    # Notes describe how or where the cue is delivered; they are not extra
    # dialogue. NFKC normalisation has already converted full-width brackets.
    remainder = re.sub(r"\([^()]*\)", "", remainder)
    remainder = re.sub(r"\[[^\[\]]*\]", "", remainder)
    remainder = re.sub(r"【[^【】]*】", "", remainder)
    remainder = re.sub(r"\b(?:and)\b|[和与]", "", remainder)

    return not re.sub(r"[\s.,，:：。;/／&＋+、;；-]", "", remainder)


def speaker_label_continues(text):
    """Return true when another PDF fragment completes a shared label."""
    cleaned = normalise(text)
    return bool(
        re.search(
            r"(?:[,，/&＋+、／;；]|\band|和|与)\s*$",
            cleaned,
            flags=re.IGNORECASE,
        )
    )


def get_leading_known_speaker_names(text, possible_characters):
    """Return a validated speaker label at the start of an inline cue.

    This is deliberately narrower than ``get_speaker_names``. It is used
    while reconstructing parallel lyric rows, before trusted page columns
    are available, so the printed fragment itself must still look like a
    complete speaker cue rather than dialogue that merely begins with a
    character's name.
    """
    if starts_with_stage_direction(text):
        return []

    speaker_names = get_speaker_names(text, possible_characters)
    if (
        not speaker_names
        or not all(
            name in possible_characters for name in speaker_names
        )
    ):
        return []

    if is_standalone_speaker_label(text, possible_characters):
        return speaker_names

    if (
        len(speaker_names) == 1
        and looks_like_speaker_label(text, speaker_names[0])
    ):
        return speaker_names

    if (
        len(speaker_names) >= 2
        and get_explicit_speaker_names(
            text,
            set(speaker_names),
        ) == speaker_names
    ):
        return speaker_names

    return []


def padded_span_visible_left_edges(page, page_text):
    """Locate visible glyphs without changing the geometry used for matching.

    Some PDFs centre a speaker with spaces inside the same text span. Its
    bounding box then starts at the margin, far before the printed name.
    Use the PDF's actual character positions, not an estimated space width.
    Exact text/bbox keys deliberately exclude any synthetic spans produced
    later when reconstructing parallel speaker columns.
    """
    padded_spans = {
        (tuple(span["bbox"]), span["text"])
        for block in page_text.get("blocks", [])
        for line in block.get("lines", [])
        if tuple(line.get("dir", (1, 0))) == (1, 0)
        for span in line.get("spans", [])
        if span.get("text", "")[:1].isspace() and span["text"].strip()
    }
    if not padded_spans:
        return {}

    visible_left_edges = {}
    for block in page.get_text("rawdict").get("blocks", []):
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                characters = span.get("chars", [])
                text = "".join(char["c"] for char in characters)
                key = (tuple(span["bbox"]), text)
                if key not in padded_spans:
                    continue
                visible = next(
                    (char for char in characters if not char["c"].isspace()),
                    None,
                )
                if visible is not None:
                    left = float(visible["bbox"][0])
                    if (
                        math.isfinite(left)
                        and span["bbox"][0] < left < span["bbox"][2]
                    ):
                        visible_left_edges[key] = left
    return visible_left_edges


def split_embedded_right_speaker_lines(
    page,
    page_text,
    possible_characters,
):
    """Expose a tabbed right-column speaker as its own synthetic line.

    Word-generated PDFs can store a visual row like this in one span::

        left lyric            RIGHT SPEAKER      right lyric

    The word boxes still retain the real x-position of ``RIGHT SPEAKER``.
    Split only when another known speaker anchors the left side of the same
    baseline and dialogue continues to the right. Those guards keep names in
    ordinary prose and stage directions out of the speaker pipeline.
    """
    words_by_line = {}
    for word in page.get_text("words"):
        if len(word) < 8:
            continue
        words_by_line.setdefault(
            (int(word[5]), int(word[6])),
            [],
        ).append(word)

    original_lines = [
        line
        for block in page_text.get("blocks", [])
        if "lines" in block
        for line in block["lines"]
        if any(
            normalise(span.get("text", ""))
            for span in line.get("spans", [])
        )
    ]
    leading_speakers_by_line = {
        id(line): get_leading_known_speaker_names(
            "".join(
                span.get("text", "")
                for span in line.get("spans", [])
            ),
            possible_characters,
        )
        for line in original_lines
    }

    for block_index, block in enumerate(page_text.get("blocks", [])):
        if "lines" not in block:
            continue

        expanded_lines = []
        for line_index, line in enumerate(block["lines"]):
            line_words = sorted(
                words_by_line.get((block_index, line_index), []),
                key=lambda word: int(word[7]),
            )
            nonempty_spans = [
                span
                for span in line.get("spans", [])
                if normalise(span.get("text", ""))
            ]
            raw_text = "".join(
                span.get("text", "")
                for span in line.get("spans", [])
            )
            split_result = None

            if (
                len(line_words) >= 2
                and len(nonempty_spans) == 1
                and not starts_with_stage_direction(raw_text)
            ):
                for word_index in range(1, len(line_words)):
                    word = line_words[word_index]
                    previous_word = line_words[word_index - 1]
                    embedded_name = match_known_speaker_label(
                        word[4],
                        possible_characters,
                    )
                    if not embedded_name:
                        continue

                    internal_gap = float(word[0]) - float(previous_word[2])
                    if (
                        internal_gap < 60
                        or float(word[0]) < page.rect.width * 0.45
                    ):
                        continue

                    left_peer_names = next(
                        (
                            leading_speakers_by_line[id(peer)]
                            for peer in original_lines
                            if peer is not line
                            and abs(
                                float(peer["bbox"][1])
                                - float(line["bbox"][1])
                            ) < 0.75
                            and float(peer["bbox"][0])
                            <= page.rect.width * 0.30
                            and float(word[0])
                            - float(peer["bbox"][0])
                            >= page.rect.width * 0.25
                            and leading_speakers_by_line[id(peer)]
                            and embedded_name
                            not in leading_speakers_by_line[id(peer)]
                        ),
                        [],
                    )
                    if not left_peer_names:
                        continue

                    has_following_word = word_index + 1 < len(line_words)
                    has_right_fragment = any(
                        peer is not line
                        and abs(
                            float(peer["bbox"][1])
                            - float(line["bbox"][1])
                        ) < 0.75
                        and float(peer["bbox"][0])
                        >= float(word[2]) + 8
                        and any(
                            normalise(span.get("text", ""))
                            for span in peer.get("spans", [])
                        )
                        for peer in original_lines
                    )
                    if not (has_following_word or has_right_fragment):
                        continue

                    name_match = next(
                        (
                            match
                            for match in re.finditer(
                                re.escape(str(word[4])),
                                raw_text,
                            )
                            if match.start() > 0
                            and re.search(
                                r"\s{3,}$",
                                raw_text[:match.start()],
                            )
                        ),
                        None,
                    )
                    if name_match is None:
                        continue

                    prefix_text = raw_text[:name_match.start()].rstrip()
                    speaker_text = raw_text[name_match.start():].lstrip()
                    if not normalise(prefix_text):
                        continue

                    source_span = nonempty_spans[0]
                    prefix_bbox = (
                        float(line["bbox"][0]),
                        float(line["bbox"][1]),
                        float(previous_word[2]),
                        float(line["bbox"][3]),
                    )
                    speaker_has_suffix = bool(
                        normalise(raw_text[name_match.end():])
                    )
                    speaker_bbox = (
                        float(word[0]),
                        float(line["bbox"][1]),
                        (
                            float(line["bbox"][2])
                            if speaker_has_suffix
                            else float(word[2])
                        ),
                        float(line["bbox"][3]),
                    )

                    prefix_span = dict(source_span)
                    prefix_span["text"] = prefix_text
                    prefix_span["bbox"] = prefix_bbox
                    prefix_line = dict(line)
                    prefix_line["bbox"] = prefix_bbox
                    prefix_line["spans"] = [prefix_span]

                    speaker_span = dict(source_span)
                    speaker_span["text"] = speaker_text
                    speaker_span["bbox"] = speaker_bbox
                    if "origin" in speaker_span:
                        speaker_span["origin"] = (
                            speaker_bbox[0],
                            speaker_span["origin"][1],
                        )
                    speaker_line = dict(line)
                    speaker_line["bbox"] = speaker_bbox
                    speaker_line["spans"] = [speaker_span]
                    split_result = (prefix_line, speaker_line)
                    break

            if split_result is None:
                expanded_lines.append(line)
            else:
                expanded_lines.extend(split_result)

        block["lines"] = expanded_lines


def find_split_english_speaker_layout(
    physical_lines,
    possible_characters,
    inherited_columns,
    page_width,
    page_is_cast_reference=False,
    tolerance=24,
):
    """Infer safe title-case speaker fragments from repeated PDF geometry.

    Some older Word PDFs place a plain ``Rowan.`` fragment in one column and
    the dialogue in a separate fragment on the same baseline.  A title-case
    full stop is not safe text evidence by itself.  It becomes strong evidence
    when the isolated, non-italic fragment has dialogue to its right and its
    x-position repeats or continues an established speaker column.  A single
    far-right column is also accepted on a page that already establishes the
    main speaker gutter, which preserves parallel duet layouts.

    Bare fragments such as ``Casey`` may use an already trusted column, but can
    never establish one on their own.
    """
    if page_is_cast_reference:
        return [], {}

    candidates = []
    for line in physical_lines:
        nonempty_spans = [
            span
            for span in line.get("spans", [])
            if normalise(span.get("text", ""))
        ]
        if not nonempty_spans or is_italic(nonempty_spans[0]):
            continue

        # Use the first physical span as the label fragment. Some PDFs place
        # ``Avery.`` in one span and ``(aside) Dialogue`` in later spans of the
        # same line rather than exposing two separate PDF lines.
        line_text = str(nonempty_spans[0].get("text", ""))
        period_names = get_split_english_speaker_fragment_names(
            line_text,
            possible_characters,
        )
        fragment_names = period_names or (
            get_split_english_speaker_fragment_names(
                line_text,
                possible_characters,
                allow_bare=True,
            )
        )
        if not fragment_names:
            continue

        has_inline_dialogue = any(
            any(character.isalpha() for character in str(span.get("text", "")))
            for span in nonempty_spans[1:]
        )
        has_right_dialogue = has_inline_dialogue or any(
            candidate is not line
            and abs(candidate["bbox"][1] - line["bbox"][1]) < 0.75
            and candidate["bbox"][0] > line["bbox"][2] + 4
            and any(
                character.isalpha()
                for character in "".join(
                    str(span.get("text", ""))
                    for span in candidate.get("spans", [])
                )
            )
            for candidate in physical_lines
        )
        # A period-ended label may continue on the following baseline. Keep
        # it as a candidate so an already established column can validate it.
        # A bare title-case name remains too ambiguous without adjacent text.
        if not has_right_dialogue and not period_names:
            continue

        candidates.append({
            "line_id": id(line),
            "names": fragment_names,
            "x": float(nonempty_spans[0]["bbox"][0]),
            "has_period": bool(period_names),
            "has_right_dialogue": has_right_dialogue,
        })

    period_candidates = [
        candidate for candidate in candidates
        if candidate["has_period"] and candidate["has_right_dialogue"]
    ]
    reliable_columns = []

    for candidate in period_candidates:
        same_column_count = sum(
            abs(candidate["x"] - other["x"]) <= 8
            for other in period_candidates
        )
        continues_inherited_column = any(
            abs(candidate["x"] - float(column)) <= tolerance
            for column in inherited_columns
        )
        if same_column_count >= 2 or continues_inherited_column:
            reliable_columns.append(candidate["x"])

    # Parallel song layouts can introduce one genuine speaker cue in a new
    # right-hand column.  Require the page to establish another reliable
    # column first, then constrain this exception to the right half.
    if reliable_columns:
        for candidate in period_candidates:
            if candidate["x"] >= float(page_width) * 0.45:
                reliable_columns.append(candidate["x"])

    trusted_for_page = list(inherited_columns) + reliable_columns
    layout_names_by_line = {}
    current_columns = []
    for candidate in candidates:
        if any(
            abs(candidate["x"] - float(column)) <= tolerance
            for column in trusted_for_page
        ):
            layout_names_by_line[candidate["line_id"]] = candidate["names"]
            current_columns.append(candidate["x"])

    return list(dict.fromkeys(current_columns)), layout_names_by_line


def fragment_extends_speaker_name(current_text, fragment_text, characters):
    """Return true when a fragment completes an unfinished template name."""
    current_key = speaker_match_key(current_text)
    if not current_key:
        return False

    character_keys = {
        speaker_match_key(character)
        for character in characters
        if speaker_match_key(character)
    }
    # Respect a real gap after a complete short name even when another name
    # in the workbook happens to begin with the same characters.
    if current_key in character_keys:
        return False

    combined_key = speaker_match_key(current_text + fragment_text)
    return any(
        character_key.startswith(current_key)
        and combined_key.startswith(character_key)
        for character_key in character_keys
    )


def join_visual_text_parts(parts, possible_characters):
    """Join positioned text parts while retaining visible word boundaries."""
    joined_parts = []
    previous_right = None

    for part in parts:
        part_text = part["text"]
        current_text = "".join(joined_parts)
        if (
            joined_parts
            and previous_right is not None
            and part["bbox"][0] - previous_right
            > VISUAL_FRAGMENT_WORD_GAP
            and not current_text[-1:].isspace()
            and not part_text[:1].isspace()
            and not fragment_extends_speaker_name(
                current_text,
                part_text,
                possible_characters,
            )
        ):
            joined_parts.append(" ")
        joined_parts.append(part_text)
        previous_right = part["bbox"][2]

    return "".join(joined_parts)


def join_visual_line_fragments(fragments, possible_characters):
    """Rebuild one printed row without losing speaker-name boundaries.

    WPS-generated PDFs may expose a bold name, a cast-count suffix, and its
    dialogue as separate physical lines or separate styled spans on the same
    baseline. Very small gaps belong inside a label (for example
    ``巡逻员`` + ``2``); a wider gap is a real word boundary. If the PDF
    splits a name such as ``林`` + ``川`` more widely, the workbook confirms
    that those parts still belong together.
    """
    fragment_parts = [
        {
            "text": join_visual_text_parts(
                fragment["spans"],
                possible_characters,
            ),
            "bbox": fragment["bbox"],
        }
        for fragment in fragments
    ]
    return join_visual_text_parts(
        fragment_parts,
        possible_characters,
    )


def read_sheet_rows(worksheet):
    header_row = None
    headers = []

    for cells in worksheet.iter_rows():
        possible_headers = [
            normalise(cell.value) if cell.value is not None else ""
            for cell in cells
        ]

        if "dca state" in possible_headers:
            header_row = cells[0].row
            headers = possible_headers
            break

    if header_row is None:
        return []

    rows = []

    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):
        row_data = {}

        for index, value in enumerate(row):
            if index < len(headers) and headers[index]:
                row_data[headers[index]] = value

        if any(value is not None for value in row):
            rows.append(row_data)

    return rows


def split_aliases(value):
    if value is None:
        return []

    return [
        normalise(alias)
        for alias in re.split(r"[,，;；|]", str(value))
        if alias.strip()
    ]


def split_role_names(value):
    """Read globally mapped script roles from Character List column B."""
    if value is None:
        return []

    return [
        normalise(role)
        for role in re.split(r"[,，、;；|\r\n]+", str(value))
        if role.strip()
    ]


def split_display_role_names(value):
    """Read mapped roles while preserving the workbook's visible spelling."""
    if value is None:
        return []

    roles = []
    for role in re.split(r"[,，、;；|\r\n]+", str(value)):
        display_role = unicodedata.normalize("NFKC", role).strip()
        if display_role and display_role not in roles:
            roles.append(display_role)

    return roles


def convert_legacy_assignments(project):
    """One-way conversion of retired assignment sets; never infer new sets."""
    definitions = project.pop("shared_groups", None) or []
    if not definitions:
        return project

    def names(value):
        # Accept old delimiters without splitting commas inside [aliases].
        parts, current, depth = [], [], 0
        for character in str(value or ""):
            if character == "[":
                depth += 1
            elif character == "]":
                depth = max(0, depth - 1)
            if depth == 0 and character in ",，、;；|\r\n":
                if "".join(current).strip():
                    parts.append("".join(current).strip())
                current = []
            else:
                current.append(character)
        if "".join(current).strip():
            parts.append("".join(current).strip())
        return parts

    def identity(value):
        entries = split_character_entries(value)
        return entries[0][0] if entries else ""

    by_name = {}
    for item in definitions:
        key = identity(item.get("name", ""))
        if key:
            by_name.setdefault(key, []).append(item)

    characters = project.setdefault("characters", [])
    for item in characters:
        item["other_characters"] = "\n".join(
            role for role in names(item.get("other_characters", ""))
            if identity(role) not in by_name
        )

    known = {identity(item.get("dca_name", "")) for item in characters}
    for item in definitions:
        for name in [str(item.get("name", "")).strip(), *names(item.get("members", ""))]:
            key = identity(name)
            if key and key not in known:
                known.add(key)
                characters.append({
                    "id": f"converted-character-{len(characters) + 1}",
                    "dca_name": name,
                    "other_characters": "",
                })

    for state in project.setdefault("states", []):
        converted_cells = []
        for cell in state.get("dca_assignments", []):
            entries = []
            seen = set()
            for line in str(cell or "").splitlines():
                line = line.strip()
                if not line:
                    continue
                if normalise(line) not in seen:
                    entries.append(line)
                    seen.add(normalise(line))
                for item in by_name.get(identity(line), []):
                    for member in names(item.get("members", "")):
                        if normalise(member) not in seen:
                            entries.append(member)
                            seen.add(normalise(member))
            converted_cells.append("\n".join(entries))
        state["dca_assignments"] = converted_cells
    return project


def read_legacy_assignment_rows(workbook):
    """Read retired worksheet data only for one-way import conversion."""
    if "Shared Groups" not in workbook.sheetnames:
        return None

    worksheet = workbook["Shared Groups"]
    header_row, columns = _find_header_row_and_columns(
        worksheet,
        {
            "shared group name",
            "shared group",
            "group name",
        },
    )
    if header_row is None:
        return []

    name_column = next(
        (
            columns[header]
            for header in (
                "shared group name",
                "shared group",
                "group name",
            )
            if header in columns
        ),
        None,
    )
    member_column = next(
        (
            columns[header]
            for header in (
                "dca members — one per line",
                "dca members one per line",
                "dca members",
                "dca names",
                "members",
            )
            if header in columns
        ),
        None,
    )
    if name_column is None or member_column is None:
        return []

    rows = []
    for index, row in enumerate(
        worksheet.iter_rows(
            min_row=header_row + 1,
            values_only=True,
        ),
        1,
    ):
        name = (
            unicodedata.normalize("NFKC", str(row[name_column])).strip()
            if name_column < len(row) and row[name_column] is not None
            else ""
        )
        members = (
            unicodedata.normalize("NFKC", str(row[member_column])).strip()
            if member_column < len(row) and row[member_column] is not None
            else ""
        )
        if name or members:
            rows.append({
                "id": f"legacy-assignment-{index}",
                "name": name,
                "members": members,
            })
    return rows


def split_character_entries(value, *, preserve_display_names=False):
    """Read DCA names and normalized state-local aliases.

    Matching callers keep normalized names by default. The inspector-loading
    path preserves the entered name until its separate display value has been
    captured; add_template_assignment still normalizes the matching identity.
    """
    if value is None:
        return []

    entries = []

    for line in str(value).splitlines():
        line = line.strip()

        if not line:
            continue

        alias_match = re.match(r"^(.*?)\s*\[([^\]]+)\]\s*$", line)

        if alias_match:
            character = alias_match.group(1).strip()
            aliases = split_aliases(alias_match.group(2))
        else:
            character = line
            aliases = []

        if character:
            entries.append((
                character if preserve_display_names else normalise(character),
                aliases,
            ))

    return entries


def split_character_cell(value):
    """Read one horizontal DCA cell, including names and [aliases]."""
    names = []

    for character, aliases in split_character_entries(value):
        names.append(character)
        names.extend(aliases)

    return names


def load_role_mappings(workbook):
    """Load optional performer-to-script-role mappings.

    Existing aliases remain local to an individual DCA cell. This separate
    mapping lets a stable DCA member or performer name such as ``Ben`` match
    several real script roles such as ``Barber``, ``Butcher``, and ``Coach``
    without repeating those role names in every state. Every mapped role has
    exactly one DCA identity; names never create implicit membership lists.
    """
    if "Character List" not in workbook.sheetnames:
        return {}, {}, {}

    worksheet = workbook["Character List"]
    canonical_headers = {
        "dca name",
        "dca name / performer",
        "dca member / performer",
        "performer / dca name",
    }
    role_headers = {
        "other script characters played",
        "script characters played",
        "script character names",
        "script roles",
    }
    header_row = None
    canonical_column = None
    role_column = None

    for cells in worksheet.iter_rows(
        # XLSX dimensions are optional; read-only sheets may have no max_row.
        max_row=min(worksheet.max_row or 20, 20)
    ):
        headers = [
            normalise(cell.value) if cell.value is not None else ""
            for cell in cells
        ]
        canonical_column = next(
            (
                index
                for index, header in enumerate(headers)
                if header in canonical_headers
            ),
            None,
        )
        role_column = next(
            (
                index
                for index, header in enumerate(headers)
                if header in role_headers
            ),
            None,
        )

        if canonical_column is not None and role_column is not None:
            header_row = cells[0].row
            break

    # Older workbooks have no role-mapping headers. Preserve their exact
    # behavior and keep Character List as a dropdown-only helper sheet.
    if header_row is None:
        return {}, {}, {}

    role_groups = {}
    role_display_groups = {}

    for row in worksheet.iter_rows(
        min_row=header_row + 1,
        values_only=True,
    ):
        canonical_value = (
            row[canonical_column]
            if canonical_column < len(row)
            else None
        )
        role_value = row[role_column] if role_column < len(row) else None
        canonical_display = unicodedata.normalize(
            "NFKC", str(canonical_value or "")
        ).strip()
        # A DCA Name may already include an inline alias. Use its primary
        # name as the identity so Jack [John, Student] can resolve the same
        # global role mapping as Jack [John]. Keep the full label for display.
        canonical_match = re.fullmatch(r"(.+?)\s*\[([^\]]+)\]", canonical_display)
        canonical = normalise(
            canonical_match.group(1) if canonical_match else canonical_display
        )
        name_aliases = (
            split_display_role_names(canonical_match.group(2))
            if canonical_match else []
        )

        if not canonical:
            continue

        roles = role_groups.setdefault(canonical, [])
        display_group = role_display_groups.setdefault(
            canonical,
            {
                "performer": canonical_display,
                "roles": [],
            },
        )
        display_roles_by_key = {
            normalise(role): role
            for role in display_group["roles"]
        }

        for display_role in [*name_aliases, *split_display_role_names(role_value)]:
            role = normalise(display_role)
            if role != canonical and role not in roles:
                roles.append(role)
            if (
                role != canonical
                and role not in display_roles_by_key
            ):
                display_group["roles"].append(display_role)
                display_roles_by_key[role] = display_role

    owners_by_speaker_key = {}
    canonical_by_speaker_key = {}

    # Register every DCA member first. A role may not silently reuse another
    # member's name because that remains genuinely ambiguous.
    for canonical in role_groups:
        key = speaker_match_key(canonical)
        existing_owner = canonical_by_speaker_key.get(key)
        if existing_owner and existing_owner != canonical:
            raise ValueError(
                "Character List contains ambiguous DCA names: "
                f'"{canonical}" conflicts with "{existing_owner}".'
            )
        canonical_by_speaker_key[key] = canonical
        owners_by_speaker_key[key] = [canonical]

    for canonical, roles in role_groups.items():
        for role in roles:
            key = speaker_match_key(role)
            canonical_owner = canonical_by_speaker_key.get(key)
            if canonical_owner and canonical_owner != canonical:
                raise ValueError(
                    "Character List role mapping conflicts with a DCA Name: "
                    f'role "{role}" under "{canonical}" is already the '
                    f'DCA Name "{canonical_owner}".'
                )
            owners = owners_by_speaker_key.setdefault(key, [])
            if owners and canonical not in owners:
                first_owner = role_display_groups[owners[0]]["performer"]
                current_owner = role_display_groups[canonical]["performer"]
                raise ValueError(
                    f'Role "{role}" is assigned to both "{first_owner}" '
                    f'and "{current_owner}". Give each mapped role one DCA '
                    "Name, or enter the printed label as its own DCA Name."
                )
            if canonical not in owners:
                owners.append(canonical)

    return role_groups, owners_by_speaker_key, role_display_groups


def role_equivalent_names(name, role_groups, owners_by_speaker_key):
    """Return a DCA identity and every globally mapped script role."""
    normalised_name = normalise(name)
    if not normalised_name:
        return []

    owners = owners_by_speaker_key.get(
        speaker_match_key(normalised_name),
        [],
    )
    owner = owners[0] if owners else normalised_name
    names = [owner]
    names.extend(role_groups.get(owner, []))

    return list(dict.fromkeys(names))


def add_dca_reference_assignment(
    reference_assignments,
    state_key,
    performer_key,
    performer_display,
    dca,
):
    """Keep one complete, display-ready DCA row for the inspector."""
    state_rows = reference_assignments.setdefault(state_key, {})
    row = state_rows.setdefault(
        performer_key,
        {
            "dca": [],
            "performer": performer_display,
        },
    )
    if dca not in row["dca"]:
        row["dca"].append(dca)


def add_template_assignment(
    assignments,
    legend_assignments,
    reference_assignments,
    state_key,
    character,
    dca,
    role_groups,
    owners_by_speaker_key,
    inline_aliases=None,
    state_display_name=None,
    performer_display_names=None,
):
    """Add one DCA identity and its explicitly mapped script names."""
    character_display = unicodedata.normalize(
        "NFKC", str(character or "")
    ).strip()
    character = normalise(character_display)
    if not character:
        return

    character_owners = owners_by_speaker_key.get(
        speaker_match_key(character),
        [],
    )
    if not character_owners or character in character_owners:
        character_owner = character
    elif len(character_owners) == 1:
        # Older or manually edited workbooks may place a uniquely mapped
        # script role in a DCA State cell. Resolving it to its one DCA Name is
        # deterministic and keeps legends / mapping cards canonical.
        character_owner = character_owners[0]
    else:
        raise ValueError(
            f'The script role "{character_display}" has more than one DCA Name.'
        )

    add_assignment(legend_assignments, state_key, character_owner, dca)
    performer_display_names = performer_display_names or {}
    add_dca_reference_assignment(
        reference_assignments,
        state_key,
        character_owner,
        performer_display_names.get(
            character_owner,
            character_display,
        ),
        dca,
    )
    matching_names = role_equivalent_names(
        character_owner,
        role_groups,
        owners_by_speaker_key,
    )

    for alias in inline_aliases or []:
        alias = normalise(alias)
        alias_owners = owners_by_speaker_key.get(
            speaker_match_key(alias),
            [],
        )
        if alias_owners and character_owner not in alias_owners:
            owner_text = ", ".join(alias_owners)
            raise ValueError(
                "A DCA cell alias conflicts with Character List role "
                f'mapping: "{alias}" belongs to "{owner_text}", not '
                f'"{character}".'
            )
        if alias and alias not in matching_names:
            matching_names.append(alias)
        if alias:
            # Inline aliases are explicitly state-local and historically
            # appeared in legends. Keep that behavior while hiding only the
            # new global performer/role expansion.
            add_assignment(legend_assignments, state_key, alias, dca)

    for matching_name in matching_names:
        add_assignment(assignments, state_key, matching_name, dca)


def add_assignment(assignments, state_key, character, dca):
    """Keep every DCA when the same cue intentionally uses more than one."""
    state_assignments = assignments.setdefault(state_key, {})
    existing = state_assignments.get(character, [])

    if not isinstance(existing, list):
        existing = [existing]

    if dca not in existing:
        existing.append(dca)

    state_assignments[character] = existing


def display_dca(dca):
    if not isinstance(dca, list):
        return str(dca)

    # A shared cue should read naturally as 3/4 rather than following the
    # order in which the two speaker names happen to appear on the PDF page.
    ordered = sorted(
        dict.fromkeys(str(value) for value in dca),
        key=lambda value: (
            0,
            int(value),
        ) if value.isdigit() else (1, value),
    )
    return "/".join(ordered)


def load_template(filename, diagnostics=None):
    workbook = load_workbook(filename, data_only=True, read_only=True)

    try:
        if read_legacy_assignment_rows(workbook):
            converted = project_to_workbook(import_excel_project(filename), include_guide=False)
            try:
                return _load_template_workbook(converted, diagnostics)
            finally:
                converted.close()
        return _load_template_workbook(workbook, diagnostics)
    finally:
        workbook.close()


PROJECT_SCHEMA_VERSION = 1
PROJECT_SETTINGS_DEFAULTS = {
    "marking_style": "Editable Full Marking",
    "mark_selected_pages": False,
    "start_page": "",
    "end_page": "",
    "number_colour": "Red",
    "number_size": "Medium",
    "number_font": "Helvetica",
    "number_position": "Standard",
    "number_vertical_position": "Default",
    "state_colour": "Blue",
    "state_size": "Medium",
    "state_font": "PingFang SC",
    "state_position": "Left Gutter",
    "legend_position": "Left Gutter",
    "page_state_display": "Header and Footer",
    "page_state_text_colour": "Blue",
    "page_state_text_size": "Medium",
    "page_state_text_font": "PingFang SC",
    "page_state_border_colour": "Blue",
    "show_performer_role_mapping": False,
}


def blank_project(name="Untitled DCA Project"):
    return {
        "schema_version": PROJECT_SCHEMA_VERSION,
        "name": name,
        "script_path": "",
        "output_folder": "",
        "source_excel_path": "",
        "settings": dict(PROJECT_SETTINGS_DEFAULTS),
        "characters": [],
        "states": [],
    }


def read_project_file(filename):
    with open(filename, encoding="utf-8") as file:
        project = json.load(file)

    if not isinstance(project, dict):
        raise ValueError("The DCA Script Marker project file is not valid.")
    schema_version = project.get("schema_version", PROJECT_SCHEMA_VERSION)
    if schema_version != PROJECT_SCHEMA_VERSION:
        raise ValueError(
            "This DCA Script Marker project uses an unsupported format "
            f"(version {schema_version})."
        )

    settings = dict(PROJECT_SETTINGS_DEFAULTS)
    raw_settings = project.get("settings", {})
    if isinstance(raw_settings, dict):
        settings.update(raw_settings)
    project["settings"] = settings
    project.setdefault("characters", [])
    project.setdefault("states", [])
    project.setdefault("name", "Untitled DCA Project")
    project.setdefault("script_path", "")
    project.setdefault("output_folder", "")
    project.setdefault("source_excel_path", "")
    return convert_legacy_assignments(project)


def _find_header_row_and_columns(worksheet, expected_headers, max_rows=20):
    # Search the normal bounded header area even when dimensions are omitted.
    for cells in worksheet.iter_rows(
        max_row=min(worksheet.max_row or max_rows, max_rows)
    ):
        headers = [
            normalise(cell.value) if cell.value is not None else ""
            for cell in cells
        ]
        columns = {
            header: index
            for index, header in enumerate(headers)
            if header
        }
        if any(header in columns for header in expected_headers):
            return cells[0].row, columns
    return None, {}


def import_excel_project(filename):
    """Convert a compatible workbook into the Version 2 project format."""
    workbook = load_workbook(filename, data_only=False, read_only=True)
    try:
        if "DCA States" not in workbook.sheetnames:
            raise ValueError('The Excel file needs a sheet named "DCA States".')

        project_name = os.path.splitext(os.path.basename(filename))[0]
        project = blank_project(project_name)
        project["source_excel_path"] = os.path.abspath(filename)

        if "Character List" in workbook.sheetnames:
            character_sheet = workbook["Character List"]
            header_row, columns = _find_header_row_and_columns(
                character_sheet,
                {
                    "dca name",
                    "dca name / performer",
                    "dca member / performer",
                    "performer / dca name",
                },
            )
            name_column = next(
                (
                    columns[header]
                    for header in (
                        "dca name",
                        "dca name / performer",
                        "dca member / performer",
                        "performer / dca name",
                    )
                    if header in columns
                ),
                None,
            )
            role_column = next(
                (
                    columns[header]
                    for header in (
                        "other script characters played",
                        "script characters played",
                        "script character names",
                        "script roles",
                    )
                    if header in columns
                ),
                None,
            )
            if header_row is not None and name_column is not None:
                for index, row in enumerate(
                    character_sheet.iter_rows(
                        min_row=header_row + 1,
                        values_only=True,
                    ),
                    1,
                ):
                    name = (
                        str(row[name_column]).strip()
                        if name_column < len(row) and row[name_column] is not None
                        else ""
                    )
                    roles = (
                        str(row[role_column]).strip()
                        if role_column is not None
                        and role_column < len(row)
                        and row[role_column] is not None
                        else ""
                    )
                    if name:
                        project["characters"].append({
                            "id": f"character-{index}",
                            "dca_name": name,
                            "other_characters": roles,
                        })

        legacy_assignments = read_legacy_assignment_rows(workbook)
        if legacy_assignments:
            project["shared_groups"] = legacy_assignments

        states_sheet = workbook["DCA States"]
        header_row, columns = _find_header_row_and_columns(
            states_sheet,
            {"dca state"},
            max_rows=50,
        )
        if header_row is None or "dca state" not in columns:
            raise ValueError(
                'The "DCA States" sheet needs a "DCA State" header.'
            )

        def row_value(row, *headers):
            for header in headers:
                column = columns.get(header)
                if column is not None and column < len(row):
                    value = row[column]
                    if value is not None:
                        return str(value).strip()
            return ""

        for index, row in enumerate(
            states_sheet.iter_rows(
                min_row=header_row + 1,
                values_only=True,
            ),
            1,
        ):
            state_name = row_value(row, "dca state")
            cue_character = row_value(
                row,
                "start line character",
                "start cue character",
                "start cue speaker",
            )
            cue_text = row_value(row, "start line text", "start cue text")
            start_position = row_value(
                row,
                "state start position",
                "start position",
            ) or "After"
            page_hint = row_value(row, "script page hint", "page hint")
            notes = row_value(row, "notes")
            assignments = []
            for dca in range(1, 13):
                assignments.append(row_value(row, f"dca {dca}"))

            if not any(
                [state_name, cue_character, cue_text, page_hint, notes]
                + assignments
            ):
                continue

            project["states"].append({
                "id": f"state-{index}",
                "name": state_name,
                "start_line_character": cue_character,
                "start_line_text": cue_text,
                "start_position": start_position.title(),
                "page_hint": page_hint,
                "notes": notes,
                "dca_assignments": assignments,
            })

        return convert_legacy_assignments(project)
    finally:
        workbook.close()


def excel_role_choice(dca_name, role):
    """Label a macro-free role shortcut without nesting inline alias brackets."""
    owner = unicodedata.normalize("NFKC", str(dca_name or "")).strip()
    role = unicodedata.normalize("NFKC", str(role or "")).strip()
    if not owner or not role:
        return ""
    owner_match = re.fullmatch(r"(.+?)\s*\[([^\]]+)\]", owner)
    primary = owner_match.group(1).strip() if owner_match else owner
    if normalise(primary) == normalise(role):
        return ""
    aliases = split_display_role_names(owner_match.group(2)) if owner_match else []
    role_match = re.fullmatch(r"(.+?)\s*\[([^\]]+)\]", role)
    additions = (
        [role_match.group(1).strip(), *split_display_role_names(role_match.group(2))]
        if role_match else [role]
    )
    keys = {normalise(alias) for alias in aliases}
    for alias in additions:
        if normalise(alias) not in keys:
            aliases.append(alias)
            keys.add(normalise(alias))
    return f"{primary} [{', '.join(aliases)}]"


def project_to_workbook(project, include_guide=True):
    """Create a clean, compatible workbook from a Version 2 project."""
    project = convert_legacy_assignments(copy.deepcopy(project))
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Keep project exports visually consistent with the bundled Excel
    # template.  These colours, fonts, borders, and row heights mirror the
    # current template rather than the darker legacy export palette.
    font_name = "Carlito"
    dark_blue = "FF2F5E86"
    mapping_dark_blue = "FF2F648B"
    medium_blue = "FF8FAFC8"
    mapping_medium_blue = "FF93B4CC"
    pale_blue = "FFF4F8FB"
    pale_yellow = "FFFFFDF5"
    info_blue = "FFE8F1F8"
    dark_text = "FF102A43"
    mapping_dark_text = "FF16324A"
    info_text = "FF29475F"
    white = "FFFFFFFF"
    header_side = Side(style="thin", color="FFC4D4E2")
    mapping_header_side = Side(style="thin", color="FFD2E0EA")
    body_side = Side(style="thin", color="FFD9E2EC")
    header_border = Border(
        left=header_side,
        right=header_side,
        top=header_side,
        bottom=header_side,
    )
    mapping_header_border = Border(
        left=mapping_header_side,
        right=mapping_header_side,
        top=mapping_header_side,
        bottom=mapping_header_side,
    )
    body_border = Border(
        left=body_side,
        right=body_side,
        top=body_side,
        bottom=body_side,
    )
    workbook._named_styles["Normal"].font = Font(
        name=font_name,
        size=11,
    )

    if include_guide:
        guide = workbook.create_sheet("How to use")
        guide.append(["How to use DCA Script Marker Version 2"])
        guide.append([
            "Edit this workbook in Excel or import it into DCA Script Marker. "
            "Character List is optional. A label such as MALE ENSEMBLE can simply "
            "be its own DCA Name; no membership list is needed. Use Other Script "
            "Characters Played only when one person plays differently named roles. "
            "In DCA States, enter the exact start line, Page Hint when needed, and "
            "active names under DCA 1–12. Put multiple names on separate lines. "
            "Special DCA-cell example: put TOM and ALL THREE in DCA 1, JERRY and "
            "ALL THREE in DCA 2, and APPLE and ALL THREE in DCA 3. The printed "
            "ALL THREE cue receives 1/2/3; confirm or ignore the intentional "
            "duplicate-assignment reminder."
        ])
        guide.append([
            "可在 Excel 中编辑此工作簿，或导入 DCA Script Marker。Character List 为可选项。"
            "MALE ENSEMBLE 等标签可直接作为独立的 DCA Name，无需填写成员名单。"
            "只有同一人饰演名称不同的角色时，才使用 Other Script Characters Played。"
            "DCA States 请填写准确的开始台词、必要的 Page Hint，并在 DCA 1–12 下填写启用的名称；"
            "多个名称请分行填写。特别示例：DCA 1 填写 TOM 和 ALL THREE，DCA 2 填写 JERRY "
            "和 ALL THREE，DCA 3 填写 APPLE 和 ALL THREE。剧本中 ALL THREE 的提示会获得 "
            "1/2/3；请确认或忽略这项有意设置的重复分配提醒。"
        ])
        guide.append([
            "Find a DCA Name by its script role: enter Jack as the DCA Name and "
            "Student and Teacher on separate lines under Other Script Characters "
            "Played. In Excel, choose Jack [Student] or Jack [Teacher]; the cell "
            "keeps the selected label, without macros. Either choice assigns Jack "
            "and all his mapped roles to this DCA. Keep the Character List mapping. "
            "This exported workbook includes the choices known at export; export "
            "again to refresh them, or type a name manually. The blank template "
            "updates its choices when Excel recalculates."
        ])
        guide.append([
            "按剧本角色选择 DCA Name：在 Character List 中，将 Jack 填入 DCA Name，"
            "Student 和 Teacher 分两行填入 Other Script Characters Played。Excel 中选择 "
            "Jack [Student] 或 Jack [Teacher]，单元格保留所选文字，无需宏。选择任一角色，"
            "均将 Jack 及其全部对应角色分配至此 DCA。请保留 Character List 中的对应关系。"
            "此导出工作簿包含导出时已有的选项；请重新导出以更新选项，也可以手动填写名称。"
            "空白模板的选项会在 Excel 重新计算后更新。"
        ])
        guide.column_dimensions["A"].width = 115
        guide.sheet_view.showGridLines = False
        for row in guide.iter_rows():
            for cell in row:
                cell.font = Font(name=font_name, size=11)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = body_border
        guide["A1"].font = Font(
            name=font_name,
            size=16,
            bold=True,
            color=white,
        )
        guide["A1"].fill = PatternFill("solid", fgColor=dark_blue)
        guide["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        guide["A1"].border = Border()
        guide.row_dimensions[1].height = 30
        guide.row_dimensions[2].height = 185
        guide.row_dimensions[3].height = 170
        guide.row_dimensions[4].height = 165
        guide.row_dimensions[5].height = 140

    characters = workbook.create_sheet("Character List")
    characters.sheet_view.showGridLines = False
    characters.merge_cells("A1:C1")
    characters["A1"] = "Character / Performer Role Mapping — 角色 / 演员映射"
    characters["A1"].font = Font(
        name=font_name,
        size=12,
        bold=True,
        color=white,
    )
    characters["A1"].fill = PatternFill(
        "solid",
        fgColor=mapping_dark_blue,
    )
    characters["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
    )
    characters["A1"].border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
    )
    characters.row_dimensions[1].height = 28
    characters.append([
        "DCA Name",
        "Other Script Characters Played",
        "Notes",
    ])
    for cell in characters[2]:
        cell.font = Font(
            name=font_name,
            size=11,
            bold=True,
            italic=True,
            color=mapping_dark_text,
        )
        cell.fill = PatternFill(
            "solid",
            fgColor=mapping_medium_blue,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = mapping_header_border
    characters.row_dimensions[2].height = 30

    for row_index, item in enumerate(project.get("characters", []), 3):
        characters.cell(row_index, 1, str(item.get("dca_name", "")))
        characters.cell(
            row_index,
            2,
            str(item.get("other_characters", "")),
        )
        characters.cell(row_index, 3, "")
        characters.cell(row_index, 1).fill = PatternFill(
            "solid", fgColor=pale_blue
        )
        characters.cell(row_index, 2).fill = PatternFill(
            "solid", fgColor=pale_yellow
        )
        characters.cell(row_index, 3).fill = PatternFill(
            "solid", fgColor=pale_blue
        )
        for cell in characters[row_index]:
            cell.font = Font(name=font_name, size=11, color="FF000000")
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
            )
            cell.border = body_border
        role_line_count = max(
            1,
            str(item.get("other_characters", "")).count("\n") + 1,
        )
        characters.row_dimensions[row_index].height = max(22, 15 * role_line_count)

    # Keep a styled entry row visible when Character List is intentionally
    # unused.  The sheet remains optional, but the exported workbook still
    # looks and behaves like the supplied template.
    if characters.max_row < 3:
        for column in range(1, 4):
            cell = characters.cell(3, column)
            cell.font = Font(name=font_name, size=11, color="FF000000")
            cell.fill = PatternFill(
                "solid",
                fgColor=pale_yellow if column == 2 else pale_blue,
            )
            cell.alignment = Alignment(
                wrap_text=True,
                vertical="center",
            )
            cell.border = body_border

    characters.column_dimensions["A"].width = 32
    characters.column_dimensions["B"].width = 46
    characters.column_dimensions["C"].width = 42
    characters.freeze_panes = None
    characters.auto_filter.ref = f"A2:C{max(3, characters.max_row)}"

    states = workbook.create_sheet("DCA States")
    states.sheet_view.showGridLines = False
    states.merge_cells("A1:R1")
    states["A1"] = "DCA Script Marker — DCA States"
    states["A1"].font = Font(
        name=font_name,
        size=16,
        bold=True,
        color=white,
    )
    states["A1"].fill = PatternFill("solid", fgColor=dark_blue)
    states["A1"].alignment = Alignment(vertical="center")
    states.row_dimensions[1].height = 30
    states.merge_cells("A2:R2")
    states["A2"] = (
        "Page Hint normally uses the page number printed inside the script. "
        "Enter multiple names or aliases on separate lines. / Page Hint 通常填写剧本内印刷页码；"
        "多个名称或别名请分行填写。"
    )
    states["A2"].font = Font(
        name=font_name,
        size=11,
        italic=True,
        color=info_text,
    )
    states["A2"].fill = PatternFill("solid", fgColor=info_blue)
    states["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    states.row_dimensions[2].height = 36
    headers = [
        "DCA State",
        "Start Line Character",
        "Start Line Text",
        "State Start Position",
        "Page Hint",
    ] + [f"DCA {index}" for index in range(1, 13)] + ["Notes"]
    states.append([])
    states.append(headers)
    for cell in states[4]:
        cell.font = Font(
            name=font_name,
            size=11,
            bold=True,
            color=dark_text,
        )
        cell.fill = PatternFill("solid", fgColor=medium_blue)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = header_border
    states.row_dimensions[4].height = 28

    for row_index, item in enumerate(project.get("states", []), 5):
        assignments = list(item.get("dca_assignments", []))[:12]
        assignments += [""] * (12 - len(assignments))
        values = [
            item.get("name", ""),
            item.get("start_line_character", ""),
            item.get("start_line_text", ""),
            item.get("start_position", "After"),
            item.get("page_hint", ""),
        ] + assignments + [item.get("notes", "")]
        for column, value in enumerate(values, 1):
            cell = states.cell(row_index, column, value)
            cell.font = Font(
                name=font_name,
                size=11,
                bold=(column == 1),
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=pale_yellow if 6 <= column <= 17 else pale_blue,
            )
            cell.alignment = Alignment(
                horizontal="center" if column in (1, 5) else None,
                vertical="center",
                wrap_text=True,
            )
            cell.border = body_border
        states.row_dimensions[row_index].height = max(
            34,
            15 * max(
                [str(value).count("\n") + 1 for value in assignments] or [1]
            ),
        )

    if states.max_row < 5:
        for column in range(1, 19):
            cell = states.cell(5, column)
            cell.font = Font(
                name=font_name,
                size=11,
                bold=(column == 1),
            )
            cell.fill = PatternFill(
                "solid",
                fgColor=pale_yellow if 6 <= column <= 17 else pale_blue,
            )
            cell.alignment = Alignment(
                horizontal="center" if column in (1, 5) else None,
                vertical="center",
                wrap_text=True,
            )
            cell.border = body_border

    widths = [20, 24, 48, 20, 14] + [18] * 12 + [24]
    for column, width in enumerate(widths, 1):
        states.column_dimensions[get_column_letter(column)].width = width
    states.freeze_panes = None
    states.auto_filter.ref = f"A4:R{max(5, states.max_row)}"

    position_validation = DataValidation(
        type="list",
        formula1='"Before,After"',
        allow_blank=False,
    )
    states.add_data_validation(position_validation)
    position_validation.add(f"D5:D{max(1000, states.max_row)}")

    validation_choices = []
    validation_choice_keys = set()
    for item in project.get("characters", []):
        dca_name = unicodedata.normalize(
            "NFKC", str(item.get("dca_name", "") or "")
        ).strip()
        dca_key = normalise(dca_name)
        if dca_key and dca_key not in validation_choice_keys:
            validation_choices.append(dca_name)
            validation_choice_keys.add(dca_key)

    # Native .xlsx dropdowns write their visible label. The bracketed role
    # remains compatible with the marker's existing inline-alias parser.
    for item in project.get("characters", []):
        for role in split_display_role_names(item.get("other_characters", "")):
            choice = excel_role_choice(item.get("dca_name", ""), role)
            choice_key = normalise(choice)
            if choice_key and choice_key not in validation_choice_keys:
                validation_choices.append(choice)
                validation_choice_keys.add(choice_key)

    # Hidden dropdown source: DCA Names followed by linked script roles.
    # Names added later in Excel can still be typed directly into a cell.
    for row_index, choice in enumerate(validation_choices, 3):
        characters.cell(row_index, 4, choice)
    characters.column_dimensions["D"].hidden = True
    validation_last_row = max(3, 2 + len(validation_choices))

    character_name_range = DefinedName(
        "DCANameList",
        attr_text=(
            "'Character List'!$D$3:"
            f"$D${validation_last_row}"
        ),
    )
    if hasattr(workbook.defined_names, "add"):
        workbook.defined_names.add(character_name_range)
    else:
        workbook.defined_names.append(character_name_range)
    name_validation = DataValidation(
        type="list",
        formula1="DCANameList",
        allow_blank=True,
    )
    states.add_data_validation(name_validation)
    name_validation.add(f"F5:Q{max(1000, states.max_row)}")
    return workbook


def load_project(filename, diagnostics=None):
    project = read_project_file(filename)
    workbook = project_to_workbook(project, include_guide=False)
    try:
        return _load_template_workbook(workbook, diagnostics)
    finally:
        workbook.close()


def export_project_excel(project_filename, output_filename):
    project = read_project_file(project_filename)
    workbook = project_to_workbook(project, include_guide=True)
    try:
        workbook.save(output_filename)
    finally:
        workbook.close()


def _load_template_workbook(workbook, diagnostics=None):

    if "DCA States" not in workbook.sheetnames:
        raise ValueError(
            'The Excel file needs a sheet named "DCA States".'
        )

    (
        role_groups,
        owner_by_speaker_key,
        role_display_groups,
    ) = load_role_mappings(workbook)
    performer_display_names = {
        performer_key: display_group["performer"]
        for performer_key, display_group in role_display_groups.items()
    }
    states_sheet = workbook["DCA States"]
    state_rows = read_sheet_rows(states_sheet)
    states = []
    assignments = {}
    legend_assignments = {}
    dca_reference_assignments = {}
    named_states_without_cues = []
    cues_without_state_names = []

    for row in state_rows:
        state_name = str(row.get("dca state", "") or "").strip()
        # New template names:
        # Start Line Character / Start Line Text / State Start Position.
        # Old names remain supported for existing templates.
        cue_speaker = str(
            row.get(
                "start line character",
                row.get(
                    "start cue character",
                    row.get("start cue speaker", ""),
                ),
            ) or ""
        ).strip()

        cue_text = str(
            row.get(
                "start line text",
                row.get("start cue text", ""),
            ) or ""
        ).strip()

        start_position = str(
            row.get(
                "state start position",
                row.get("start position", "after"),
            ) or "after"
        ).strip().lower()
        # "Script Page Hint" means the number printed on the script page,
        # not necessarily the PDF's internal page number. Keep "Page Hint"
        # as a fallback so existing templates continue to work.
        page_hint = row.get(
            "script page hint",
            row.get("page hint"),
        )

        if state_name and not cue_text:
            named_states_without_cues.append(state_name)
        elif cue_text and not state_name:
            cues_without_state_names.append(cue_text)

        if state_name and cue_text:
            cue_speaker_names = role_equivalent_names(
                cue_speaker,
                role_groups,
                owner_by_speaker_key,
            )
            states.append({
                "name": state_name,
                "key": normalise(state_name),
                "cue": cue_match_key(cue_text),
                # Optional. When provided, the cue must be spoken by this
                # character, which removes ambiguity when the same line is
                # sung or spoken by more than one person.
                "cue_speaker": normalise(cue_speaker),
                "cue_speaker_names": cue_speaker_names,
                "position": start_position,
                "page_hint": str(page_hint).strip()
                if page_hint is not None else "",
            })

    if "Assignments" in workbook.sheetnames:
        # Keeps compatibility with the original vertical template.
        assignments_sheet = workbook["Assignments"]

        for row in read_sheet_rows(assignments_sheet):
            state_name = str(row.get("dca state", "") or "").strip()
            character = str(row.get("character", "") or "").strip()
            dca = str(row.get("dca", "") or "").strip()

            if state_name and character and dca:
                state_key = normalise(state_name)
                add_template_assignment(
                    assignments,
                    legend_assignments,
                    dca_reference_assignments,
                    state_key,
                    character,
                    dca,
                    role_groups,
                    owner_by_speaker_key,
                    inline_aliases=split_aliases(
                        row.get("aliases", "")
                    ),
                    state_display_name=state_name,
                    performer_display_names=performer_display_names,
                )
    else:
        # Reads the new horizontal template: DCA 1, DCA 2, etc.
        for row in state_rows:
            state_name = str(row.get("dca state", "") or "").strip()

            if not state_name:
                continue

            state_key = normalise(state_name)

            for header, cell_value in row.items():
                dca_match = re.fullmatch(r"dca\s*(\d+)", header)

                if not dca_match:
                    continue

                dca = dca_match.group(1)

                for character, inline_aliases in split_character_entries(
                    cell_value,
                    preserve_display_names=True,
                ):
                    add_template_assignment(
                        assignments,
                        legend_assignments,
                        dca_reference_assignments,
                        state_key,
                        character,
                        dca,
                        role_groups,
                        owner_by_speaker_key,
                        inline_aliases=inline_aliases,
                        state_display_name=state_name,
                        performer_display_names=performer_display_names,
                        )

    for state in states:
        # Matching uses the expanded role names, while legends continue to
        # show the single DCA identity the user selected in the workbook.
        state["legend_assignments"] = legend_assignments.get(
            state["key"],
            {},
        )
        state["performer_role_rows"] = []

        for performer_key, display_group in role_display_groups.items():
            roles = display_group.get("roles", [])
            dca = state["legend_assignments"].get(performer_key)

            # The reference card is deliberately limited to genuine global
            # Character List mappings. Ordinary DCA assignments and inline
            # [aliases] remain out of the card so its meaning stays clear.
            if not roles or dca is None:
                continue

            state["performer_role_rows"].append({
                "dca": list(dca) if isinstance(dca, list) else [dca],
                "performer": display_group["performer"],
                "roles": list(roles),
            })

        state["performer_role_rows"].sort(
            key=lambda row: (
                int(row["dca"][0])
                if row["dca"] and str(row["dca"][0]).isdigit()
                else 999,
                str(row["performer"]).casefold(),
            )
        )

        state["dca_reference_rows"] = []
        for performer_key, reference_row in (
            dca_reference_assignments.get(state["key"], {}).items()
        ):
            display_group = role_display_groups.get(performer_key, {})
            state["dca_reference_rows"].append({
                "dca": list(reference_row["dca"]),
                "performer": display_group.get(
                    "performer",
                    reference_row["performer"],
                ),
                "roles": list(display_group.get("roles", [])),
            })

        state["dca_reference_rows"].sort(
            key=lambda row: (
                int(row["dca"][0])
                if row["dca"] and str(row["dca"][0]).isdigit()
                else 999,
                str(row["performer"]).casefold(),
            )
        )

    if diagnostics is not None:
        configured_state_keys = {
            state["key"] for state in states
        }
        state_display_names = {
            state["key"]: state["name"] for state in states
        }
        assignment_gaps = []
        duplicate_dca_assignments = []
        duplicate_assignment_keys = set()

        for state_key, state_rows in dca_reference_assignments.items():
            for reference_row in state_rows.values():
                dcas = list(dict.fromkeys(
                    str(dca).strip()
                    for dca in reference_row.get("dca", [])
                    if str(dca).strip()
                ))
                if len(dcas) < 2:
                    continue
                duplicate_key = (
                    state_key,
                    speaker_match_key(reference_row.get("performer", "")),
                )
                duplicate_assignment_keys.add(duplicate_key)
                duplicate_dca_assignments.append({
                    "state_key": state_key,
                    "state_name": state_display_names.get(
                        state_key,
                        state_key,
                    ),
                    "dca_name": reference_row.get("performer", ""),
                    "dcas": sorted(
                        dcas,
                        key=lambda value: (
                            0,
                            int(value),
                        ) if value.isdigit() else (1, value),
                    ),
                })

        # Also check resolved DCA identities for repeated assignments.
        # Warnings remain advisory so intentional overlaps are allowed.
        for state_key, state_assignments in assignments.items():
            for performer_key, dca_values in state_assignments.items():
                if performer_key not in performer_display_names:
                    continue
                dcas = list(dict.fromkeys(
                    str(dca).strip()
                    for dca in (
                        dca_values
                        if isinstance(dca_values, list)
                        else [dca_values]
                    )
                    if str(dca).strip()
                ))
                duplicate_key = (
                    state_key,
                    speaker_match_key(performer_key),
                )
                if (
                    len(dcas) < 2
                    or duplicate_key in duplicate_assignment_keys
                ):
                    continue
                duplicate_assignment_keys.add(duplicate_key)
                duplicate_dca_assignments.append({
                    "state_key": state_key,
                    "state_name": state_display_names.get(
                        state_key,
                        state_key,
                    ),
                    "dca_name": performer_display_names.get(
                        performer_key,
                        performer_key,
                    ),
                    "dcas": sorted(
                        dcas,
                        key=lambda value: (
                            0,
                            int(value),
                        ) if value.isdigit() else (1, value),
                    ),
                })

        for state_key, state_assignments in legend_assignments.items():
            assigned_dcas = set()

            for dca_values in state_assignments.values():
                values = (
                    dca_values
                    if isinstance(dca_values, list)
                    else [dca_values]
                )

                for dca in values:
                    dca_text = str(dca).strip()
                    if dca_text.isdigit():
                        assigned_dcas.add(int(dca_text))

            if len(assigned_dcas) < 2:
                continue

            first_dca = min(assigned_dcas)
            last_dca = max(assigned_dcas)
            missing_dcas = [
                dca
                for dca in range(first_dca, last_dca + 1)
                if dca not in assigned_dcas
            ]

            if missing_dcas:
                assignment_gaps.append({
                    "state_key": state_key,
                    "state_name": state_display_names.get(
                        state_key,
                        state_key,
                    ),
                    "missing_dcas": missing_dcas,
                    "first_dca": first_dca,
                    "last_dca": last_dca,
                })

        diagnostics["assignment_gaps"] = assignment_gaps
        diagnostics["duplicate_dca_assignments"] = (
            duplicate_dca_assignments
        )
        diagnostics["named_states_without_cues"] = list(
            dict.fromkeys(named_states_without_cues)
        )
        diagnostics["cues_without_state_names"] = list(
            dict.fromkeys(cues_without_state_names)
        )
        diagnostics["assignment_states_without_start_cues"] = sorted(
            state_key
            for state_key, state_assignments in assignments.items()
            if state_assignments and state_key not in configured_state_keys
        )
        diagnostics["role_mapping_members"] = sum(
            bool(roles) for roles in role_groups.values()
        )
        diagnostics["role_mapping_roles"] = sum(
            len(roles) for roles in role_groups.values()
        )

    return states, assignments



def is_italic(span):
    return bool(span["flags"] & 2)


def is_bold(span):
    """Return true when a PDF span carries reliable bold styling."""
    font_name = str(span.get("font", "")).lower()
    return bool(span.get("flags", 0) & 16) or any(
        weight in font_name
        for weight in ("bold", "semibold", "demibold", "black")
    )


def has_bold_speaker_prefix(line, speaker_names):
    """Use exact bold name spans as evidence for a title-case cue.

    Some English plays print ``Robin Dialogue`` rather than ``ROBIN`` or
    ``Robin: Dialogue``. Text alone is ambiguous with prose such as
    ``Robin enters``. A bold, non-italic first span that contains only the
    workbook name is a strong layout signal while keeping ordinary prose and
    stage directions excluded. A standalone shared cue such as
    ``Avery and Casey`` is equally strong when every recognised name is a
    separate bold span and the complete line contains only that known group.
    """
    if not speaker_names:
        return False

    nonempty_spans = [
        span
        for span in line.get("spans", [])
        if normalise(span.get("text", ""))
    ]
    first_span = nonempty_spans[0] if nonempty_spans else None
    if (
        first_span is None
        or is_italic(first_span)
        or not is_bold(first_span)
    ):
        return False

    if len(speaker_names) == 1:
        return (
            speaker_match_key(first_span.get("text", ""))
            == speaker_match_key(speaker_names[0])
        )

    line_text = "".join(
        span.get("text", "") for span in line.get("spans", [])
    )
    if split_known_speaker_group(
        line_text,
        set(speaker_names),
    ) != speaker_names:
        return False

    bold_name_keys = {
        speaker_match_key(span.get("text", ""))
        for span in nonempty_spans
        if is_bold(span) and not is_italic(span)
    }
    return all(
        speaker_match_key(name) in bold_name_keys
        for name in speaker_names
    )


def cue_speaker_matches(state, speaker_names):
    """Return true if a state has no speaker requirement, or it matches."""
    required_speaker = state.get("cue_speaker", "")

    if not required_speaker:
        return True

    if not speaker_names:
        return False

    required_names = state.get("cue_speaker_names") or [required_speaker]

    return any(
        speaker_match_key(name) == speaker_match_key(required_name)
        or speaker_base_key(name) == speaker_base_key(required_name)
        for name in speaker_names
        for required_name in required_names
    )


def get_matching_state(
    states,
    text,
    page_hints,
    speaker_names=None,
    diagnostics=None,
    pdf_page_number=None,
):
    if isinstance(page_hints, (str, int)):
        page_hints = {str(page_hints)}
    else:
        page_hints = {str(hint) for hint in page_hints}

    text_key = cue_match_key(text)
    cue_matches = [
        state for state in states
        if state["cue"] in text_key
    ]

    # Embedded fonts can make a title unreadable to the PDF text layer while
    # preserving its leading cue code (for example M0 or M1). Use that code
    # only when the workbook also supplies a matching page hint. This recovers
    # the real in-script cue while reducing the risk from contents-page lists.
    if not cue_matches:
        text_identifier = cue_identifier(text)
        if text_identifier:
            cue_matches = [
                state for state in states
                if state.get("page_hint") in page_hints
                and cue_identifier(state.get("cue", "")) == text_identifier
            ]

    # If the template specifies who says the cue, keep only the state whose
    # speaker matches the script's current speaker label. States with a blank
    # Start Cue Speaker retain the original cue-text-only behaviour.
    cue_matches = [
        state for state in cue_matches
        if cue_speaker_matches(state, speaker_names)
    ]

    # A Script Page Hint is an explicit safety boundary. A cue listed in a contents
    # page, song list, or other reference material must not accidentally
    # activate its DCA State before the real script page.
    exact_page_matches = [
        state for state in cue_matches
        if state["page_hint"] in page_hints
    ]

    if exact_page_matches:
        cue_matches = exact_page_matches
    else:
        # Keep Script Page Hint matching strict, but explain the common setup
        # error where the complete Start Cue Text is present on a different
        # PDF page. Substring matches and cue-identifier recovery are excluded
        # because they are not strong enough evidence that the hint is wrong.
        if diagnostics is not None:
            mismatch_candidates = [
                state for state in cue_matches
                if state.get("page_hint")
                and state.get("cue") == text_key
            ]
            if mismatch_candidates:
                mismatches = diagnostics.setdefault(
                    "page_hint_mismatches",
                    [],
                )
                for state in mismatch_candidates:
                    mismatch = {
                        "state_key": state.get("key", ""),
                        "state_name": state.get("name", ""),
                        "cue": state.get("cue", ""),
                        "page_hint": state.get("page_hint", ""),
                        "pdf_page": pdf_page_number,
                        "observed_page_hints": sorted(page_hints),
                    }
                    duplicate = any(
                        existing.get("state_key")
                        == mismatch["state_key"]
                        and existing.get("page_hint")
                        == mismatch["page_hint"]
                        and existing.get("pdf_page")
                        == mismatch["pdf_page"]
                        for existing in mismatches
                    )
                    if not duplicate:
                        mismatches.append(mismatch)

        # States without a Page Hint remain flexible, which is useful for
        # templates where the source PDF has unpredictable pagination.
        cue_matches = [
            state for state in cue_matches
            if not state["page_hint"]
        ]

    if len(cue_matches) == 1:
        return cue_matches[0]

    return None


def find_page_hints(page, pdf_page_number, page_text=None):
    """Return the PDF index plus printed script-page numbers, when available.

    Theatre scripts commonly include cover sheets or contents pages, meaning
    the visible page number can differ from the PDF page index. Printed page
    numbers can appear in either the top header or bottom footer. Both remain
    safe, strict Script Page Hint choices. The PDF index remains as a
    compatibility fallback for templates that do not use printed pages.
    """
    hints = {str(pdf_page_number)}
    # Keep the header band deliberately narrow. Some musical scores contain
    # standalone DCA or bar numbers near the top of the body; treating those
    # as page hints can activate a state on the wrong page.
    upper_page_area = page.rect.height * 0.10
    lower_page_area = page.rect.height * 0.75
    decorated_footer_area = page.rect.height * 0.90

    page_text = page_text or page.get_text("dict")

    for block in page_text["blocks"]:
        if "lines" not in block:
            continue

        for line in block["lines"]:
            line_value = unicodedata.normalize(
                "NFKC",
                "".join(
                    str(span["text"])
                    for span in line["spans"]
                ),
            ).strip()

            # A header page number must occupy the complete visual line. This
            # accepts fragmented digits while rejecting dates and headings
            # such as ``10/22/16`` or ``Scene 2``.
            if (
                re.fullmatch(r"[0-9]{1,4}", line_value)
                and line["bbox"][3] <= upper_page_area
            ):
                hints.add(line_value)

            # Some Word scripts decorate the printed footer number, for
            # example ``~ 1 ~``. Accept it only when the complete footer line
            # consists of one number plus whitespace, tildes, or dash marks.
            # This keeps dates, scene labels, bar numbers, and body text out
            # of the page-hint safety boundary.
            decorated_footer_number = re.fullmatch(
                r"[\s~〜～\-‐‑‒–—―−]*"
                r"([0-9]{1,4})"
                r"[\s~〜～\-‐‑‒–—―−]*",
                line_value,
            )
            if (
                decorated_footer_number
                and line["bbox"][1] >= decorated_footer_area
            ):
                hints.add(decorated_footer_number.group(1))

            # Some scripts print the current script page together with the
            # document total, for example ``4/47`` or ``Page 4 of 47``. The
            # first number is still the user's Script Page Hint. Accept this
            # only as a complete line in the guarded header/footer bands so a
            # fraction or cue number inside the script body cannot activate a
            # state on the wrong page.
            page_counter = re.fullmatch(
                r"(?:page\s*)?([0-9]{1,4})\s*(?:/|of)\s*[0-9]{1,4}",
                line_value,
                flags=re.IGNORECASE,
            )
            if (
                page_counter
                and (
                    line["bbox"][3] <= upper_page_area
                    or line["bbox"][1] >= decorated_footer_area
                )
            ):
                hints.add(page_counter.group(1))

            for span in line["spans"]:
                value = str(span["text"]).strip()
                is_footer_number = (
                    value.isdigit()
                    and span["bbox"][1] >= lower_page_area
                )

                if is_footer_number:
                    hints.add(value)

    return hints


def page_has_cast_reference_heading(page_text):
    """Identify a cast/character reference page by a strong heading.

    A role list can contain many bare workbook names that are not dialogue
    cues. Keep this deliberately heading-based so a script remains free to
    switch between cast-tagged and ordinary dialogue labels on real scenes.
    """
    values = []
    for block in page_text.get("blocks", []):
        for line in block.get("lines", []):
            values.append(
                "".join(
                    str(span.get("text", ""))
                    for span in line.get("spans", [])
                )
            )

    page_value = unicodedata.normalize("NFKC", "\n".join(values))
    compact_value = re.sub(r"\s+", "", page_value)
    if any(
        heading in compact_value
        for heading in (
            "演员分配",
            "演员表",
            "角色表",
            "人物表",
            "主要角色",
            "译名对照",
        )
    ):
        return True

    english_value = normalise(page_value)
    return bool(
        re.search(
            r"\b(?:cast|character)\s+list\b"
            r"|\bcast\s+of\s+characters\b"
            r"|\bdramatis\s+person(?:ae|æ)\b",
            english_value,
        )
    )


def load_ocr_pages(ocr_json_file, document):
    """Load Vision OCR lines exported by the Mac app.

    The OCR file has one list item per PDF page. Each recognised line keeps
    its original PDF-coordinate bounding box, so the normal marker can still
    place DCA numbers beside the correct printed line.
    """
    with open(ocr_json_file, encoding="utf-8") as file:
        raw_pages = json.load(file)

    if not isinstance(raw_pages, list):
        raise ValueError("OCR data must contain a list of PDF pages.")
    if len(raw_pages) != len(document):
        raise ValueError(
            "OCR data does not match this PDF. Please choose the same PDF "
            "again and rerun OCR."
        )

    pages = []
    for raw_page in raw_pages:
        raw_lines = raw_page.get("lines", []) if isinstance(raw_page, dict) else []
        lines = []
        for raw_line in raw_lines:
            text = str(raw_line.get("text", "")).strip()
            bbox = raw_line.get("bbox", [])
            if not text or not isinstance(bbox, list) or len(bbox) != 4:
                continue
            try:
                x0, y0, width, height = (float(value) for value in bbox)
            except (TypeError, ValueError):
                continue
            if width <= 0 or height <= 0:
                continue
            line_bbox = (x0, y0, x0 + width, y0 + height)
            lines.append(
                {
                    "bbox": line_bbox,
                    "spans": [
                        {
                            "text": text,
                            "bbox": line_bbox,
                            "size": max(8, height * 0.82),
                            "flags": 0,
                        }
                    ],
                }
            )
        pages.append({"blocks": [{"lines": lines}]})
    return pages


def build_performer_role_mapping_text(state):
    """Build the compact reference card shown for one activated state."""
    rows = state.get("performer_role_rows", [])
    if not rows:
        return ""

    lines = [f'{state["name"]} - Performer / Role Mapping']
    for row in rows:
        lines.append(
            f'DCA {display_dca(row["dca"])} | {row["performer"]}: '
            + " / ".join(row["roles"])
        )

    return "\n".join(lines)


def add_bordered_freetext_annotation(
    document,
    page,
    rectangle,
    text,
    fontsize,
    font_name,
    font_file,
    text_colour,
    border_colour,
    align=0,
    bold=False,
):
    """Create one movable annotation containing both text and its border."""
    if border_colour == text_colour:
        annotation = page.add_freetext_annot(
            rectangle,
            text,
            fontsize=fontsize,
            fontname=font_name,
            text_color=text_colour,
            fill_color=None,
            align=align,
        )
        annotation.set_border(width=0.8)
        annotation.update()
        return annotation

    alignment = {0: "left", 1: "center", 2: "right"}.get(align, "left")
    font_weight = "bold" if bold else "normal"
    text_style = (
        "font-family: "
        f"{rich_text_font_family(font_name, font_file)}; "
        f"font-size: {fontsize:g}pt; "
        f"font-weight: {font_weight}; "
        f"color: {css_colour(text_colour)}; "
        f"text-align: {alignment}; "
        "margin: 0; padding: 0; line-height: 1.15;"
    )
    escaped_html = "<br/>".join(
        html.escape(line) for line in text.splitlines()
    )
    annotation = page.add_freetext_annot(
        rectangle,
        escaped_html,
        richtext=True,
        style=text_style,
        border_width=0.8,
        fill_color=None,
        align=align,
    )
    rich_content = (
        '<?xml version="1.0"?>'
        '<body xmlns="http://www.w3.org/1999/xhtml" '
        'xmlns:xfa="http://www.xfa.org/schema/xfa-data/1.0/" '
        'xfa:contentType="text/html" '
        'xfa:APIVersion="Acrobat:8.0.0" '
        'xfa:spec="2.4">'
        f"{escaped_html}</body>"
    )
    document.xref_set_key(
        annotation.xref,
        "RC",
        fitz.get_pdf_str(rich_content),
    )
    annotation.update(text_color=border_colour)
    document.xref_set_key(
        annotation.xref,
        "Contents",
        fitz.get_pdf_str(text),
    )
    return annotation


def rectangle_intersection_area(first, second):
    intersection = fitz.Rect(first)
    intersection.intersect(fitz.Rect(second))
    if intersection.is_empty:
        return 0
    return intersection.width * intersection.height


def choose_performer_role_card_box(page, width, height):
    """Prefer the upper-left, while avoiding existing text and annotations."""
    page_rect = fitz.Rect(page.rect)
    horizontal_margin = 18
    vertical_margin = 42
    width = min(width, max(80, page_rect.width - 2 * horizontal_margin))
    height = min(height, max(40, page_rect.height - 2 * vertical_margin))
    right_x = max(horizontal_margin, page_rect.width - horizontal_margin - width)
    bottom_y = max(vertical_margin, page_rect.height - vertical_margin - height)

    occupied = []
    for block in page.get_text("blocks"):
        block_rect = fitz.Rect(block[:4])
        if block_rect.is_empty:
            continue
        occupied.append(
            fitz.Rect(
                max(page_rect.x0, block_rect.x0 - 3),
                max(page_rect.y0, block_rect.y0 - 3),
                min(page_rect.x1, block_rect.x1 + 3),
                min(page_rect.y1, block_rect.y1 + 3),
            )
        )

    for annotation in page.annots() or []:
        occupied.append(fitz.Rect(annotation.rect))

    candidates = [
        fitz.Rect(horizontal_margin, vertical_margin, horizontal_margin + width, vertical_margin + height),
        fitz.Rect(right_x, vertical_margin, right_x + width, vertical_margin + height),
        fitz.Rect(horizontal_margin, bottom_y, horizontal_margin + width, bottom_y + height),
        fitz.Rect(right_x, bottom_y, right_x + width, bottom_y + height),
    ]

    x_positions = list(dict.fromkeys([
        horizontal_margin,
        right_x,
        max(horizontal_margin, (page_rect.width - width) / 2),
    ]))
    maximum_y = max(vertical_margin, page_rect.height - vertical_margin - height)
    y = vertical_margin
    while y <= maximum_y + 0.1:
        for x in x_positions:
            candidate = fitz.Rect(x, y, x + width, y + height)
            if candidate not in candidates:
                candidates.append(candidate)
        y += 12

    best_candidate = candidates[0]
    best_score = math.inf
    for candidate in candidates:
        score = sum(
            rectangle_intersection_area(candidate, rectangle)
            for rectangle in occupied
        )
        if score == 0:
            return candidate
        if score < best_score:
            best_candidate = candidate
            best_score = score

    return best_candidate


def add_performer_role_mapping_card(document, page, state, state_style):
    """Add one editable performer/role reference card to a PDF page."""
    text = build_performer_role_mapping_text(state)
    if not text:
        return None

    text_colour = state_style.get(
        "page_header_footer_text_colour",
        state_style.get("colour", STATE_COLOUR),
    )
    border_colour = state_style.get(
        "page_header_footer_border_colour",
        text_colour,
    )
    font_name = state_style.get(
        "page_header_footer_font_name",
        state_style.get("font_name", "heiti"),
    )
    font_file = state_style.get(
        "page_header_footer_font_file",
        state_style.get("font_file", CHINESE_FONT_FILE),
    )
    font_size = max(
        7,
        min(
            9.5,
            state_style.get(
                "page_header_footer_size",
                state_style.get("size", 12),
            ) * 0.68,
        ),
    )

    if contains_cjk(text):
        font_name = "heiti"
        font_file = CHINESE_FONT_FILE

    text_font = (
        fitz.Font(fontfile=font_file)
        if font_file
        else fitz.Font(fontname=font_name)
    )
    card_width = min(
        250,
        max(80, page.rect.width - 36),
        max(205, page.rect.width * 0.42),
    )
    usable_width = max(60, card_width - 14)
    visual_line_count = 0
    for line in text.splitlines():
        line_width = text_font.text_length(line or " ", fontsize=font_size)
        visual_line_count += max(1, math.ceil(line_width / usable_width))
    card_height = max(42, 14 + visual_line_count * font_size * 1.28)
    card_box = choose_performer_role_card_box(
        page,
        card_width,
        card_height,
    )

    return add_bordered_freetext_annotation(
        document,
        page,
        card_box,
        text,
        font_size,
        font_name,
        font_file,
        text_colour,
        border_colour,
        align=0,
    )

def build_legend_text(state, assignments):
    state_assignments = state.get("legend_assignments")
    if state_assignments is None:
        state_assignments = assignments.get(state["key"], {})

    legend_items = sorted(
        state_assignments.items(),
        key=lambda item: (
            int(item[1][0])
            if isinstance(item[1], list) and item[1][0].isdigit()
            else int(item[1])
            if str(item[1]).isdigit()
            else 999
        ),
    )
    lines = [state["name"]]

    for character, dca in legend_items:
        lines.append(f"{display_dca(dca)}: {character}")

    return "\n".join(lines)


def save_document_atomically(document, output_file):
    """Save a complete PDF before replacing an existing marked output.

    Preview may still need an already-open PDF to be closed and reopened, but
    presenting the finished file in one replacement event avoids a transient
    missing or half-written output and preserves the previous PDF if saving
    or validation fails.
    """
    output_directory = os.path.dirname(os.path.abspath(output_file))
    output_name = os.path.basename(output_file)
    page_count = len(document)
    descriptor, temporary_output = tempfile.mkstemp(
        prefix=f".{output_name}.",
        suffix=".tmp.pdf",
        dir=output_directory,
    )
    os.close(descriptor)

    try:
        os.remove(temporary_output)
        document.save(
            temporary_output,
            garbage=4,
            clean=True,
            deflate=True,
            deflate_images=True,
            deflate_fonts=True,
        )
        document.close()

        saved_document = fitz.open(temporary_output)
        try:
            if len(saved_document) != page_count:
                raise RuntimeError(
                    "The replacement PDF did not preserve every page."
                )
        finally:
            saved_document.close()

        os.replace(temporary_output, output_file)
    finally:
        if not document.is_closed:
            document.close()
        if os.path.exists(temporary_output):
            os.remove(temporary_output)


def mark_pdf(
    states,
    assignments,
    pdf_file,
    output_file,
    editable=False,
    first_appearance=False,
    legend_only=False,
    legend_overrides=None,
    legend_style=None,
    number_style=None,
    state_style=None,
    start_page=None,
    end_page=None,
    ocr_json_file=None,
    diagnostics=None,
):
    document = fitz.open(pdf_file)
    pdf_page_count = len(document)
    ocr_pages = load_ocr_pages(ocr_json_file, document) if ocr_json_file else None
    current_state = None
    marked_count = 0
    unmatched_names = []
    activated_states = set()
    marked_speakers = set()
    marked_cue_lines = set()
    marked_pages = set()
    marked_page_counts = {}
    marked_cue_counts = {}
    state_activation_pages = {}
    performer_role_mapping_pages = {}
    performer_role_mapping_states = set()
    unassigned_known_speakers = []
    unassigned_known_speaker_keys = set()
    number_style = number_style or {}
    state_style = state_style or {}
    page_state_display = state_style.get("page_state_display")
    if page_state_display not in {"off", "header", "footer", "both"}:
        # Preserve direct callers that use the original Boolean setting.
        page_state_display = (
            "both"
            if state_style.get("page_header_footer", False)
            else "off"
        )

    number_colour = number_style.get(
        "colour",
        NUMBER_COLOUR,
    )
    number_scale = number_style.get(
        "scale",
        NUMBER_SCALE,
    )
    number_font = number_style.get(
        "font_name",
        "helv",
    )
    # The label ends a fixed distance before the speaker name. Long labels
    # therefore grow left into the gutter instead of across the script text.
    number_gap = number_style.get("gap", 16)
    number_vertical_offset = number_style.get("vertical_offset", 0)
    state_colour = state_style.get(
        "colour",
        STATE_COLOUR,
    )
    state_font_name = state_style.get(
        "font_name",
        "heiti",
    )
    state_font_file = state_style.get(
        "font_file",
        CHINESE_FONT_FILE,
    )
    state_names = {state["key"]: state["name"] for state in states}
    states_by_key = {state["key"]: state for state in states}
    show_performer_role_mapping = bool(
        state_style.get("show_performer_role_mapping", False)
    )
    all_template_characters = {
        character
        for state_assignments in assignments.values()
        for character in state_assignments
    }
    # A speaker name and its dialogue can occupy separate PDF lines. Keep the
    # most recent real speaker label so Start Cue Speaker still works there.
    current_cue_speakers = []
    previous_speaker_columns = []
    cast_track_labels_seen = False

    for page_number, page in enumerate(document, start=1):
        # The selected page range is a strict export boundary.  We still
        # inspect pages before a selected range so the active DCA State can
        # be established, but once the end page has been reached there is
        # nothing later to mark.  Breaking here guarantees every following
        # PDF page is copied without any DCA numbers, state labels, headers,
        # footers, or editable annotations.
        if end_page is not None and page_number > end_page:
            break

        page_is_selected = (
            (start_page is None or page_number >= start_page)
            and (end_page is None or page_number <= end_page)
        )
        page_text = (
            ocr_pages[page_number - 1]
            if ocr_pages is not None
            else page.get_text("dict")
        )
        visible_span_left_edges = {}
        if ocr_pages is None:
            visible_span_left_edges = padded_span_visible_left_edges(
                page, page_text
            )
            split_embedded_right_speaker_lines(
                page,
                page_text,
                all_template_characters,
            )
        page_is_cast_reference = page_has_cast_reference_heading(page_text)
        page_hint_values = find_page_hints(
            page, page_number, page_text=page_text
        )
        # PDF extraction can split one printed line into separate fragments.
            # For example: ``ROBIN &`` and ``AVERY`` may share the same
        # baseline but arrive as two independent lines. Keep a combined view
        # for speaker recognition and state headings split across same-row
        # fragments.
        physical_lines = [
            line
            for block in page_text["blocks"]
            if "lines" in block
            for line in block["lines"]
            # Some Word-generated PDFs emit a line containing only one space
            # at every tab stop. Those invisible fragments must not bridge
            # two genuine speaker columns into one visual cue group.
            if any(
                normalise(span.get("text", ""))
                for span in line.get("spans", [])
            )
        ]
        visual_line_texts = {}
        visual_row_left_edges = {}

        for physical_line in physical_lines:
            row_y = physical_line["bbox"][1]
            same_row = [
                candidate
                for candidate in physical_lines
                if abs(candidate["bbox"][1] - row_y) < 0.75
            ]
            same_row.sort(key=lambda candidate: candidate["bbox"][0])

            # Most PDFs split one printed row into nearby fragments (speaker
            # name + dialogue), which should remain one visual cue. A duet
            # may instead place two independent speaker columns on the same
            # row: ALEX. on the left and BLAIR. on the right. A large empty
            # horizontal gap means those are separate cue groups and both
            # labels must be eligible for a DCA number. Narrower theatre
            # columns are split only when the incoming fragment is itself a
            # standalone speaker label from the template.
            cue_groups = []
            current_group = []
            previous_right = None
            for candidate in same_row:
                candidate_left = candidate["bbox"][0]
                gap = (
                    candidate_left - previous_right
                    if previous_right is not None
                    else 0
                )
                current_group_text = "".join(
                    "".join(span["text"] for span in grouped["spans"])
                    for grouped in current_group
                )
                candidate_text = "".join(
                    span["text"] for span in candidate["spans"]
                )
                candidate_nonempty_spans = [
                    span
                    for span in candidate.get("spans", [])
                    if normalise(span.get("text", ""))
                ]
                candidate_is_split_period_label = bool(
                    candidate_nonempty_spans
                    and not is_italic(candidate_nonempty_spans[0])
                    and get_split_english_speaker_fragment_names(
                        candidate_text,
                        all_template_characters,
                    )
                )
                current_group_has_speaker = any(
                    is_standalone_speaker_label(
                        "".join(
                            span["text"]
                            for span in grouped["spans"]
                        ),
                        all_template_characters,
                    )
                    for grouped in current_group
                )
                current_group_leading_names = [
                    name
                    for grouped in current_group
                    for name in get_leading_known_speaker_names(
                        "".join(
                            span["text"]
                            for span in grouped["spans"]
                        ),
                        all_template_characters,
                    )
                ]
                candidate_leading_names = (
                    get_leading_known_speaker_names(
                        candidate_text,
                        all_template_characters,
                    )
                )
                close_speaker_columns = (
                    current_group
                    and gap > 45
                    and current_group_has_speaker
                    and is_standalone_speaker_label(
                        candidate_text,
                        all_template_characters,
                    )
                    and not speaker_label_continues(current_group_text)
                )
                tight_parallel_speaker_columns = (
                    current_group
                    and gap > 24
                    and candidate_left >= page.rect.width * 0.45
                    and candidate_left
                    - float(current_group[0]["bbox"][0])
                    >= page.rect.width * 0.30
                    and current_group_leading_names
                    and candidate_leading_names
                    and not set(current_group_leading_names).intersection(
                        candidate_leading_names
                    )
                    and not speaker_label_continues(current_group_text)
                )
                right_hand_split_column = (
                    current_group
                    and gap > 45
                    and candidate_left >= page.rect.width * 0.45
                    and candidate_is_split_period_label
                    and not speaker_label_continues(current_group_text)
                )
                if current_group and (
                    gap > 60
                    or close_speaker_columns
                    or tight_parallel_speaker_columns
                    or right_hand_split_column
                ):
                    cue_groups.append(current_group)
                    current_group = []
                current_group.append(candidate)
                previous_right = candidate["bbox"][2]
            if current_group:
                cue_groups.append(current_group)

            cue_group = next(
                group for group in cue_groups if physical_line in group
            )
            visual_line_texts[id(physical_line)] = (
                join_visual_line_fragments(
                    cue_group,
                    all_template_characters,
                )
            )
            # Only the leftmost fragment within its own cue group can receive
            # a number. With a two-column duet, each column has its own
            # group and therefore its own leftmost speaker label.
            visual_row_left_edges[id(physical_line)] = min(
                candidate["bbox"][0] for candidate in cue_group
            )

        split_speaker_columns, split_layout_names_by_line = (
            find_split_english_speaker_layout(
                physical_lines,
                all_template_characters,
                previous_speaker_columns,
                page.rect.width,
                page_is_cast_reference=page_is_cast_reference,
            )
        )

        # Colons provide unambiguous examples of the page's true speaker
        # column. Use them to distinguish an aligned ``林青。`` cue from the
        # same words indented inside another character's dialogue. Pages that
        # contain no colon labels inherit the most recently established
        # columns, while scripts that never use colons retain old behaviour.
        page_speaker_columns = []
        for physical_line in physical_lines:
            row_left = visual_row_left_edges.get(
                id(physical_line),
                physical_line["bbox"][0],
            )
            if abs(physical_line["bbox"][0] - row_left) >= 0.75:
                continue

            row_text = visual_line_texts.get(
                id(physical_line),
                "".join(
                    span["text"] for span in physical_line["spans"]
                ),
            )
            if get_explicit_speaker_names(
                row_text,
                all_template_characters,
            ):
                page_speaker_columns.append(row_left)

        current_speaker_columns = list(dict.fromkeys(
            page_speaker_columns + split_speaker_columns
        ))
        trusted_speaker_columns = list(dict.fromkeys(
            previous_speaker_columns + current_speaker_columns
        ))
        if current_speaker_columns:
            previous_speaker_columns = current_speaker_columns
        # This is the state already active as the page begins. On the first
        # page it may be unknown until the first cue has been found.
        page_start_state = current_state

        for block in page_text["blocks"]:
            if "lines" not in block:
                continue

            for line in block["lines"]:
                line_text = "".join(
                    span["text"] for span in line["spans"]
                )
                speaker_line_text = visual_line_texts.get(
                    id(line),
                    line_text,
                )
                is_visual_row_anchor = abs(
                    line["bbox"][0]
                    - visual_row_left_edges.get(
                        id(line), line["bbox"][0]
                    )
                ) < 0.75
                speaker_row_x = visual_row_left_edges.get(
                    id(line),
                    line["bbox"][0],
                )
                line_clean_text = normalise(line_text)
                line_speaker_names = get_speaker_names(
                    speaker_line_text,
                    all_template_characters,
                )
                line_layout_speaker_names = (
                    split_layout_names_by_line.get(id(line), [])
                )
                raw_speaker_line_text = unicodedata.normalize(
                    "NFKC",
                    str(speaker_line_text),
                ).strip()
                line_has_cast_track_prefix = (
                    strip_cast_track_prefix(raw_speaker_line_text)
                    != raw_speaker_line_text
                )

                line_is_known_speaker_label = bool(
                    line_speaker_names
                    and all(
                        name in all_template_characters
                        for name in line_speaker_names
                    )
                    and looks_like_positioned_speaker_label(
                        speaker_line_text,
                        line_speaker_names,
                        speaker_row_x,
                        trusted_speaker_columns,
                        bold_prefix=has_bold_speaker_prefix(
                            line,
                            line_speaker_names,
                        ),
                        layout_speaker_names=line_layout_speaker_names,
                    )
                )
                if (
                    line_is_known_speaker_label
                    and line_has_cast_track_prefix
                ):
                    cast_track_labels_seen = True

                # A cast/translation table can repeat every workbook name as
                # an isolated cell. Ignore those cells only when this script
                # has established cast-track dialogue and this page carries a
                # strong cast-reference heading. Genuine later scenes remain
                # free to use ordinary untagged speaker labels.
                ignore_untagged_bare_name = bool(
                    cast_track_labels_seen
                    and page_is_cast_reference
                    and is_untagged_punctuationless_speaker_label(
                        speaker_line_text,
                        line_speaker_names,
                    )
                )
                if ignore_untagged_bare_name:
                    line_speaker_names = []
                    line_is_known_speaker_label = False

                if line_is_known_speaker_label:
                    current_cue_speakers = line_speaker_names

                line_next_state = get_matching_state(
                    states,
                    line_clean_text,
                    page_hint_values,
                    current_cue_speakers,
                    diagnostics=diagnostics,
                    pdf_page_number=page_number,
                )
                # Word and WPS PDFs can expose one printed state heading as
                # separate same-baseline lines, for example ``I.`` and
                # ``黎明之前``. The physical fragments cannot match the
                # workbook cue individually. Retry with the already guarded
                # visual cue group, but only at its leftmost anchor so the
                # state is activated once and separate columns stay separate.
                if (
                    line_next_state is None
                    and is_visual_row_anchor
                    and normalise(speaker_line_text) != line_clean_text
                ):
                    line_next_state = get_matching_state(
                        states,
                        speaker_line_text,
                        page_hint_values,
                        current_cue_speakers,
                        diagnostics=diagnostics,
                        pdf_page_number=page_number,
                    )
                # A short cue can occur again in prose after its state has
                # already started (for example ECHO and later "Echo music").
                # Do not draw or re-activate the state that is
                # already current. Returning to it after a different state
                # remains supported because current_state will then differ.
                if (
                    line_next_state
                    and line_next_state["key"] == current_state
                ):
                    line_next_state = None
                state_anchor_index = next(
                    (
                        index
                        for index, candidate in enumerate(line["spans"])
                        if normalise(candidate["text"])
                    ),
                    None,
                )

                for span_index, span in enumerate(line["spans"]):
                    text = span["text"].strip()
                    clean_text = normalise(text)
                    # A state cue may be split into several PDF text spans.
                    # Match the complete printed line, then act only once.
                    next_state = (
                        line_next_state
                        if span_index == state_anchor_index
                        else None
                    )

                    mapping_state = current_state
                    possible_characters = assignments.get(
                        mapping_state or "", {}
                    )
                    # Some PDFs split one speaker name across several spans
                    # (for example 林 / · / 海). Match the complete line.
                    speaker_names = get_speaker_names(
                        speaker_line_text, possible_characters
                    )
                    if ignore_untagged_bare_name:
                        speaker_names = []
                    speaker_name = speaker_names[0] if speaker_names else ""

                    # An "After" cue can occur inside the first line spoken
                    # in the new DCA State. In that case mark the speaker from
                    # the new state, while still placing the state label after
                    # the cue line.
                    if (
                        line_next_state
                        and line_next_state["position"] == "after"
                    ):
                        candidate_state = line_next_state["key"]
                        candidate_characters = assignments.get(
                            candidate_state, {}
                        )
                        candidate_speakers = get_speaker_names(
                            speaker_line_text, candidate_characters
                        )
                        if ignore_untagged_bare_name:
                            candidate_speakers = []

                        if (
                            candidate_speakers
                            and all(
                                name in candidate_characters
                                for name in candidate_speakers
                            )
                            and looks_like_positioned_speaker_label(
                                speaker_line_text,
                                candidate_speakers,
                                speaker_row_x,
                                trusted_speaker_columns,
                                bold_prefix=has_bold_speaker_prefix(
                                    line,
                                    candidate_speakers,
                                ),
                                layout_speaker_names=(
                                    line_layout_speaker_names
                                ),
                            )
                        ):
                            mapping_state = candidate_state
                            possible_characters = candidate_characters
                            speaker_names = candidate_speakers
                            speaker_name = speaker_names[0]

                    if not clean_text:
                        continue

                    diagnostic_state = mapping_state
                    if next_state and next_state["position"] == "before":
                        diagnostic_state = next_state["key"]

                    if (
                        diagnostics is not None
                        and span_index == state_anchor_index
                        and page_is_selected
                        and not legend_only
                        and not is_italic(span)
                        and is_visual_row_anchor
                        and diagnostic_state
                        and line_is_known_speaker_label
                    ):
                        missing_speakers = [
                            name
                            for name in line_speaker_names
                            if name not in assignments.get(
                                diagnostic_state, {}
                            )
                        ]
                        diagnostic_key = (
                            page_number,
                            diagnostic_state,
                            round(line["bbox"][1], 1),
                            round(line["bbox"][0], 1),
                            tuple(missing_speakers),
                        )
                        if (
                            missing_speakers
                            and diagnostic_key
                            not in unassigned_known_speaker_keys
                        ):
                            unassigned_known_speaker_keys.add(
                                diagnostic_key
                            )
                            unassigned_known_speakers.append({
                                "page": page_number,
                                "state": diagnostic_state,
                                "speakers": missing_speakers,
                                "label": speaker_line_text.strip(),
                            })

                    if next_state and page_is_selected:
                        cue_box = fitz.Rect(line["bbox"])

                        if legend_only:
                            if legend_style is None:
                                legend_style = {}

                            legend_position = legend_style.get(
                                "position", "Left Gutter"
                            )

                            # Legends always live on the left. "Near Script"
                            # keeps the list near the script boundary, while
                            # "Left Gutter" starts it at the page edge.
                            legend_right = max(95, cue_box.x0 - 8)
                            legend_left = 18
                            if legend_position == "Near Script":
                                legend_left = max(18, legend_right - 120)

                            legend_box = fitz.Rect(
                                legend_left,
                                cue_box.y1 + 4,
                                legend_right,
                                cue_box.y1 + 180,
                            )

                            legend_text = build_legend_text(
                                next_state, assignments
                            )

                            if legend_overrides:
                                legend_text = legend_overrides.get(
                                    next_state["key"],
                                    legend_text,
                                )

                            legend_colour = legend_style.get(
                                "colour",
                                STATE_COLOUR,
                            )
                            legend_size = legend_style.get(
                                "size",
                                8,
                            )
                            legend_font = legend_style.get(
                                "font_name",
                                "heiti",
                            )
                            legend_font_file = legend_style.get(
                                "font_file",
                                CHINESE_FONT_FILE,
                            )

                            # Built-in Times / Helvetica / Courier fonts do
                            # not contain Chinese glyphs. Keep the selected
                            # colour and size, but safely fall back for text.
                            if contains_cjk(legend_text):
                                legend_font = "heiti"
                                legend_font_file = CHINESE_FONT_FILE

                            legend_annotation = page.add_freetext_annot(
                                legend_box,
                                legend_text,
                                fontsize=legend_size,
                                fontname=legend_font,
                                text_color=legend_colour,
                                fill_color=None,
                                align=0,
                            )

                            legend_annotation.set_border(width=0)
                            legend_annotation.update()

                        else:
                            state_font_size = state_style.get(
                                "size",
                                max(
                                    10,
                                    span["size"] * number_scale,
                                ),
                            )

                            draw_state_font_name = state_font_name
                            draw_state_font_file = state_font_file

                            if contains_cjk(next_state["name"]):
                                draw_state_font_name = "heiti"
                                draw_state_font_file = CHINESE_FONT_FILE

                            if draw_state_font_file:
                                state_font = fitz.Font(
                                    fontfile=draw_state_font_file
                                )
                            else:
                                state_font = fitz.Font(
                                    fontname=draw_state_font_name
                                )

                            state_position = state_style.get(
                                "position", "Left Gutter"
                            )
                            # The placement follows the meaning of the
                            # template's Start position.  A Before state is
                            # shown beside the cue that starts it; an After
                            # state is shown below the cue that finishes the
                            # previous state.
                            state_placement = (
                                "Below Cue"
                                if next_state["position"] == "after"
                                else "Beside Cue"
                            )
                            # A DCA State can begin on the same printed line
                            # as a character cue.  In that situation the state
                            # label and DCA number would compete for
                            # the same gutter.  Move only the state label above
                            # the cue automatically, keeping both readable.
                            state_line_speakers = get_speaker_names(
                                speaker_line_text,
                                assignments.get(next_state["key"], {}),
                            )
                            if ignore_untagged_bare_name:
                                state_line_speakers = []
                            state_shares_cue_line = bool(
                                state_line_speakers
                                and looks_like_positioned_speaker_label(
                                    speaker_line_text,
                                    state_line_speakers,
                                    speaker_row_x,
                                    trusted_speaker_columns,
                                    bold_prefix=has_bold_speaker_prefix(
                                        line,
                                        state_line_speakers,
                                    ),
                                    layout_speaker_names=(
                                        line_layout_speaker_names
                                    ),
                                )
                            )
                            state_width = state_font.text_length(
                                next_state["name"], fontsize=state_font_size
                            )

                            if state_position == "Left Gutter":
                                # Keep the state in the left gutter, but
                                # right-align its *end* before the DCA number
                                # number column.  This matters for longer
                                # labels such as "Scene 8": a fixed left
                                # starting point lets the label grow right
                                # into the number.  A fixed right edge keeps
                                # both markings readable on every script.
                                # Do not follow the cue's x-position here.
                                # A first cue is often a centred scene title,
                                # which would pull its state label into the
                                # middle of the page. The fixed right edge
                                # makes Left Gutter a real, stable column.
                                state_right = 82
                                state_x = max(8, state_right - state_width)
                            elif state_position == "Near Script":
                                # Retained for old templates that explicitly
                                # want the label to follow each cue's indent.
                                state_x = max(
                                    12,
                                    cue_box.x0 - number_gap - state_width,
                                )
                            else:
                                # Far from Script is a visibly separate,
                                # far-left gutter.
                                state_x = 12

                            avoid_following_speaker_number = False
                            if (
                                state_position == "Left Gutter"
                                and next_state["position"] == "after"
                            ):
                                proposed_state_y = (
                                    cue_box.y1 + state_font_size + 4
                                )
                                proposed_state_box = fitz.Rect(
                                    state_x - 3,
                                    proposed_state_y - state_font_size - 4,
                                    state_x + state_width + 5,
                                    proposed_state_y + 4,
                                )
                                next_assignments = assignments.get(
                                    next_state["key"],
                                    {},
                                )

                                for candidate_line in sorted(
                                    physical_lines,
                                    key=lambda candidate: (
                                        candidate["bbox"][1],
                                        candidate["bbox"][0],
                                    ),
                                ):
                                    if (
                                        candidate_line["bbox"][1]
                                        <= cue_box.y1 - 0.5
                                    ):
                                        continue
                                    if (
                                        candidate_line["bbox"][1]
                                        >= proposed_state_box.y1
                                    ):
                                        break

                                    candidate_spans = [
                                        candidate
                                        for candidate in candidate_line.get(
                                            "spans", []
                                        )
                                        if normalise(
                                            candidate.get("text", "")
                                        )
                                    ]
                                    if (
                                        not candidate_spans
                                        or is_italic(candidate_spans[0])
                                    ):
                                        continue

                                    candidate_row_x = (
                                        visual_row_left_edges.get(
                                            id(candidate_line),
                                            candidate_line["bbox"][0],
                                        )
                                    )
                                    if abs(
                                        candidate_line["bbox"][0]
                                        - candidate_row_x
                                    ) >= 0.75:
                                        continue

                                    candidate_text = visual_line_texts.get(
                                        id(candidate_line),
                                        "".join(
                                            candidate["text"]
                                            for candidate in candidate_line[
                                                "spans"
                                            ]
                                        ),
                                    )
                                    candidate_names = get_speaker_names(
                                        candidate_text,
                                        next_assignments,
                                    )
                                    if (
                                        not candidate_names
                                        or not all(
                                            name in next_assignments
                                            for name in candidate_names
                                        )
                                        or not looks_like_positioned_speaker_label(
                                            candidate_text,
                                            candidate_names,
                                            candidate_row_x,
                                            trusted_speaker_columns,
                                            bold_prefix=(
                                                has_bold_speaker_prefix(
                                                    candidate_line,
                                                    candidate_names,
                                                )
                                            ),
                                            layout_speaker_names=(
                                                split_layout_names_by_line.get(
                                                    id(candidate_line),
                                                    [],
                                                )
                                            ),
                                        )
                                    ):
                                        continue

                                    dca_values = []
                                    for candidate_name in candidate_names:
                                        values = next_assignments[
                                            candidate_name
                                        ]
                                        if not isinstance(values, list):
                                            values = [values]
                                        for value in values:
                                            if value not in dca_values:
                                                dca_values.append(value)

                                    candidate_dca = display_dca(dca_values)
                                    candidate_box = fitz.Rect(
                                        candidate_spans[0]["bbox"]
                                    )
                                    candidate_number_right = max(
                                        36,
                                        candidate_box.x0 - number_gap,
                                    )
                                    candidate_number_size = max(
                                        12,
                                        candidate_spans[0]["size"]
                                        * number_scale,
                                    )
                                    candidate_number_width = fitz.Font(
                                        fontname=number_font
                                    ).text_length(
                                        candidate_dca,
                                        fontsize=candidate_number_size,
                                    )
                                    candidate_annotation_width = max(
                                        56,
                                        candidate_number_width + 10,
                                    )
                                    predicted_number_box = fitz.Rect(
                                        max(
                                            8,
                                            candidate_number_right
                                            - candidate_annotation_width,
                                        ),
                                        candidate_box.y0
                                        + number_vertical_offset,
                                        candidate_number_right,
                                        candidate_box.y1
                                        + 4
                                        + number_vertical_offset,
                                    )
                                    avoid_following_speaker_number = bool(
                                        proposed_state_box.intersects(
                                            predicted_number_box
                                        )
                                        and proposed_state_box.x1 + 4
                                        <= cue_box.x0
                                    )
                                    break

                            if state_position == "Left Gutter":
                                # A label in the far-left gutter must not
                                # share a dialogue row. Long state names can
                                # otherwise run into a speaker name even when
                                # their x-position is safely in the gutter.
                                if avoid_following_speaker_number:
                                    state_y = cue_box.y1 - 2
                                else:
                                    state_y = (
                                        cue_box.y1 + state_font_size + 4
                                        if next_state["position"] == "after"
                                        else cue_box.y0 - 4
                                    )
                            elif (
                                state_placement == "Beside Cue"
                                and not state_shares_cue_line
                            ):
                                state_y = cue_box.y1 - 2
                            elif state_placement == "Below Cue":
                                state_y = cue_box.y1 + state_font_size + 4
                            else:
                                # The default keeps the DCA State label above
                                # the cue line, irrespective of whether its
                                # assignment begins Before or After that cue.
                                state_y = cue_box.y0 - 4

                            if editable:
                                # Use a FreeText annotation so the DCA State
                                # label can be changed, moved, or removed in a
                                # PDF editor after export.
                                state_box = fitz.Rect(
                                    state_x - 3,
                                    state_y - state_font_size - 4,
                                    state_x + state_width + 5,
                                    state_y + 4,
                                )
                                annotation = page.add_freetext_annot(
                                    state_box,
                                    next_state["name"],
                                    fontsize=state_font_size,
                                    fontname=draw_state_font_name,
                                    text_color=state_colour,
                                    fill_color=None,
                                    align=0,
                                )
                                annotation.set_border(width=0)
                                annotation.update()
                            else:
                                page.insert_text(
                                    (state_x, state_y),
                                    next_state["name"],
                                    fontsize=state_font_size,
                                    fontname=draw_state_font_name,
                                    fontfile=draw_state_font_file,
                                    color=state_colour,
                                )

                    if next_state and next_state["position"] == "before":
                        current_state = next_state["key"]
                        if page_start_state is None:
                            page_start_state = current_state
                        activated_states.add(current_state)
                        state_activation_pages.setdefault(
                            current_state, page_number
                        )
                        marked_speakers.clear()

                    if (
                        mapping_state
                        and page_is_selected
                        and not legend_only
                        and not is_italic(span)
                        and speaker_names
                        and is_visual_row_anchor
                        and all(
                            name in assignments.get(mapping_state, {})
                            for name in speaker_names
                        )
                        and looks_like_positioned_speaker_label(
                            speaker_line_text,
                            speaker_names,
                            speaker_row_x,
                            trusted_speaker_columns,
                            bold_prefix=has_bold_speaker_prefix(
                                line,
                                speaker_names,
                            ),
                            layout_speaker_names=(
                                line_layout_speaker_names
                            ),
                        )
                        and (
                            page_number,
                            round(line["bbox"][1], 1),
                            round(line["bbox"][0], 1),
                        ) not in marked_cue_lines
                        and (
                            not first_appearance
                            or any(
                                name not in marked_speakers
                                for name in speaker_names
                            )
                        )
                    ):
                        dca_values = []
                        for name in speaker_names:
                            values = assignments[mapping_state][name]
                            if not isinstance(values, list):
                                values = [values]
                            for value in values:
                                if value not in dca_values:
                                    dca_values.append(value)
                        dca = display_dca(dca_values)
                        name_box = fitz.Rect(span["bbox"])
                        name_box.x0 = visible_span_left_edges.get(
                            (tuple(span["bbox"]), span.get("text", "")),
                            name_box.x0,
                        )

                        number_right = max(36, name_box.x0 - number_gap)
                        number_font_size = max(
                            12, span["size"] * number_scale
                        )
                        number_width = fitz.Font(
                            fontname=number_font
                        ).text_length(dca, fontsize=number_font_size)
                        number_left = max(8, number_right - number_width)

                        if editable:
                            annotation_width = max(56, number_width + 10)

                            number_box = fitz.Rect(
                                max(8, number_right - annotation_width),
                                name_box.y0 + number_vertical_offset,
                                number_right,
                                name_box.y1 + 4 + number_vertical_offset,
                            )

                            annotation = page.add_freetext_annot(
                                number_box,
                                dca,
                                fontsize=number_font_size,
                                fontname=number_font,
                                text_color=number_colour,
                                fill_color=None,
                                align=2,
                            )

                        else:
                            page.insert_text(
                                (number_left, name_box.y1 - 2 + number_vertical_offset),
                                dca,
                                fontsize=number_font_size,
                                fontname=number_font,
                                color=number_colour,
                            )

                        marked_count += 1
                        marked_pages.add(page_number)
                        marked_page_counts[page_number] = (
                            marked_page_counts.get(page_number, 0) + 1
                        )
                        cue_identity = (
                            page_number,
                            mapping_state,
                            tuple(sorted(speaker_names, key=str.casefold)),
                            dca,
                        )
                        marked_cue_counts[cue_identity] = (
                            marked_cue_counts.get(cue_identity, 0) + 1
                        )
                        marked_speakers.update(speaker_names)
                        marked_cue_lines.add(
                            (
                                page_number,
                                round(line["bbox"][1], 1),
                                round(line["bbox"][0], 1),
                            )
                        )

                    elif (
                        page_is_selected
                        and current_state
                        and is_visual_row_anchor
                        and speaker_name != clean_text
                        and not starts_with_stage_direction(
                            speaker_line_text
                        )
                        and looks_like_positioned_speaker_label(
                            speaker_line_text,
                            speaker_names,
                            speaker_row_x,
                            trusted_speaker_columns,
                            bold_prefix=has_bold_speaker_prefix(
                                line,
                                speaker_names,
                            ),
                            layout_speaker_names=(
                                line_layout_speaker_names
                            ),
                        )
                        and (
                            page_number,
                            round(line["bbox"][1], 1),
                            round(line["bbox"][0], 1),
                        ) not in marked_cue_lines
                    ):
                        unmatched_names.append(
                            (page_number, current_state, text)
                        )

                    if next_state and next_state["position"] != "before":
                        current_state = next_state["key"]
                        if page_start_state is None:
                            page_start_state = current_state
                        activated_states.add(current_state)
                        state_activation_pages.setdefault(
                            current_state, page_number
                        )
                        marked_speakers.clear()

        if (
            page_is_selected
            and page_state_display != "off"
        ):
            header_state = page_start_state or current_state
            footer_state = current_state or header_state

            page_labels = []
            if page_state_display in {"header", "both"}:
                page_labels.append(
                    (
                        state_names.get(header_state, ""),
                        (36, 24),
                    )
                )

            if page_state_display in {"footer", "both"}:
                footer_label = state_names.get(footer_state, "")
                page_labels.append((footer_label, None))

            page_state_text_colour = state_style.get(
                "page_header_footer_text_colour",
                state_colour,
            )
            page_state_border_colour = state_style.get(
                "page_header_footer_border_colour",
                page_state_text_colour,
            )
            draw_state_font_name = state_style.get(
                "page_header_footer_font_name",
                state_font_name,
            )
            draw_state_font_file = state_style.get(
                "page_header_footer_font_file",
                state_font_file,
            )
            state_font_size = state_style.get(
                "page_header_footer_size",
                state_style.get("size", 12),
            )

            # Use the Chinese-capable font whenever a state label contains
            # Chinese text, even if the user chose an English font.
            state_text = " ".join(
                name for name, _ in page_labels
                if name
            )
            if contains_cjk(state_text):
                draw_state_font_name = "heiti"
                draw_state_font_file = CHINESE_FONT_FILE

            # Page header/footer labels are deliberately more prominent than
            # an in-script state-change label: bold text plus a thin box.
            if not draw_state_font_file:
                bold_font_map = {
                    "helv": "hebo",
                    "tiro": "tibo",
                    "cour": "cobo",
                }
                draw_state_font_name = bold_font_map.get(
                    draw_state_font_name,
                    "hebo",
                )

            if draw_state_font_file:
                footer_font = fitz.Font(fontfile=draw_state_font_file)
            else:
                footer_font = fitz.Font(fontname=draw_state_font_name)

            if page_state_display in {"footer", "both"}:
                footer_label = state_names.get(footer_state, "")
                footer_width = footer_font.text_length(
                    footer_label,
                    fontsize=state_font_size,
                )
                # Keep 90 points clear at the far right for the PDF's own page
                # number, while still placing the DCA State in the lower-right.
                footer_x = max(
                    36,
                    page.rect.width - 90 - footer_width,
                )
                page_labels[-1] = (
                    footer_label,
                    (footer_x, page.rect.height - 22),
                )

            for page_label, point in page_labels:
                if page_label:
                    label_width = footer_font.text_length(
                        page_label,
                        fontsize=state_font_size,
                    )
                    label_box = fitz.Rect(
                        point[0] - 4,
                        point[1] - state_font_size - 4,
                        point[0] + label_width + 4,
                        point[1] + 4,
                    )
                    if editable:
                        # Keep the text and border in one FreeText annotation
                        # so they move, resize, and delete together in a PDF
                        # editor.
                        if (
                            page_state_border_colour
                            == page_state_text_colour
                        ):
                            # Keep the long-standing plain FreeText output
                            # when both colours match. This preserves today's
                            # default appearance and older CLI behaviour.
                            annotation = page.add_freetext_annot(
                                label_box,
                                page_label,
                                fontsize=state_font_size,
                                fontname=draw_state_font_name,
                                text_color=page_state_text_colour,
                                fill_color=None,
                                align=1,
                            )
                            annotation.set_border(width=0.8)
                            annotation.update()
                        else:
                            # A plain PDF FreeText annotation exposes only one
                            # colour for its text and border. Rich text keeps
                            # the selected text colour in CSS while /DA holds
                            # the border colour, still as one movable object.
                            text_style = (
                                "font-family: "
                                f"{rich_text_font_family(draw_state_font_name, draw_state_font_file)}; "
                                f"font-size: {state_font_size:g}pt; "
                                "font-weight: bold; "
                                f"color: {css_colour(page_state_text_colour)}; "
                                "text-align: center; "
                                "margin: 0; padding: 0; line-height: 1;"
                            )
                            escaped_label = html.escape(page_label)
                            annotation = page.add_freetext_annot(
                                label_box,
                                escaped_label,
                                richtext=True,
                                style=text_style,
                                border_width=0.8,
                                fill_color=None,
                                align=1,
                            )
                            rich_content = (
                                '<?xml version="1.0"?>'
                                '<body xmlns="http://www.w3.org/1999/xhtml" '
                                'xmlns:xfa="http://www.xfa.org/schema/xfa-data/1.0/" '
                                'xfa:contentType="text/html" '
                                'xfa:APIVersion="Acrobat:8.0.0" '
                                'xfa:spec="2.4">'
                                f"{escaped_label}</body>"
                            )
                            # PyMuPDF 1.28 omits the closing body tag in its
                            # generated /RC value. Replace it with valid XHTML
                            # so cleaning or regenerating the PDF cannot drop
                            # the font glyphs.
                            document.xref_set_key(
                                annotation.xref,
                                "RC",
                                fitz.get_pdf_str(rich_content),
                            )
                            annotation.update(
                                text_color=page_state_border_colour,
                            )
                            # Retain a normal searchable /Contents value for
                            # Preview and other PDF editors. Set it only after
                            # generating the appearance: set_info() removes
                            # rich /RC data, while setting Contents before the
                            # update makes some renderers lose the glyphs.
                            document.xref_set_key(
                                annotation.xref,
                                "Contents",
                                fitz.get_pdf_str(page_label),
                            )
                    else:
                        page.draw_rect(
                            label_box,
                            color=page_state_border_colour,
                            width=0.8,
                            overlay=True,
                        )
                        page.insert_text(
                            point,
                            page_label,
                            fontsize=state_font_size,
                            fontname=draw_state_font_name,
                            fontfile=draw_state_font_file,
                            color=page_state_text_colour,
                        )
                        # Some Chinese system fonts do not expose a separate
                        # bold face to PDF output. A subtle second pass gives
                        # the page header/footer a reliably stronger weight.
                        page.insert_text(
                            (point[0] + 0.35, point[1]),
                            page_label,
                            fontsize=state_font_size,
                            fontname=draw_state_font_name,
                            fontfile=draw_state_font_file,
                            color=page_state_text_colour,
                        )

        if page_is_selected and show_performer_role_mapping:
            mapping_states_for_page = []
            first_selected_page = start_page or 1

            # A partial export may begin after a state was activated. Carry
            # that active state's reference card onto the first selected page
            # so the excerpt remains understandable without the workbook.
            if page_number == first_selected_page and page_start_state:
                mapping_states_for_page.append(page_start_state)

            mapping_states_for_page.extend(
                state_key
                for state_key, activation_page
                in state_activation_pages.items()
                if activation_page == page_number
            )

            for state_key in dict.fromkeys(mapping_states_for_page):
                if state_key in performer_role_mapping_states:
                    continue

                state = states_by_key.get(state_key)
                if not state or not state.get("performer_role_rows"):
                    continue

                annotation = add_performer_role_mapping_card(
                    document,
                    page,
                    state,
                    state_style,
                )
                if annotation is None:
                    continue

                performer_role_mapping_states.add(state_key)
                performer_role_mapping_pages[state_key] = page_number

    # Rebuild and compress the finished PDF, then atomically place the complete
    # file at its destination. This keeps the previous output intact if saving
    # fails and gives PDF viewers one replacement event instead of a delete
    # followed by a second file appearing at the same path.
    save_document_atomically(document, output_file)

    if diagnostics is not None:
        diagnostics.update({
            "pdf_page_count": pdf_page_count,
            "selected_start_page": start_page or 1,
            "selected_end_page": min(
                end_page or pdf_page_count,
                pdf_page_count,
            ),
            "full_document": start_page is None and end_page is None,
            "state_activation_pages": state_activation_pages,
            "performer_role_mapping_pages": (
                performer_role_mapping_pages
            ),
            "marked_pages": sorted(marked_pages),
            "marked_page_counts": {
                str(page): marked_page_counts[page]
                for page in sorted(marked_page_counts)
            },
            "marked_cue_counts": [
                {
                    "page": page,
                    "state": state,
                    "speakers": list(speakers),
                    "dca": dca,
                    "count": marked_cue_counts[
                        (page, state, speakers, dca)
                    ],
                }
                for page, state, speakers, dca in sorted(
                    marked_cue_counts,
                    key=lambda item: (
                        item[0],
                        item[1],
                        tuple(name.casefold() for name in item[2]),
                        item[3],
                    ),
                )
            ],
            "known_speakers_without_active_assignment": (
                unassigned_known_speakers
            ),
        })

    return marked_count, unmatched_names, activated_states


def build_review_notices(
    states,
    assignments,
    marked_count,
    activated_states,
    diagnostics=None,
    legend_only=False,
):
    """Return high-confidence safety notices without guessing from density."""
    diagnostics = diagnostics or {}
    notices = []
    full_document = diagnostics.get("full_document", True)

    def add_notice(code, severity, message):
        notices.append({
            "code": code,
            "severity": severity,
            "message": message,
        })

    if not states:
        add_notice(
            "NO_STATES_CONFIGURED",
            "critical",
            "No usable DCA States were found in the template. Add a state "
            "name and Start Line Text before using this PDF.",
        )
    elif not activated_states:
        add_notice(
            "NO_STATES_ACTIVATED",
            "critical" if full_document else "warning",
            "No DCA State start cue was found in the selected script pages, "
            "so no state-based marking could be verified. Confirm that the "
            "PDF text is selectable and check the exact Start Line Text, "
            "Start Line Character, and Page Hint.",
        )
    else:
        first_state = states[0]
        if first_state["key"] not in activated_states:
            marked_pages = diagnostics.get("marked_pages", [])
            page_detail = (
                f" The first DCA number appears on PDF page "
                f"{marked_pages[0]}."
                if marked_pages else ""
            )
            add_notice(
                "FIRST_STATE_NOT_ACTIVATED",
                "warning",
                f'The first configured state, "{first_state["name"]}", '
                f"was not activated.{page_detail} Early pages may be unmarked. "
                "Check that Page Hint uses the printed script page number "
                "when one is printed, and otherwise the sequential PDF page "
                "position.",
            )

        if full_document:
            missing_states = [
                state
                for state in states
                if state["key"] not in activated_states
                and state["key"] != first_state["key"]
            ]
            if missing_states:
                add_notice(
                    "MISSING_STATE_CUES",
                    "warning",
                    f"{len(missing_states)} later DCA State start cue(s) "
                    "were not found. Check the state list in this report.",
                )

    if (
        states
        and activated_states
        and marked_count == 0
        and not legend_only
    ):
        add_notice(
            "ZERO_CUES_MARKED",
            "critical" if full_document else "warning",
            "A DCA State was activated, but no dialogue DCA numbers were "
            "placed. The speaker-label layout may not be recognised, or "
            "script names and aliases, or Performer / Role Mappings, may not "
            "match the workbook's DCA assignments. Confirm that the PDF "
            "text is selectable before use.",
        )

    cues_without_names = diagnostics.get(
        "cues_without_state_names", []
    )
    if cues_without_names:
        add_notice(
            "START_CUES_WITHOUT_STATE_NAMES",
            "warning",
            f"{len(cues_without_names)} template row(s) have Start Line Text "
            "but no DCA State name and were not used.",
        )

    assignment_states_without_cues = diagnostics.get(
        "assignment_states_without_start_cues", []
    )
    if assignment_states_without_cues:
        add_notice(
            "ASSIGNMENTS_WITHOUT_START_CUES",
            "warning",
            f"DCA assignments exist for {len(assignment_states_without_cues)} "
            "state row(s) that have no usable start cue.",
        )

    assignment_gaps = diagnostics.get("assignment_gaps", [])
    if assignment_gaps:
        examples = []
        for gap in assignment_gaps[:3]:
            missing = ", ".join(
                f"DCA {dca}" for dca in gap.get("missing_dcas", [])
            )
            examples.append(
                f'{gap.get("state_name", "Unknown state")}: {missing} '
                f'is blank between DCA {gap.get("first_dca", "?")} and '
                f'DCA {gap.get("last_dca", "?")}'
            )

        remaining = len(assignment_gaps) - len(examples)
        remaining_text = (
            f"; and {remaining} more state(s)"
            if remaining > 0 else ""
        )
        add_notice(
            "DCA_ASSIGNMENT_GAPS",
            "warning",
            f"{len(assignment_gaps)} DCA State row(s) contain an empty DCA "
            "column between populated DCA columns. This may be intentional, "
            "but it can also mean an assignment was accidentally missed. "
            f'Confirm: {"; ".join(examples)}{remaining_text}.'
        )

    duplicate_dca_assignments = diagnostics.get(
        "duplicate_dca_assignments", []
    )
    if duplicate_dca_assignments:
        examples = []
        for duplicate in duplicate_dca_assignments[:4]:
            dcas = "/".join(
                f'DCA {dca}' for dca in duplicate.get("dcas", [])
            )
            examples.append(
                f'{duplicate.get("state_name", "Unknown state")}: '
                f'{duplicate.get("dca_name", "Unknown DCA Name")} '
                f'appears in {dcas}'
            )
        remaining = len(duplicate_dca_assignments) - len(examples)
        remaining_text = (
            f"; and {remaining} more assignment(s)"
            if remaining > 0 else ""
        )
        add_notice(
            "DUPLICATE_DCA_ASSIGNMENTS",
            "warning",
            f"{len(duplicate_dca_assignments)} DCA Name assignment(s) "
            "appear in more than one DCA column in the same state. This may "
            "be intentional, so generation continued. Confirm before use: "
            f'{"; ".join(examples)}{remaining_text}.'
        )

    # A repeated cue can legitimately appear on another page while the state
    # still activates at its configured Page Hint. Report only conflicts for
    # states that never activated, which is the actionable setup failure.
    page_hint_mismatches = [
        mismatch
        for mismatch in diagnostics.get("page_hint_mismatches", [])
        if mismatch.get("state_key") not in activated_states
    ]
    if page_hint_mismatches:
        examples = []
        for mismatch in page_hint_mismatches[:3]:
            state_name = (
                mismatch.get("state_name")
                or mismatch.get("state_key")
                or "Unknown state"
            )
            pdf_page = mismatch.get("pdf_page")
            page_found = (
                f"PDF page {pdf_page}"
                if pdf_page is not None
                else "another PDF page"
            )
            examples.append(
                f'{state_name} (found on {page_found}; '
                f'Page Hint {mismatch.get("page_hint", "")})'
            )
        add_notice(
            "PAGE_HINT_MISMATCH",
            "warning",
            "Exact Start Cue Text was found at a different sequential PDF "
            f"page position for {len(page_hint_mismatches)} DCA State(s). "
            "Page Hint normally uses the number printed inside the script; "
            "use the PDF page position only when no printed number exists. "
            "Correct the Page Hint, or leave it blank only when the cue text "
            "is unique. Examples: "
            f'{"; ".join(examples)}.',
        )

    unassigned_labels = diagnostics.get(
        "known_speakers_without_active_assignment", []
    )
    state_display_names = {
        state["key"]: state["name"] for state in states
    }

    def unassigned_example(item):
        state_name = state_display_names.get(
            item.get("state", ""),
            item.get("state", "Unknown state"),
        )
        speakers = "/".join(item.get("speakers", []))
        return (
            f'PDF page {item.get("page", "?")} | '
            f"{state_name} | {speakers}"
        )

    def unassigned_examples(items, limit=3):
        examples = []
        seen = set()
        for item in items:
            key = (
                item.get("page"),
                item.get("state"),
                tuple(item.get("speakers", [])),
            )
            if key in seen:
                continue
            seen.add(key)
            examples.append(unassigned_example(item))
            if len(examples) >= limit:
                break
        return "; ".join(examples)

    final_state_key = states[-1]["key"] if states else ""
    final_state_page = diagnostics.get(
        "state_activation_pages", {}
    ).get(final_state_key)
    final_state_labels = [
        item
        for item in unassigned_labels
        if final_state_page is not None
        and item.get("state") == final_state_key
        and item.get("page", 0) >= final_state_page
    ]

    if final_state_labels:
        examples = unassigned_examples(final_state_labels)
        add_notice(
            "POSSIBLE_INCOMPLETE_FINAL_STATE",
            "warning",
            "Known speaker labels without an assignment were found after "
            "the final configured state began. Confirm that the workbook "
            f"includes every later state. Examples: {examples}.",
        )

    other_unassigned_labels = [
        item for item in unassigned_labels
        if item not in final_state_labels
    ]
    if other_unassigned_labels:
        examples = unassigned_examples(other_unassigned_labels)
        add_notice(
            "KNOWN_SPEAKERS_UNASSIGNED",
            "warning",
            f"{len(other_unassigned_labels)} positioned speaker label(s) "
            "were recognised but have no DCA assignment in the active "
            f"state. Examples: {examples}.",
        )

    return notices


def safety_level_for_notices(notices):
    if any(notice["severity"] == "critical" for notice in notices):
        return "critical"
    if notices:
        return "warning"
    return "ok"


def write_review_report(
    states,
    marked_count,
    unmatched_names,
    activated_states,
    report_file,
    notices=None,
    diagnostics=None,
):
    notices = notices or []
    diagnostics = diagnostics or {}

    with open(report_file, "w", encoding="utf-8") as file:
        file.write("DCA Script Marker - Review Report\n")
        file.write("=" * 40 + "\n\n")
        file.write("Automatic safety check\n")
        file.write("-" * 40 + "\n")

        if notices:
            file.write("Status: REVIEW REQUIRED\n\n")
            for notice in notices:
                file.write(
                    f'- [{notice["severity"].upper()}] '
                    f'{notice["message"]}\n'
                )
        else:
            file.write("Status: No high-risk automatic warning found.\n")

        file.write("\n")
        file.write(
            f"DCA States found: {len(activated_states)} of {len(states)}\n"
        )
        marked_pages = diagnostics.get("marked_pages", [])
        if marked_pages:
            file.write(
                "PDF pages receiving DCA numbers: "
                f"{marked_pages[0]} to {marked_pages[-1]}\n"
            )
        else:
            file.write("PDF pages receiving DCA numbers: none\n")
        file.write(
            "Automatic checks cannot guarantee every script layout. "
            "Human review is always required.\n\n"
        )
        file.write(f"Marked character cues: {marked_count}\n\n")

        known_unassigned = diagnostics.get(
            "known_speakers_without_active_assignment", []
        )
        if known_unassigned:
            state_names = {
                state["key"]: state["name"] for state in states
            }
            grouped_unassigned = {}
            for item in known_unassigned:
                group_key = (
                    item.get("state", ""),
                    tuple(item.get("speakers", [])),
                )
                group = grouped_unassigned.setdefault(group_key, {
                    "count": 0,
                    "pages": [],
                    "label": item.get("label", ""),
                })
                group["count"] += 1
                page = item.get("page")
                if page not in group["pages"]:
                    group["pages"].append(page)

            file.write(
                "Recognised speaker labels without an active DCA "
                "assignment:\n\n"
            )
            for (state_key, speakers), group in grouped_unassigned.items():
                state_name = state_names.get(
                    state_key,
                    state_key or "Unknown state",
                )
                speaker_names = "/".join(speakers)
                pages = ", ".join(
                    str(page) for page in group["pages"][:8]
                )
                if len(group["pages"]) > 8:
                    pages += f" and {len(group['pages']) - 8} more"
                file.write(
                    f"- {state_name} | {speaker_names} | PDF page(s) "
                    f"{pages} | {group['count']} label(s) | "
                    f"Example: {group['label']}\n"
                )
            file.write("\n")

        missing_states = [
            state["name"]
            for state in states
            if state["key"] not in activated_states
        ]

        if missing_states:
            file.write("DCA States whose start cue was not found:\n\n")

            for state_name in missing_states:
                file.write(f"- {state_name}\n")

            file.write("\n")

        if unmatched_names:
            file.write("Possible character names without a DCA assignment:\n\n")

            for page, state, name in unmatched_names:
                file.write(f"- Page {page} | {state} | {name}\n")
        else:
            file.write("No possible unmatched character names found.\n")


def write_result_json(result_json_file, result):
    """Atomically write the optional machine-readable completion summary."""
    result_directory = os.path.dirname(
        os.path.abspath(result_json_file)
    )
    temporary_file = None

    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            dir=result_directory,
            prefix=".dca-marker-result-",
            suffix=".json",
            delete=False,
        ) as file:
            temporary_file = file.name
            json.dump(result, file, ensure_ascii=False, indent=2)
            file.write("\n")
        os.replace(temporary_file, result_json_file)
        temporary_file = None
    finally:
        if temporary_file and os.path.exists(temporary_file):
            os.remove(temporary_file)


def run_marker(
    template_file,
    pdf_file,
    output_folder,
    editable=False,
    first_appearance=False,
    legend_only=False,
    legend_overrides=None,
    legend_style=None,
    number_style=None,
    state_style=None,
    start_page=None,
    end_page=None,
    ocr_json_file=None,
    output_mode="replace",
    result_json_file=None,
    result_data=None,
    project_file=None,
):
    original_name = os.path.splitext(
        os.path.basename(pdf_file)
    )[0]

    today = date.today().isoformat()

    output_file = os.path.join(
        output_folder,
        f"{original_name}_marked_{today}.pdf",
    )

    report_file = os.path.join(
        output_folder,
        f"{original_name}_review_{today}.txt",
    )

    if output_mode == "new":
        pdf_root, pdf_extension = os.path.splitext(output_file)
        report_root, report_extension = os.path.splitext(report_file)

        version = 2
        while os.path.exists(output_file):
            output_file = f"{pdf_root}_{version}{pdf_extension}"
            report_file = f"{report_root}_{version}{report_extension}"
            version += 1

    # Some Excel templates contain a data-validation extension that openpyxl
    # does not preserve when saving. This app only reads the workbook, so the
    # warning is harmless and should not obscure the user's completion message.
    diagnostics = {}

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported.*",
            category=UserWarning,
        )
        if project_file:
            states, assignments = load_project(
                project_file,
                diagnostics=diagnostics,
            )
        else:
            states, assignments = load_template(
                template_file,
                diagnostics=diagnostics,
            )
    marked_count, unmatched_names, activated_states = mark_pdf(
        states,
        assignments,
        pdf_file,
        output_file,
        editable,
        first_appearance,
        legend_only,
        legend_overrides,
        legend_style,
        number_style,
        state_style,
        start_page,
        end_page,
        ocr_json_file,
        diagnostics,
    )
    notices = build_review_notices(
        states,
        assignments,
        marked_count,
        activated_states,
        diagnostics=diagnostics,
        legend_only=legend_only,
    )
    write_review_report(
        states,
        marked_count,
        unmatched_names,
        activated_states,
        report_file,
        notices=notices,
        diagnostics=diagnostics,
    )

    missing_state_count = sum(
        state["key"] not in activated_states
        for state in states
    )
    activated_state_keys = [
        state["key"]
        for state in states
        if state["key"] in activated_states
    ]
    missing_state_keys = [
        state["key"]
        for state in states
        if state["key"] not in activated_states
    ]
    completion_result = {
        "schema_version": 1,
        "marked_count": marked_count,
        "output_pdf": output_file,
        "review_report": report_file,
        "safety_level": safety_level_for_notices(notices),
        "safety_warning_count": len(notices),
        "safety_warnings": notices,
        "configured_state_count": len(states),
        "activated_state_count": len(activated_states),
        "missing_state_count": missing_state_count,
        "unmatched_name_count": len(unmatched_names),
        "activated_states": activated_state_keys,
        "missing_states": missing_state_keys,
        "state_activation_pages": {
            state_key: diagnostics.get(
                "state_activation_pages", {}
            )[state_key]
            for state_key in activated_state_keys
            if state_key in diagnostics.get(
                "state_activation_pages", {}
            )
        },
        "performer_role_mapping_pages": diagnostics.get(
            "performer_role_mapping_pages", {}
        ),
        "marked_pages": diagnostics.get("marked_pages", []),
        "marked_page_counts": diagnostics.get(
            "marked_page_counts", {}
        ),
        "marked_cue_counts": diagnostics.get(
            "marked_cue_counts", []
        ),
        "pdf_page_count": diagnostics.get("pdf_page_count", 0),
    }

    if result_data is not None:
        result_data.clear()
        result_data.update(completion_result)

    if result_json_file:
        write_result_json(result_json_file, completion_result)

    return marked_count, output_file, report_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Create a DCA-marked rehearsal script.",
        allow_abbrev=False,
    )
    parser.add_argument(
        "--self-test",
        action="store_true",
        help="Report the bundled runtime versions and architecture, then exit",
    )
    parser.add_argument("--template", help="Path to the DCA Excel template")
    parser.add_argument(
        "--project",
        help="Path to a Version 2 DCA Script Marker project (.dcamarker)",
    )
    parser.add_argument(
        "--import-excel",
        action="store_true",
        help="Convert --template into Version 2 project JSON and exit",
    )
    parser.add_argument(
        "--export-excel",
        help="Export --project to the specified Excel workbook and exit",
    )
    parser.add_argument("--script", help="Path to the script PDF")
    parser.add_argument("--output", help="Folder for the marked PDF")
    parser.add_argument(
        "--result-json-file",
        help=(
            "Optional path for a machine-readable completion and safety "
            "summary"
        ),
    )
    parser.add_argument(
        "--output-mode",
        choices=["replace", "new"],
        default="replace",
        help="Replace an existing marked PDF or create a new version",
    )
    parser.add_argument(
        "--ocr-json",
        help="Vision OCR data produced by the Mac app for a scanned PDF",
    )
    parser.add_argument(
        "--list-legends",
        action="store_true",
        help="Print DCA State legend text as JSON and exit",
    )
    parser.add_argument(
        "--list-role-mappings",
        action="store_true",
        help=(
            "Print every active DCA Name in each state, including optional "
            "Other Script Characters Played, as JSON and exit"
        ),
    )
    parser.add_argument(
        "--legend-overrides-file",
        help="Path to a JSON file containing edited legend text",
    )
    parser.add_argument(
        "--number-colour",
        choices=ANNOTATION_COLOUR_CHOICES,
        help="Colour for DCA numbers",
    )
    parser.add_argument("--number-scale", type=float, help="DCA number size scale")
    parser.add_argument("--number-font", help="DCA number font: Helvetica, Times, or Courier")
    parser.add_argument("--number-x", type=float, help="Legacy DCA number position")
    parser.add_argument(
        "--number-gap",
        type=float,
        help="Space between a DCA number and the speaker name",
    )
    parser.add_argument(
        "--number-y-offset",
        type=float,
        default=0,
        help="Move DCA numbers vertically in points; negative is up",
    )
    parser.add_argument(
        "--state-colour",
        choices=ANNOTATION_COLOUR_CHOICES,
        help="Colour for DCA State labels",
    )
    parser.add_argument("--state-scale", type=float, help="DCA state size scale")
    parser.add_argument("--state-font", help="DCA state font: Helvetica, Times, or Courier")
    parser.add_argument(
        "--state-position",
        choices=["Left Gutter", "Far from Script", "Near Script"],
        default="Left Gutter",
        help="Position for the DCA State label",
    )
    parser.add_argument(
        "--state-placement",
        choices=["Above Cue", "Beside Cue", "Below Cue"],
        default="Beside Cue",
        help="Legacy setting; DCA State labels default beside their cue",
    )
    parser.add_argument(
        "--page-state-header-footer",
        action="store_true",
        help=(
            "Legacy option: show the active DCA State at both the top and "
            "bottom of each page"
        ),
    )
    parser.add_argument(
        "--page-state-display",
        choices=["off", "header", "footer", "both"],
        help=(
            "Choose whether the active DCA State appears in the page "
            "header, footer, both, or neither"
        ),
    )
    parser.add_argument(
        "--page-state-text-colour",
        choices=ANNOTATION_COLOUR_CHOICES,
        help="Text colour for page header/footer DCA State labels",
    )
    parser.add_argument(
        "--page-state-scale",
        type=float,
        help="Text size scale for page header/footer DCA State labels",
    )
    parser.add_argument(
        "--page-state-font",
        choices=[
            "PingFang SC",
            "Chinese System",
            "Helvetica",
            "Times",
            "Courier",
        ],
        help="Text font for page header/footer DCA State labels",
    )
    parser.add_argument(
        "--page-state-border-colour",
        choices=ANNOTATION_COLOUR_CHOICES,
        help="Border colour for page header/footer DCA State labels",
    )
    parser.add_argument(
        "--show-performer-role-mapping",
        action="store_true",
        help=(
            "Show a movable Character List performer/role mapping on the "
            "first selected page where each DCA State is active"
        ),
    )
    parser.add_argument(
        "--legend-position",
        choices=["Left Gutter", "Near Script"],
        default="Left Gutter",
        help="Position for the DCA State Legend",
    )
    parser.add_argument(
        "--start-page",
        type=int,
        help="First sequential PDF file page to mark (cover is page 1)",
    )
    parser.add_argument(
        "--end-page",
        type=int,
        help="Last sequential PDF file page to mark (optional)",
    )
    parser.add_argument(
        "--style",
        default="Full Marking",
        choices=[
            "Full Marking",
            "Editable Full Marking",
            "First Appearance Only",
            "DCA State Legend",
        ],
        help="Marking style to use",
    )
    arguments = parser.parse_args()

    if arguments.import_excel:
        if not arguments.template:
            parser.error("--template is required with --import-excel")
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            project = import_excel_project(arguments.template)
        print(json.dumps(project, ensure_ascii=False, indent=2))
        raise SystemExit(0)

    if arguments.export_excel:
        if not arguments.project:
            parser.error("--project is required with --export-excel")
        export_project_excel(arguments.project, arguments.export_excel)
        print(f"Excel workbook exported: {arguments.export_excel}")
        raise SystemExit(0)

    if arguments.self_test:
        print(json.dumps({
            "ok": True,
            "architecture": platform.machine(),
            "python": platform.python_version(),
            "pymupdf": getattr(fitz, "VersionBind", "unknown"),
            "openpyxl": openpyxl.__version__,
            "frozen": bool(getattr(sys, "frozen", False)),
            "executable": os.path.realpath(sys.executable),
            "module_paths": {
                "pymupdf": os.path.realpath(fitz.__file__),
                "openpyxl": os.path.realpath(openpyxl.__file__),
            },
        }, ensure_ascii=False))
        raise SystemExit(0)

    if arguments.list_legends:
        if not arguments.template and not arguments.project:
            parser.error(
                "--template or --project is required with --list-legends"
            )
        # Excel may report an unsupported validation-extension warning. The
        # Mac app expects clean JSON here, so do not send that warning to it.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            if arguments.project:
                states, assignments = load_project(arguments.project)
            else:
                states, assignments = load_template(arguments.template)
        legends = [
            {
                "key": state["key"],
                "name": state["name"],
                "text": build_legend_text(state, assignments),
            }
            for state in states
        ]
        print(json.dumps(legends, ensure_ascii=False))
        raise SystemExit(0)

    if arguments.list_role_mappings:
        if not arguments.template and not arguments.project:
            parser.error(
                "--template or --project is required with "
                "--list-role-mappings"
            )
        # The floating Mac inspector needs only workbook data. Keep this
        # command read-only and independent from a script PDF or output folder.
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            if arguments.project:
                states, _assignments = load_project(arguments.project)
            else:
                states, _assignments = load_template(arguments.template)
        role_mappings = [
            {
                "id": f'{state["key"]}::{index}',
                "key": state["key"],
                "name": state["name"],
                "page_hint": state.get("page_hint", ""),
                "rows": [
                    {
                        "dca": display_dca(row["dca"]),
                        "performer": row["performer"],
                        "roles": row["roles"],
                    }
                    for row in state.get("dca_reference_rows", [])
                ],
            }
            for index, state in enumerate(states)
        ]
        print(json.dumps(role_mappings, ensure_ascii=False))
        raise SystemExit(0)

    completion_result = {}

    # With no command-line options, keep the original simple test behaviour.
    if not arguments.template and not arguments.project:
        marked_count, output_file, report_file = run_marker(
            TEMPLATE_FILE,
            PDF_FILE,
            os.getcwd(),
            result_data=completion_result,
        )
    else:
        if not arguments.script or not arguments.output:
            parser.error(
                "--template/--project, --script, and --output must be "
                "used together"
            )
        if arguments.start_page and arguments.start_page < 1:
            parser.error("--start-page must be 1 or greater")
        if arguments.end_page and arguments.end_page < 1:
            parser.error("--end-page must be 1 or greater")
        if (
            arguments.start_page
            and arguments.end_page
            and arguments.start_page > arguments.end_page
        ):
            parser.error("--start-page cannot be after --end-page")

        colour_map = ANNOTATION_COLOURS
        font_map = {
            "Helvetica": "helv",
            "Times": "tiro",
            "Courier": "cour",
        }
        number_style = {}
        state_style = {}

        if arguments.number_colour:
            number_style["colour"] = colour_map[arguments.number_colour]
        if arguments.number_scale:
            number_style["scale"] = arguments.number_scale
        if arguments.number_font:
            number_style["font_name"] = font_map[arguments.number_font]
        if arguments.number_gap is not None:
            number_style["gap"] = arguments.number_gap
        if arguments.number_y_offset is not None:
            number_style["vertical_offset"] = arguments.number_y_offset
        elif arguments.number_x is not None:
            # Older Mac UI versions send 72/60/36 for near/standard/far.
            number_style["gap"] = max(8, 76 - arguments.number_x)
        if arguments.state_colour:
            state_style["colour"] = colour_map[arguments.state_colour]
        if arguments.state_scale:
            state_style["size"] = 12 * arguments.state_scale
        # Keep the existing Chinese-capable font for DCA State labels. The
        # selected state font is used only when it can safely render the text.
        if arguments.state_font == "PingFang SC":
            state_style["font_name"] = "heiti"
            state_style["font_file"] = CHINESE_FONT_FILE
        elif arguments.state_font and arguments.state_font != "Chinese System":
            state_style["font_name"] = font_map[arguments.state_font]
            state_style["font_file"] = None
            state_style["font_name"] = font_map[arguments.state_font]
            state_style["font_file"] = None
        state_style["position"] = arguments.state_position
        state_style["placement"] = arguments.state_placement
        page_state_display = arguments.page_state_display
        if page_state_display is None:
            page_state_display = (
                "both"
                if arguments.page_state_header_footer
                else "off"
            )
        state_style["page_state_display"] = page_state_display
        state_style["page_header_footer"] = page_state_display != "off"
        if arguments.page_state_text_colour:
            state_style["page_header_footer_text_colour"] = colour_map[
                arguments.page_state_text_colour
            ]
        if arguments.page_state_scale:
            state_style["page_header_footer_size"] = (
                12 * arguments.page_state_scale
            )
        if arguments.page_state_font in {
            "PingFang SC",
            "Chinese System",
        }:
            state_style["page_header_footer_font_name"] = "heiti"
            state_style["page_header_footer_font_file"] = (
                CHINESE_FONT_FILE
            )
        elif arguments.page_state_font:
            state_style["page_header_footer_font_name"] = font_map[
                arguments.page_state_font
            ]
            state_style["page_header_footer_font_file"] = None
        if arguments.page_state_border_colour:
            state_style["page_header_footer_border_colour"] = colour_map[
                arguments.page_state_border_colour
            ]
        state_style["show_performer_role_mapping"] = (
            arguments.show_performer_role_mapping
        )

        legend_overrides = None
        if arguments.legend_overrides_file:
            with open(arguments.legend_overrides_file, encoding="utf-8") as file:
                legend_overrides = json.load(file)

        legend_style = {
            "colour": state_style.get("colour", STATE_COLOUR),
            "size": 8 * (arguments.state_scale or 1.2),
            "font_name": state_style.get("font_name", "heiti"),
            "font_file": state_style.get("font_file", CHINESE_FONT_FILE),
            "position": arguments.legend_position,
        }

        marked_count, output_file, report_file = run_marker(
            arguments.template,
            arguments.script,
            arguments.output,
            editable=arguments.style in {
                "Editable Full Marking",
                "First Appearance Only",
                "DCA State Legend",
            },
            first_appearance=arguments.style == "First Appearance Only",
            legend_only=arguments.style == "DCA State Legend",
            legend_overrides=legend_overrides,
            legend_style=legend_style,
            number_style=number_style,
            state_style=state_style,
            start_page=arguments.start_page,
            end_page=arguments.end_page,
            ocr_json_file=arguments.ocr_json,
            output_mode=arguments.output_mode,
            result_json_file=arguments.result_json_file,
            result_data=completion_result,
            project_file=arguments.project,
        )

    print(f"Finished! Marked {marked_count} cues.")
    print(f"PDF: {output_file}")
    print(f"Review report: {report_file}")
    warning_count = completion_result.get("safety_warning_count", 0)
    if warning_count:
        print(
            "Safety check: REVIEW REQUIRED - "
            f"{warning_count} automatic warning(s)."
        )
    else:
        print(
            "Safety check: No high-risk automatic warning found. "
            "Manual review is still required."
        )
