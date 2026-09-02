import importlib.util
import pathlib
import unittest
import warnings
import zipfile
from xml.etree import ElementTree

from openpyxl import load_workbook


ROOT = pathlib.Path(__file__).resolve().parents[1]
MARKER_FILE = ROOT / "dca_script_marker.py"
TEMPLATE_FILE = ROOT / "DCA Script Marker — DCA State Template.xlsx"

SPEC = importlib.util.spec_from_file_location("dca_script_marker", MARKER_FILE)
MARKER = importlib.util.module_from_spec(SPEC)
SPEC.loader.exec_module(MARKER)


class CanonicalTemplateTests(unittest.TestCase):
    def test_blank_template_loads_without_placeholder_states(self):
        diagnostics = {}
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            states, assignments = MARKER.load_template(
                TEMPLATE_FILE,
                diagnostics=diagnostics,
            )

        self.assertEqual(states, [])
        self.assertEqual(assignments, {})
        self.assertEqual(diagnostics["role_mapping_members"], 0)
        self.assertEqual(diagnostics["role_mapping_roles"], 0)

    def test_beta_template_keeps_the_supported_horizontal_schema(self):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(
                TEMPLATE_FILE,
                read_only=True,
                data_only=True,
            )
        self.addCleanup(workbook.close)

        self.assertEqual(
            workbook.sheetnames,
            [
                "How to use",
                "Character List",
                "DCA States",
            ],
        )
        headers = [
            workbook["DCA States"].cell(row=4, column=column).value
            for column in range(1, 19)
        ]
        self.assertEqual(
            headers,
            [
                "DCA State",
                "Start Line Character",
                "Start Line Text",
                "State Start Position",
                "Page Hint",
                *[f"DCA {number}" for number in range(1, 13)],
                "Notes",
            ],
        )
        page_hint_note = workbook["How to use"]["C19"].value
        states_reminder = workbook["DCA States"]["A2"].value
        self.assertIn("page number printed inside the script", page_hint_note)
        self.assertIn("sequential PDF page position", page_hint_note)
        self.assertIn("指定页码范围", page_hint_note)
        self.assertIn("Selected-page ranges", states_reminder)
        self.assertIn("剧本内印刷页码", states_reminder)
        self.assertIn(
            "performer / role mapping",
            states_reminder.lower(),
        )

        character_headers = [
            workbook["Character List"].cell(row=2, column=column).value
            for column in range(1, 4)
        ]
        self.assertEqual(
            [header.strip() for header in character_headers],
            [
                "DCA Name",
                "Other Script Characters Played",
                "Notes",
            ],
        )
        for character_column, state_cell in (("A", "A5"), ("B", "F5")):
            expected_fill = workbook["DCA States"][state_cell].fill
            expected_signature = (
                expected_fill.patternType,
                expected_fill.fgColor.type,
                expected_fill.fgColor.rgb,
            )
            for row in (3, 168):
                actual_fill = workbook["Character List"][f"{character_column}{row}"].fill
                self.assertEqual(
                    (
                        actual_fill.patternType,
                        actual_fill.fgColor.type,
                        actual_fill.fgColor.rgb,
                    ),
                    expected_signature,
                )
        dca_name_help = workbook["How to use"]["B10"].value
        mapping_help = workbook["How to use"]["B11"].value
        mapping_safety = workbook["How to use"]["C11"].value
        alias_help = workbook["How to use"]["B20"].value
        special_dca_example = workbook["How to use"]["C20"].value
        self.assertIn("ALL THREE is simply another DCA Name", dca_name_help)
        self.assertIn("TOM", dca_name_help)
        self.assertIn("JERRY", dca_name_help)
        self.assertIn("APPLE", dca_name_help)
        self.assertIn("Optional", mapping_help)
        self.assertIn("Leave this column blank", mapping_help)
        self.assertIn("Ben [B., Young Ben]", alias_help)
        self.assertIn("MALE ENSEMBLE", alias_help)
        self.assertIn("just like any other DCA Name", alias_help)
        self.assertIn("One DCA Name can play several roles", mapping_safety)
        self.assertIn("Jack", mapping_safety)
        self.assertIn("Student", mapping_safety)
        self.assertIn("Teacher", mapping_safety)
        self.assertIn(
            "one DCA Name in multiple DCA columns",
            special_dca_example,
        )
        self.assertIn("DCA 1 = TOM", special_dca_example)
        self.assertIn("DCA 2 = JERRY", special_dca_example)
        self.assertIn("DCA 3 = APPLE", special_dca_example)
        self.assertIn("ALL THREE", special_dca_example)
        self.assertIn("1/2/3", special_dca_example)
        self.assertIn("duplicate-assignment reminder", special_dca_example)


    def test_character_dropdowns_cover_all_dca_assignment_cells(self):
        workbook = load_workbook(TEMPLATE_FILE, data_only=False)
        validations = list(
            workbook["DCA States"].data_validations.dataValidation
        )
        if validations:
            self.assertEqual(len(validations), 1)
            self.assertEqual(validations[0].type, "list")
            validation_formula = validations[0].formula1
            validation_range = str(validations[0].sqref)
        else:
            # Current Excel versions may save list validation in the x14
            # extension namespace, which openpyxl deliberately does not load.
            with zipfile.ZipFile(TEMPLATE_FILE) as archive:
                states_xml = ElementTree.fromstring(
                    archive.read("xl/worksheets/sheet3.xml")
                )
            x14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
            xm = "http://schemas.microsoft.com/office/excel/2006/main"
            validation_formula = states_xml.findtext(
                f".//{{{x14}}}dataValidation/{{{x14}}}formula1/{{{xm}}}f"
            )
            validation_range = states_xml.findtext(
                f".//{{{x14}}}dataValidation/{{{xm}}}sqref"
            )
        self.assertEqual(
            validation_formula,
            "DCANameList",
        )
        self.assertEqual(validation_range, "F5:Q1072")
        # Use a workbook-scoped, formula-backed list instead of a direct
        # cross-sheet reference or an export-time list of fixed strings.
        dropdown_name = workbook.defined_names["DCANameList"]
        self.assertIsNone(dropdown_name.localSheetId)
        self.assertEqual(
            dropdown_name.attr_text.lstrip("="),
            "'Character List'!$D$3:$D$2328",
        )
        character_sheet = workbook["Character List"]
        self.assertEqual(character_sheet["O1"].value, '=COUNTIF($A$3:$A$168,"<>")')
        self.assertEqual(character_sheet["O2"].value, '=ROWS($F$3:$F$1994)-COUNTBLANK($F$3:$F$1994)')
        self.assertEqual(character_sheet["G3"].value, '=IF(F3<>"",ROWS($F$3:F3)-COUNTBLANK($F$3:F3),"")')
        self.assertIn("$O$1", character_sheet["D3"].value)
        self.assertIn("$O$2", character_sheet["D3"].value)
        self.assertNotIn("$H$", character_sheet["D3"].value)
        self.assertEqual(character_sheet["H3"].value, '=SUM($J$3:J3)-J3')
        self.assertIn("$F$3:$F$1994", character_sheet["D3"].value)
        self.assertIn("COUNTIF($A$3:A3", character_sheet["I3"].value)
        for column in "DEFGHIJKLMN":
            dimension = character_sheet.column_dimensions[column]
            self.assertTrue(
                dimension.hidden or (dimension.width or 99) <= 0.11
            )
        workbook.close()

    def test_role_dropdowns_and_bilingual_how_to_example(self):
        workbook = load_workbook(TEMPLATE_FILE, data_only=False)
        self.addCleanup(workbook.close)
        characters = workbook["Character List"]
        self.assertIn("CHAR(10)", characters["J3"].value)
        self.assertIn("COUNTIF", characters["K1994"].value)
        self.assertIn("$B$3:$B$168", characters["E1994"].value)
        self.assertIn('IF(N3<=M3,""', characters["E3"].value)
        self.assertIn('RIGHT(', characters["F3"].value)
        self.assertIn('COUNTBLANK($F$3:F1994', characters["G1994"].value)
        guide = workbook["How to use"]
        self.assertIn("Jack [Student]", guide["B24"].value)
        self.assertIn("普通 .xlsx", guide["B24"].value)
        self.assertIsNone(guide["C24"].value)
        self.assertIsNone(guide["C25"].value)
        self.assertIn("Jack [Teacher]", guide["B24"].value)
        self.assertIn("Blank rows do not stop the list", guide["B25"].value)
        self.assertIn("空行不会截断列表", guide["B25"].value)
        self.assertIn("Automatic", guide["B25"].value)

    def test_dropdown_helpers_remain_live_after_new_character_rows(self):
        workbook = load_workbook(TEMPLATE_FILE, data_only=False)
        self.addCleanup(workbook.close)
        characters = workbook["Character List"]
        # In particular, the thirteenth name (row 15) must not be outside a
        # fixed list ending at row 14, as it was in the reported workbook.
        for row in (3, 15, 16, 168):
            self.assertEqual(characters[f"I{row}"].data_type, "f")
            self.assertIn(f"A{row}", characters[f"I{row}"].value)
            self.assertIn(f"B{row}", characters[f"J{row}"].value)
            self.assertIsNone(characters[f"A{row}"].value)
            self.assertIsNone(characters[f"B{row}"].value)
        for row in (3, 15, 168, 2328):
            self.assertEqual(characters[f"D{row}"].data_type, "f")
            self.assertIn("$A$3:$A$168", characters[f"D{row}"].value)
            self.assertIn("$F$3:$F$1994", characters[f"D{row}"].value)

    def test_template_has_no_retired_sheet_text_or_formula_dependencies(self):
        workbook = load_workbook(TEMPLATE_FILE, data_only=False)
        self.addCleanup(workbook.close)
        self.assertEqual(workbook.sheetnames, ["How to use", "Character List", "DCA States"])
        for sheet in workbook:
            self.assertIsNone(sheet.freeze_panes)
            for row in sheet:
                for cell in row:
                    if isinstance(cell.value, str):
                        self.assertNotIn("Shared Group", cell.value)
                        self.assertNotIn("共享群组", cell.value)
                        self.assertNotIn("#REF!", cell.value)


    def test_internal_dropdown_headers_do_not_spill_into_visible_columns(self):
        workbook = load_workbook(TEMPLATE_FILE, data_only=False)
        self.addCleanup(workbook.close)
        for column in "DEFGHIJKL":
            self.assertIsNone(workbook["Character List"][f"{column}2"].value)


if __name__ == "__main__":
    unittest.main()
