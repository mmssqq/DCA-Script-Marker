import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

import fitz
from openpyxl import Workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MARKER_FILE = PROJECT_ROOT / "dca_script_marker.py"
CONTENT_VIEW_FILE = (
    PROJECT_ROOT
    / "macOS App"
    / "DCA Script Marker"
    / "DCA Script Marker"
    / "ContentView.swift"
)
sys.path.insert(0, str(PROJECT_ROOT))

from tests.test_pdf_annotations import appearance_colours, includes_colour


class PageStateStyleCLITests(unittest.TestCase):
    def test_macos_app_exposes_only_the_three_editable_styles(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            'let styles = [\n'
            '        "Editable Full Marking",\n'
            '        "First Appearance Only",\n'
            '        "DCA State Legend"\n'
            '    ]',
            content_view,
        )
        self.assertNotIn('\n        "Full Marking",', content_view)

    def test_macos_app_exposes_performer_role_mapping_toggle(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            '@State private var showPerformerRoleMapping = false',
            content_view,
        )
        self.assertIn(
            '"Show DCA Name / Other Script Characters"',
            content_view,
        )
        self.assertIn(
            'arguments.append("--show-performer-role-mapping")',
            content_view,
        )

    def test_macos_app_exposes_floating_role_mapping_inspector(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn('"DCA States Inspector"', content_view)
        self.assertIn('"--list-role-mappings"', content_view)
        self.assertIn("newPanel.isFloatingPanel = true", content_view)
        self.assertIn("newPanel.hidesOnDeactivate = false", content_view)
        self.assertIn(".canJoinAllSpaces", content_view)

    def test_dca_states_inspector_has_one_click_state_navigation(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn('model.language.label("Previous")', content_view)
        self.assertIn('model.language.label("Next")', content_view)
        self.assertIn("selectState(offset: -1)", content_view)
        self.assertIn("selectState(offset: 1)", content_view)
        self.assertIn('Text(model.language.label("DCA State"))', content_view)
        self.assertIn('systemImage: "list.number"', content_view)

    def test_role_mapping_search_covers_every_dca_state(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            '"Search all states by DCA Name or other character"',
            content_view,
        )
        self.assertIn('"Other Script Characters Played"', content_view)
        self.assertIn("var resultIndexes: [String: Int]", content_view)
        self.assertIn('Text(model.language.label("DCA States"))', content_view)
        self.assertIn("appearance.stateName", content_view)
        self.assertIn('"Whole-project search', content_view)
        self.assertIn('"Each matching DCA Name is shown once.', content_view)

    def test_inspector_headers_show_names_and_optional_roles_only(self):
        content = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertNotIn("Shared Groups", content)
        self.assertNotIn("共享群组", content)
        self.assertGreaterEqual(content.count('"DCA Name"'), 4)
        self.assertEqual(content.count('"Other Script Characters Played"'), 2)
        self.assertEqual(content.count('"饰演的其他剧本角色"'), 2)

    def test_macos_app_surfaces_workbook_setup_failures(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            '"Check DCA project setup"',
            content_view,
        )
        self.assertIn('alert.addButton(withTitle: t("Edit DCA Project", "编辑 DCA 项目"))', content_view)
        self.assertIn("showMarkerFailureAlert(errorMessage)", content_view)

    def test_macos_app_exposes_version_two_project_workflow(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn('Button(action: createNewProject)', content_view)
        self.assertIn('Button(action: openProject)', content_view)
        self.assertIn('"Import Excel"', content_view)
        self.assertIn('"Export Excel"', content_view)
        self.assertIn('"--import-excel"', content_view)
        self.assertIn('"--export-excel"', content_view)
        self.assertIn('"--project", projectPath', content_view)
        self.assertIn('struct DCAProjectDocument: Codable', project_editor)
        self.assertIn('struct DCAProjectEditor: View', project_editor)
        self.assertIn('ScrollView(.horizontal)', project_editor)
        self.assertIn('stateHeaderCell(language.label("Page Hint")', project_editor)
        self.assertIn('stateHeaderCell(language.label("Start Line Text")', project_editor)
        self.assertIn('"DCA \\(dcaIndex + 1)"', project_editor)

    def test_main_page_places_export_on_editor_row_below_file_actions(self):
        content = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        start = content.index('Text(appLanguage.label("DCA Project"))')
        end = content.index("\n                FileRow(", start)
        block = content[start:end]
        new_button = block.index("Button(action: createNewProject)")
        open_button = block.index("Button(action: openProject)")
        import_button = block.index("importExcelProject()")
        edit_button = block.index("showProjectEditor = true")
        export_button = block.index("exportProjectExcel()")
        self.assertLess(new_button, open_button)
        self.assertLess(open_button, import_button)
        self.assertLess(import_button, edit_button)
        self.assertLess(edit_button, export_button)
        self.assertIn("HStack(spacing: 10)", block)
        self.assertIn('"Edit Character List and DCA States"', block)
        self.assertIn('"Main project setup"', block)
        self.assertIn(".frame(maxWidth: .infinity)", block)
        self.assertIn(".frame(height: 46)", block)
        self.assertIn("colors: [.blue, .cyan]", block)
        self.assertIn("RoundedRectangle(cornerRadius: 12)", block)
        self.assertGreaterEqual(block.count(".frame(width: 78)"), 4)

    def test_main_page_keeps_each_marking_style_explanation_visible(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        style_start = content_view.index(
            "HStack(alignment: .top, spacing: 28)"
        )
        style_end = content_view.index(
            "\n            if !message.isEmpty {",
            style_start,
        )
        style_block = content_view[style_start:style_end]

        self.assertIn('Text(appLanguage.label("Choose Marking Style"))', style_block)
        self.assertIn("ForEach(styles, id: \\.self)", style_block)
        self.assertIn(".frame(width: 500", style_block)
        self.assertIn("Text(helpText(for: style))", style_block)
        self.assertIn("markingStyleColour(for: style)", style_block)
        self.assertNotIn("Text(helpText(for: selectedStyle))", style_block)
        self.assertNotIn('Text("Selected:', content_view)
        self.assertIn('.frame(width: 168, height: 168)', style_block)
        self.assertLess(
            style_block.index('"Generate\\nMarked Script"'),
            style_block.index('appLanguage.label("DCA States")'),
        )
        self.assertIn(".padding(.top, 10)", style_block)
        self.assertIn("minWidth: 980", content_view)
        self.assertIn("idealWidth: 1040", content_view)
        self.assertIn("minHeight: 680", content_view)
        self.assertIn("idealHeight: 740", content_view)

        colour_start = content_view.index(
            "private func markingStyleColour(for style: String) -> Color"
        )
        colour_end = content_view.index(
            "\n    func chooseFile(allowedTypes:",
            colour_start,
        )
        colour_block = content_view[colour_start:colour_end]
        self.assertIn('case "Editable Full Marking":\n            return Color(red: 0.58, green: 0.75, blue: 0.88)', colour_block)
        self.assertIn('case "First Appearance Only":\n            return Color(red: 0.94, green: 0.84, blue: 0.58)', colour_block)
        self.assertIn('default:\n            return Color(red: 0.90, green: 0.70, blue: 0.75)', colour_block)
        self.assertIn(
            '.black.opacity(0.76)',
            colour_block,
        )
        self.assertIn("Color.indigo,", style_block)
        self.assertIn("RoundedRectangle(cornerRadius: 16)", style_block)
        self.assertIn(
            ".shadow(color: .indigo.opacity(0.24), radius: 6, y: 2)",
            style_block,
        )

    def test_primary_main_page_labels_are_larger_than_button_text(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            'Text(appLanguage.label("DCA Project"))\n'
            '                            .font(.system(size: 16, weight: .bold))',
            content_view,
        )
        self.assertIn(
            'Text(appLanguage.label("Choose Marking Style"))\n'
            '                        .font(.system(size: 16, weight: .bold))',
            content_view,
        )
        self.assertIn('title: "Script PDF"', content_view)
        self.assertIn('title: "Output Folder"', content_view)

        file_row_start = content_view.index("struct FileRow: View")
        file_row = content_view[file_row_start:]
        self.assertIn(
            'Text(language.label(title))\n'
            '                .font(.system(size: 16, weight: .bold))',
            file_row,
        )

    def test_excel_import_automatically_saves_project_beside_workbook(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        import_start = content_view.index(
            "    private func importExcelProject()"
        )
        import_end = content_view.index(
            "    private func exportProjectExcel()",
            import_start,
        )
        import_block = content_view[import_start:import_end]
        self.assertIn(
            "automaticImportedProjectURL(",
            import_block,
        )
        self.assertNotIn("projectSaveURL(", import_block)
        self.assertIn(
            'let baseName = excelURL.deletingPathExtension().lastPathComponent',
            content_view,
        )
        self.assertIn(
            '!FileManager.default.fileExists(atPath: candidate.path)',
            content_view,
        )

    def test_internal_project_character_list_is_optional(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        user_guide = (PROJECT_ROOT / "USER_GUIDE.md").read_text(
            encoding="utf-8"
        )
        self.assertIn(
            "Character List is optional.",
            project_editor,
        )
        self.assertNotIn(
            "is not in Character List.",
            project_editor,
        )
        self.assertIn("Character List is optional:", content_view)
        self.assertIn("Character List is optional.", user_guide)
        self.assertIn("Character List 为可选项", user_guide)

    def test_primary_project_add_buttons_have_large_click_targets(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn(
            ".frame(minWidth: 200, minHeight: 34)",
            project_editor,
        )
        for label in (
            "Add DCA Name",
            "Add DCA State",
        ):
            self.assertIn(
                f'title: language.label("{label}")',
                project_editor,
            )
        self.assertGreaterEqual(
            project_editor.count(".controlSize(.large)"),
            2,
        )

    def test_project_editor_autosave_does_not_rewrite_active_fields(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        self.assertIn(
            "saveCurrentProject(capturingInterface: false)",
            content_view,
        )
        self.assertIn(
            "private func saveCurrentProject(\n"
            "        capturingInterface: Bool = true\n"
            "    ) -> Bool",
            content_view,
        )
        self.assertIn(
            "if capturingInterface {\n"
            "            captureInterfaceInProject()\n"
            "        }",
            content_view,
        )

    def test_dca_assignment_focus_request_is_consumed_once(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn(
            "let focusBinding = _requestedFocus",
            project_editor,
        )
        self.assertIn(
            "if focusBinding.wrappedValue == focusRequest {\n"
            "                focusBinding.wrappedValue = nil\n"
            "            }",
            project_editor,
        )
        self.assertNotIn(
            "parent.requestedFocus = parent.focus",
            project_editor,
        )

    def test_internal_dca_cells_support_tab_navigation(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn("override func insertTab", project_editor)
        self.assertIn("override func insertBacktab", project_editor)
        self.assertIn("moveDCAAssignmentFocus", project_editor)
        self.assertIn("stateIndex * 12 + focus.dcaIndex + offset", project_editor)
        self.assertIn(
            'language.label("Delete Row")',
            project_editor,
        )

    def test_internal_dca_state_identity_stays_pinned_while_scrolling(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn(
            "final class StateTableHorizontalOffsetView",
            project_editor,
        )
        self.assertIn(
            "observedClipView.bounds.minX",
            project_editor,
        )
        self.assertIn(
            "StateTableHorizontalOffsetReader(",
            project_editor,
        )
        self.assertGreaterEqual(
            project_editor.count(".offset(x: frozenStateIdentityOffset)"),
            2,
        )
        self.assertIn(
            "private var frozenStateIdentityOffset: CGFloat {\n"
            "        stateTableHorizontalOffset\n"
            "    }",
            project_editor,
        )

    def test_special_multi_column_dca_name_example_is_in_user_guides(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        user_guide = (PROJECT_ROOT / "USER_GUIDE.md").read_text(
            encoding="utf-8"
        )
        self.assertIn(
            "put TOM and ALL THREE in DCA 1",
            content_view,
        )
        self.assertIn(
            "The printed ALL THREE cue then receives 1/2/3",
            content_view,
        )
        self.assertIn(
            "Special DCA-cell example",
            user_guide,
        )
        self.assertIn(
            "特别 DCA 单元格示例",
            user_guide,
        )
        self.assertIn(
            "| `DCA 3` | `APPLE` + new line + `ALL THREE` |",
            user_guide,
        )

    def test_internal_editor_has_no_membership_controls(self):
        editor = CONTENT_VIEW_FILE.with_name("DCAProjectEditor.swift").read_text(encoding="utf-8")
        self.assertNotIn("struct DCAProjectGroup", editor)
        self.assertNotIn("sharedGroupsEditor", editor)
        self.assertNotIn("Add Shared Group", editor)
        self.assertNotIn('case groups', editor)
        self.assertIn("convertLegacyAssignments()", editor)
        self.assertIn("writeConvertedCopy", editor)
        self.assertIn("transformBeforeCommit: { $0 }", editor)

    def test_dca_picker_lists_names_then_optional_script_roles(self):
        editor = CONTENT_VIEW_FILE.with_name("DCAProjectEditor.swift").read_text(encoding="utf-8")
        start = editor.index('Text(t("Add DCA Names"')
        picker = editor[start:editor.index("private func addDCAName", start)]
        self.assertLess(picker.index("ForEach(characterListDCANames,"),
                        picker.index("ForEach(characterListRoleChoices,"))
        self.assertNotIn("characterListSharedGroups", picker)
        self.assertIn("addDCAName(choice.assignmentLabel, to: text)", picker)
        self.assertIn('"Adds \\(choice.assignmentLabel)"', picker)
        self.assertIn('"填入 \\(choice.assignmentLabel)"', picker)
        self.assertIn("choice.dcaName,", picker)
        self.assertIn("dcaCellContains(", picker)

    def test_internal_dca_state_rows_are_more_compact(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn(".frame(width: 48, height: 76)", project_editor)
        self.assertIn(".frame(width: 140, height: 76)", project_editor)
        self.assertIn(".frame(width: width, height: 76)", project_editor)
        self.assertGreaterEqual(project_editor.count("height: 54"), 2)

    def test_dca_picker_selection_recognises_square_bracket_aliases(self):
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn(
            "private func dcaAssignmentNameKey(_ value: String)",
            project_editor,
        )
        self.assertIn(
            "let newKey = dcaAssignmentNameKey(name)",
            project_editor,
        )
        self.assertIn(
            "let nameKey = dcaAssignmentNameKey(name)",
            project_editor,
        )
        self.assertIn(
            ".contains { dcaAssignmentNameKey(String($0)) == nameKey }",
            project_editor,
        )

    def test_duplicate_dca_assignments_warn_without_blocking_generation(self):
        content_view = CONTENT_VIEW_FILE.read_text(encoding="utf-8")
        project_editor = CONTENT_VIEW_FILE.with_name(
            "DCAProjectEditor.swift"
        ).read_text(encoding="utf-8")
        self.assertIn("blockingValidationIssues", project_editor)
        self.assertIn("advisoryIssues", project_editor)
        self.assertIn("ignoredAdvisorySignatures", project_editor)
        self.assertIn('systemName: "xmark.circle.fill"', project_editor)
        self.assertIn('t("Ignore", "忽略")', project_editor)
        self.assertIn("confirmProjectAdvisories", content_view)
        self.assertIn('"Ignore and Continue"', content_view)
        self.assertIn("ignoredProjectAdvisorySignatures", content_view)
        self.assertIn("It will not stop generation.", content_view)

    def test_cli_lists_state_role_mappings_for_floating_inspector(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            template_file = temporary_path / "template.xlsx"

            workbook = Workbook()
            states_sheet = workbook.active
            states_sheet.title = "DCA States"
            states_sheet.append([
                "DCA State",
                "Start Line Text",
                "State Start Position",
                "DCA 1",
                "DCA 2",
            ])
            states_sheet.append([
                "Scene 1",
                "START ONE",
                "Before",
                "Ben",
                "Alex",
            ])
            states_sheet.append([
                "Scene 2",
                "START TWO",
                "Before",
                "Alex",
                "Ben",
            ])
            characters = workbook.create_sheet("Character List")
            characters.append([None, None])
            characters.append([
                "DCA Name",
                "Other Script Characters Played",
            ])
            characters.append([
                "Ben",
                "Barber\nButcher\nCoach",
            ])
            characters.append([
                "Alex",
                "",
            ])
            workbook.save(template_file)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(template_file),
                    "--list-role-mappings",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
            states = json.loads(result.stdout)
            self.assertEqual([state["name"] for state in states], [
                "Scene 1",
                "Scene 2",
            ])
            self.assertEqual(states[0]["rows"], [
                {
                    "dca": "1",
                    "performer": "Ben",
                    "roles": ["Barber", "Butcher", "Coach"],
                },
                {
                    "dca": "2",
                    "performer": "Alex",
                    "roles": [],
                },
            ])
            self.assertEqual(states[1]["rows"], [
                {
                    "dca": "1",
                    "performer": "Alex",
                    "roles": [],
                },
                {
                    "dca": "2",
                    "performer": "Ben",
                    "roles": ["Barber", "Butcher", "Coach"],
                },
            ])

    def test_cli_performer_role_mapping_flag_adds_one_card(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_pdf = temporary_path / "source.pdf"
            template_file = temporary_path / "template.xlsx"
            output_folder = temporary_path / "output"
            output_folder.mkdir()

            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((72, 110), "START", fontsize=12)
            source.save(source_pdf)
            source.close()

            workbook = Workbook()
            states_sheet = workbook.active
            states_sheet.title = "DCA States"
            states_sheet.append([
                "DCA State",
                "Start Line Text",
                "State Start Position",
                "DCA 1",
            ])
            states_sheet.append([
                "Scene 1",
                "START",
                "Before",
                "Ben",
            ])
            characters = workbook.create_sheet("Character List")
            characters.append([None, None])
            characters.append([
                "DCA Name",
                "Other Script Characters Played",
            ])
            characters.append([
                "Ben",
                "Barber\nButcher\nCoach",
            ])
            workbook.save(template_file)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(template_file),
                    "--script",
                    str(source_pdf),
                    "--output",
                    str(output_folder),
                    "--style",
                    "Editable Full Marking",
                    "--show-performer-role-mapping",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

            marked_pdf = next(output_folder.glob("*_marked_*.pdf"))
            document = fitz.open(marked_pdf)
            cards = [
                annotation
                for annotation in document[0].annots() or []
                if "Performer / Role Mapping"
                in annotation.info.get("content", "")
            ]
            self.assertEqual(len(cards), 1)
            self.assertIn("DCA 1 | Ben", cards[0].info["content"])
            document.close()

    def test_page_state_display_argument_selects_footer_only(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_pdf = temporary_path / "source.pdf"
            template_file = temporary_path / "template.xlsx"
            output_folder = temporary_path / "output"
            output_folder.mkdir()

            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((72, 72), "START", fontsize=12)
            source.save(source_pdf)
            source.close()

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "DCA States"
            worksheet.append(
                [
                    "DCA State",
                    "Start Line Text",
                    "State Start Position",
                ]
            )
            worksheet.append(["Scene 1", "START", "Before"])
            workbook.save(template_file)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(template_file),
                    "--script",
                    str(source_pdf),
                    "--output",
                    str(output_folder),
                    "--style",
                    "Editable Full Marking",
                    "--page-state-display",
                    "footer",
                    "--state-font",
                    "Helvetica",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

            marked_pdf = next(output_folder.glob("*_marked_*.pdf"))
            document = fitz.open(marked_pdf)
            page = document[0]
            margin_labels = [
                annotation
                for annotation in page.annots() or []
                if annotation.info.get("content") == "Scene 1"
                and (
                    annotation.rect.y1 < 50
                    or annotation.rect.y0 > page.rect.height - 50
                )
            ]

            self.assertEqual(len(margin_labels), 1)
            self.assertGreater(
                margin_labels[0].rect.y0,
                page.rect.height - 50,
            )
            document.close()

    def test_page_state_style_arguments_reach_the_pdf(self):
        text_colour = (0.85, 0.0, 0.35)
        border_colour = (0.0, 0.45, 0.25)

        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_pdf = temporary_path / "source.pdf"
            template_file = temporary_path / "template.xlsx"
            output_folder = temporary_path / "output"
            output_folder.mkdir()

            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((72, 72), "START", fontsize=12)
            source.save(source_pdf)
            source.close()

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "DCA States"
            worksheet.append(
                [
                    "DCA State",
                    "Start Line Text",
                    "State Start Position",
                ]
            )
            worksheet.append(["Scene 1", "START", "Before"])
            workbook.save(template_file)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(template_file),
                    "--script",
                    str(source_pdf),
                    "--output",
                    str(output_folder),
                    "--style",
                    "Editable Full Marking",
                    "--state-colour",
                    "blue",
                    "--state-scale",
                    "1.2",
                    "--state-font",
                    "Helvetica",
                    "--page-state-header-footer",
                    "--page-state-text-colour",
                    "red",
                    "--page-state-scale",
                    "1.45",
                    "--page-state-font",
                    "Times",
                    "--page-state-border-colour",
                    "green",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

            marked_pdf = next(output_folder.glob("*_marked_*.pdf"))
            document = fitz.open(marked_pdf)
            page = document[0]
            page_labels = [
                annotation
                for annotation in page.annots() or []
                if annotation.info.get("content") == "Scene 1"
                and (
                    annotation.rect.y1 < 50
                    or annotation.rect.y0 > page.rect.height - 50
                )
            ]

            self.assertEqual(len(page_labels), 2)
            for annotation in page_labels:
                strokes, fills = appearance_colours(document, annotation)
                self.assertTrue(includes_colour(strokes, border_colour))
                self.assertTrue(includes_colour(fills, text_colour))
                _, default_style = document.xref_get_key(
                    annotation.xref,
                    "DS",
                )
                self.assertIn("Times New Roman", default_style)
                self.assertIn("17.4pt", default_style)
            document.close()

    def test_extended_palette_reaches_every_annotation_group(self):
        number_colour = (0.78, 0.24, 0.0)
        state_colour = (0.50, 0.20, 0.65)
        page_text_colour = (0.35, 0.35, 0.35)
        page_border_colour = (0.45, 0.25, 0.10)

        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_pdf = temporary_path / "source.pdf"
            template_file = temporary_path / "template.xlsx"
            output_folder = temporary_path / "output"
            output_folder.mkdir()

            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((140, 72), "START", fontsize=12)
            page.insert_text((140, 120), "ALICE:", fontsize=12)
            source.save(source_pdf)
            source.close()

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "DCA States"
            worksheet.append([
                "DCA State",
                "Start Line Text",
                "State Start Position",
                "DCA 1",
            ])
            worksheet.append(["Scene 1", "START", "Before", "ALICE"])
            workbook.save(template_file)
            workbook.close()

            result = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(template_file),
                    "--script",
                    str(source_pdf),
                    "--output",
                    str(output_folder),
                    "--style",
                    "Editable Full Marking",
                    "--number-colour",
                    "orange",
                    "--state-colour",
                    "purple",
                    "--state-font",
                    "Helvetica",
                    "--page-state-display",
                    "both",
                    "--page-state-text-colour",
                    "grey",
                    "--page-state-font",
                    "Helvetica",
                    "--page-state-border-colour",
                    "brown",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(
                result.returncode,
                0,
                result.stdout + result.stderr,
            )

            marked_pdf = next(output_folder.glob("*_marked_*.pdf"))
            document = fitz.open(marked_pdf)
            page = document[0]
            number_seen = False
            body_state_seen = False
            margin_state_count = 0

            for annotation in page.annots() or []:
                content = annotation.info.get("content", "")
                strokes, fills = appearance_colours(document, annotation)

                if content == "1":
                    number_seen = includes_colour(fills, number_colour)
                elif content == "Scene 1" and (
                    annotation.rect.y1 < 50
                    or annotation.rect.y0 > page.rect.height - 50
                ):
                    margin_state_count += 1
                    self.assertTrue(
                        includes_colour(strokes, page_border_colour)
                    )
                    self.assertTrue(
                        includes_colour(fills, page_text_colour)
                    )
                elif content == "Scene 1":
                    body_state_seen = includes_colour(fills, state_colour)

            document.close()

            self.assertTrue(number_seen)
            self.assertTrue(body_state_seen)
            self.assertEqual(margin_state_count, 2)

    def test_retained_special_styles_create_only_editable_output(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            temporary_path = Path(temporary_directory)
            source_pdf = temporary_path / "source.pdf"
            template_file = temporary_path / "template.xlsx"

            source = fitz.open()
            page = source.new_page(width=595, height=842)
            page.insert_text((140, 72), "START", fontsize=12)
            page.insert_text((140, 120), "ALICE:", fontsize=12)
            page.insert_text((140, 180), "ALICE:", fontsize=12)
            source.save(source_pdf)
            source.close()

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "DCA States"
            worksheet.append([
                "DCA State",
                "Start Line Text",
                "State Start Position",
                "DCA 1",
            ])
            worksheet.append(["Scene 1", "START", "Before", "ALICE"])
            workbook.save(template_file)
            workbook.close()

            for style, expected_dca_count, expected_legend_count in (
                ("First Appearance Only", 1, 0),
                ("DCA State Legend", 0, 1),
            ):
                with self.subTest(style=style):
                    output_folder = temporary_path / style.replace(" ", "-")
                    output_folder.mkdir()
                    result = subprocess.run(
                        [
                            sys.executable,
                            str(MARKER_FILE),
                            "--template",
                            str(template_file),
                            "--script",
                            str(source_pdf),
                            "--output",
                            str(output_folder),
                            "--style",
                            style,
                            "--number-font",
                            "Helvetica",
                            "--state-font",
                            "Helvetica",
                            "--page-state-display",
                            "both",
                        ],
                        capture_output=True,
                        text=True,
                        check=False,
                    )
                    self.assertEqual(
                        result.returncode,
                        0,
                        result.stdout + result.stderr,
                    )

                    marked_pdf = next(output_folder.glob("*_marked_*.pdf"))
                    moved_pdf = output_folder / "moved.pdf"
                    annotations_removed_pdf = (
                        output_folder / "annotations-removed.pdf"
                    )
                    document = fitz.open(marked_pdf)
                    page = document[0]
                    annotations = list(page.annots() or [])
                    self.assertTrue(annotations)
                    self.assertTrue(all(
                        annotation.type[1] == "FreeText"
                        for annotation in annotations
                    ))

                    margin_labels = [
                        annotation
                        for annotation in annotations
                        if annotation.info.get("content") == "Scene 1"
                        and (
                            annotation.rect.y1 < 50
                            or annotation.rect.y0 > page.rect.height - 50
                        )
                    ]
                    self.assertEqual(len(margin_labels), 2)
                    for annotation in margin_labels:
                        self.assertAlmostEqual(
                            annotation.border["width"],
                            0.8,
                            places=3,
                        )

                    dca_numbers = [
                        annotation
                        for annotation in annotations
                        if annotation.info.get("content") == "1"
                    ]
                    legends = [
                        annotation
                        for annotation in annotations
                        if annotation.info.get("content", "")
                        .casefold()
                        .startswith("scene 1\n1: alice")
                    ]
                    self.assertEqual(len(dca_numbers), expected_dca_count)
                    self.assertEqual(len(legends), expected_legend_count)

                    movable = (dca_numbers or legends)[0]
                    movable_content = movable.info.get("content")
                    moved_rect = movable.rect + (30, 35, 30, 35)
                    movable.set_rect(moved_rect)
                    movable.update()
                    document.save(moved_pdf)
                    document.close()

                    moved_document = fitz.open(moved_pdf)
                    moved_page = moved_document[0]
                    moved_annotation = next(
                        annotation
                        for annotation in moved_page.annots() or []
                        if annotation.info.get("content") == movable_content
                    )
                    self.assertAlmostEqual(
                        moved_annotation.rect.x0,
                        moved_rect.x0,
                        places=1,
                    )
                    self.assertAlmostEqual(
                        moved_annotation.rect.y0,
                        moved_rect.y0,
                        places=1,
                    )
                    self.assertEqual(
                        moved_document.xref_get_key(
                            moved_annotation.xref,
                            "AP",
                        )[0],
                        "dict",
                    )
                    for annotation in list(moved_page.annots() or []):
                        moved_page.delete_annot(annotation)
                    moved_document.save(annotations_removed_pdf)
                    moved_document.close()

                    cleaned_document = fitz.open(annotations_removed_pdf)
                    cleaned_page = cleaned_document[0]
                    self.assertEqual(list(cleaned_page.annots() or []), [])
                    self.assertNotIn("Scene 1", cleaned_page.get_text())
                    self.assertNotIn(
                        "1: alice",
                        cleaned_page.get_text().casefold(),
                    )
                    self.assertEqual(cleaned_page.get_drawings(), [])
                    cleaned_document.close()


if __name__ == "__main__":
    unittest.main()
