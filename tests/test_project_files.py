import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
MARKER_FILE = PROJECT_ROOT / "dca_script_marker.py"
sys.path.insert(0, str(PROJECT_ROOT))

from dca_script_marker import (
    excel_role_choice,
    export_project_excel,
    import_excel_project,
    load_project,
    load_template,
    project_to_workbook,
)


def build_workbook(path):
    workbook = Workbook()
    states = workbook.active
    states.title = "DCA States"
    states.append([
        "DCA State",
        "Start Line Character",
        "Start Line Text",
        "State Start position",
        "Page Hint",
        "Notes",
        "DCA 1",
        "DCA 2",
    ])
    states.append([
        "Scene 1",
        "Barber",
        "Welcome",
        "Before",
        "3",
        "Opening",
        "Ben",
        "Alex",
    ])
    states.append([
        "Scene 2",
        "Alex",
        "Goodbye",
        "After",
        "8",
        "",
        "Alex",
        "Ben",
    ])
    characters = workbook.create_sheet("Character List")
    characters.append(["Heading", None])
    characters.append([
        "DCA Name",
        "Other Script Characters Played",
    ])
    characters.append(["Ben", "Barber\nButcher\nCoach"])
    characters.append(["Alex", ""])
    workbook.save(path)
    workbook.close()


class ProjectFileTests(unittest.TestCase):
    def test_excel_import_and_project_loader_match_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            directory = Path(directory)
            workbook_path = directory / "show.xlsx"
            project_path = directory / "show.dcamarker"
            build_workbook(workbook_path)

            project = import_excel_project(workbook_path)
            project_path.write_text(
                json.dumps(project, ensure_ascii=False),
                encoding="utf-8",
            )
            workbook_states, workbook_assignments = load_template(
                workbook_path
            )
            project_states, project_assignments = load_project(project_path)

            self.assertEqual(project["name"], "show")
            self.assertEqual(len(project["characters"]), 2)
            self.assertEqual(len(project["states"]), 2)
            self.assertEqual(
                project["characters"][0]["other_characters"],
                "Barber\nButcher\nCoach",
            )
            self.assertEqual(project_states, workbook_states)
            self.assertEqual(project_assignments, workbook_assignments)

    def test_legacy_excel_converts_selected_labels_without_changing_source(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "legacy.xlsx"
            build_workbook(path)
            workbook = load_workbook(path)
            groups = workbook.create_sheet("Shared Groups")
            groups.append(["Shared Group Name", "DCA Members — one per line"])
            groups.append(["ALL", "Ben\nAlex"])
            workbook["DCA States"]["G2"] = "ALL"
            workbook.save(path)
            workbook.close()
            original = path.read_bytes()
            project = import_excel_project(path)
            self.assertNotIn("shared_groups", project)
            self.assertEqual(project["states"][0]["dca_assignments"][0], "ALL\nBen\nAlex")
            self.assertIn("ALL", [c["dca_name"] for c in project["characters"]])
            self.assertEqual(path.read_bytes(), original)

    def test_project_exports_a_compatible_styled_workbook(self):
        with tempfile.TemporaryDirectory() as directory:
            directory = Path(directory)
            source_path = directory / "source.xlsx"
            project_path = directory / "show.dcamarker"
            exported_path = directory / "exported.xlsx"
            build_workbook(source_path)
            project = import_excel_project(source_path)
            project_path.write_text(
                json.dumps(project, ensure_ascii=False),
                encoding="utf-8",
            )

            export_project_excel(project_path, exported_path)
            exported_states, exported_assignments = load_template(
                exported_path
            )
            source_states, source_assignments = load_template(source_path)
            self.assertEqual(exported_states, source_states)
            self.assertEqual(exported_assignments, source_assignments)

            workbook = load_workbook(exported_path, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "How to use",
                    "Character List",
                    "DCA States",
                ],
            )
            self.assertEqual(
                [cell.value for cell in workbook["DCA States"][4]],
                [
                    "DCA State",
                    "Start Line Character",
                    "Start Line Text",
                    "State Start Position",
                    "Page Hint",
                    *[f"DCA {number}" for number in range(1, 13)],
                    "Notes",
                ],
            )
            self.assertEqual(
                workbook["Character List"]["B3"].value,
                "Barber\nButcher\nCoach",
            )
            self.assertEqual(workbook["DCA States"]["F5"].value, "Ben")
            self.assertEqual(
                workbook["How to use"]["A1"].font.name,
                "Carlito",
            )
            self.assertEqual(
                workbook["How to use"]["A1"].fill.fgColor.rgb,
                "FF2F5E86",
            )
            self.assertEqual(
                workbook["Character List"]["A2"].fill.fgColor.rgb,
                "FF93B4CC",
            )
            self.assertEqual(
                workbook["Character List"]["A3"].fill.fgColor.rgb,
                "FFF4F8FB",
            )
            self.assertEqual(
                workbook["Character List"]["B3"].fill.fgColor.rgb,
                "FFFFFDF5",
            )
            self.assertEqual(
                workbook["DCA States"]["A4"].font.name,
                "Carlito",
            )
            self.assertEqual(
                workbook["DCA States"]["A4"].fill.fgColor.rgb,
                "FF8FAFC8",
            )
            self.assertEqual(
                workbook["DCA States"]["A5"].fill.fgColor.rgb,
                "FFF4F8FB",
            )
            self.assertEqual(
                workbook["DCA States"]["F5"].fill.fgColor.rgb,
                "FFFFFDF5",
            )
            self.assertEqual(
                workbook["DCA States"]["F5"].border.bottom.style,
                "thin",
            )
            self.assertFalse(
                workbook["DCA States"].sheet_view.showGridLines,
            )
            self.assertIsNone(
                workbook["Character List"].freeze_panes,
            )
            self.assertIsNone(
                workbook["DCA States"].freeze_panes,
            )
            self.assertTrue(
                workbook["Character List"].column_dimensions["D"].hidden,
            )
            workbook.close()

    def test_role_shortcut_labels_keep_a_single_alias_bracket(self):
        self.assertEqual(excel_role_choice("Jack", "Student"), "Jack [Student]")
        self.assertEqual(excel_role_choice("Jack [John]", "Student"), "Jack [John, Student]")
        self.assertEqual(excel_role_choice("Jack", "Student [Kid]"), "Jack [Student, Kid]")
        self.assertEqual(excel_role_choice("Jack [John]", "John"), "Jack [John]")
        self.assertEqual(excel_role_choice("Jack", "jack"), "")
        self.assertEqual(excel_role_choice("", "Student"), "")
        self.assertEqual(excel_role_choice("Jack", ""), "")

    def test_exported_role_choices_keep_dca_identity_after_import(self):
        project = {
            "schema_version": 1,
            "name": "Role shortcut",
            "characters": [
                {"id": "jack", "dca_name": "Jack", "other_characters": "Student\nTeacher"},
                {"id": "jane", "dca_name": "Jane [Jay]", "other_characters": "Doctor"},
                {"id": "blank", "dca_name": "", "other_characters": "Orphan"},
            ],
            "states": [{
                "id": "s1", "name": "S1", "start_line_character": "Student",
                "start_line_text": "Hello", "start_position": "Before",
                "page_hint": "1", "notes": "", "dca_assignments": ["Jack [Student]", "Jane [Jay, Doctor]"] + [""] * 10,
            }],
        }
        workbook = project_to_workbook(project)
        self.addCleanup(workbook.close)
        self.assertEqual(
            [workbook["Character List"][f"D{r}"].value for r in range(3,8)],
            ["Jack", "Jane [Jay]", "Jack [Student]", "Jack [Teacher]", "Jane [Jay, Doctor]"],
        )
        self.assertIn("Jack [Student]", workbook["How to use"]["A4"].value)
        self.assertIn("Jack [Teacher]", workbook["How to use"]["A5"].value)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "roles.xlsx"
            workbook.save(path)
            _, assignments = load_template(path)
            self.assertEqual(assignments["s1"]["jack"], ["1"])
            self.assertEqual(assignments["s1"]["student"], ["1"])
            self.assertEqual(assignments["s1"]["teacher"], ["1"])
            self.assertEqual(assignments["s1"]["jane"], ["2"])
            self.assertEqual(assignments["s1"]["jay"], ["2"])
            self.assertEqual(assignments["s1"]["doctor"], ["2"])

    def test_legacy_export_uses_only_ordinary_names_and_preserves_original(self):
        with tempfile.TemporaryDirectory() as directory:
            directory = Path(directory)
            project_path = directory / "legacy.dcamarker"
            exported_path = directory / "converted.xlsx"
            project = {
                "schema_version": 1, "name": "Legacy",
                "characters": [
                    {"id": "tom", "dca_name": "Tom", "other_characters": "ALL\nStudent"},
                    {"id": "jerry", "dca_name": "Jerry", "other_characters": "ALL"},
                ],
                "shared_groups": [{"name": "ALL", "members": "Tom\nJerry"}],
                "states": [{
                    "id": "s1", "name": "S1", "start_line_character": "ALL",
                    "start_line_text": "Opening", "start_position": "After",
                    "page_hint": "1", "notes": "",
                    "dca_assignments": ["ALL"] + [""] * 11,
                }],
            }
            project_path.write_text(json.dumps(project), encoding="utf-8")
            original = project_path.read_bytes()
            export_project_excel(project_path, exported_path)
            self.assertEqual(project_path.read_bytes(), original)
            workbook = load_workbook(exported_path)
            self.assertEqual(workbook.sheetnames, ["How to use", "Character List", "DCA States"])
            self.assertEqual(workbook["DCA States"]["F5"].value, "ALL\nTom\nJerry")
            self.assertEqual(workbook["Character List"]["B3"].value, "Student")
            self.assertEqual([workbook["Character List"][f"D{r}"].value for r in range(3,7)],
                             ["Tom", "Jerry", "ALL", "Tom [Student]"])
            workbook.close()
            states, assignments = load_template(exported_path)
            self.assertEqual(assignments["s1"]["tom"], ["1"])
            self.assertEqual(assignments["s1"]["student"], ["1"])
            self.assertEqual(assignments["s1"]["jerry"], ["1"])
            self.assertEqual(assignments["s1"]["all"], ["1"])
            self.assertEqual({row["performer"] for row in states[0]["dca_reference_rows"]},
                             {"Tom", "Jerry", "ALL"})
            self.assertEqual(load_project(project_path), (states, assignments))

    def test_cli_import_and_export_round_trip(self):
        with tempfile.TemporaryDirectory() as directory:
            directory = Path(directory)
            source_path = directory / "source.xlsx"
            project_path = directory / "show.dcamarker"
            exported_path = directory / "exported.xlsx"
            build_workbook(source_path)

            imported = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--template",
                    str(source_path),
                    "--import-excel",
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(
                imported.returncode,
                0,
                imported.stdout + imported.stderr,
            )
            json.loads(imported.stdout)
            project_path.write_text(imported.stdout, encoding="utf-8")

            exported = subprocess.run(
                [
                    sys.executable,
                    str(MARKER_FILE),
                    "--project",
                    str(project_path),
                    "--export-excel",
                    str(exported_path),
                ],
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(
                exported.returncode,
                0,
                exported.stdout + exported.stderr,
            )
            self.assertTrue(exported_path.is_file())


if __name__ == "__main__":
    unittest.main()
