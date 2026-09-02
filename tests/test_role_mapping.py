import json
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

import dca_script_marker as marker


class PerformerRoleMappingTests(unittest.TestCase):
    def setUp(self):
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.temporary_path = Path(self.temporary_directory.name)

    def tearDown(self):
        self.temporary_directory.cleanup()

    def save_workbook(self, workbook, name="template.xlsx"):
        template_file = self.temporary_path / name
        workbook.save(template_file)
        workbook.close()
        return template_file

    def add_character_list(
        self,
        workbook,
        rows,
        canonical_header="DCA Name",
        role_header="Other Script Characters Played",
    ):
        worksheet = workbook.create_sheet("Character List")
        worksheet.cell(row=2, column=1, value=canonical_header)
        worksheet.cell(row=2, column=2, value=role_header)

        for row_number, (performer, roles) in enumerate(rows, start=3):
            worksheet.cell(row=row_number, column=1, value=performer)
            worksheet.cell(row=row_number, column=2, value=roles)

    def add_shared_groups(self, workbook, rows):
        worksheet = workbook.create_sheet("Shared Groups")
        worksheet.cell(row=2, column=1, value="Shared Group Name")
        worksheet.cell(
            row=2,
            column=2,
            value="DCA Members — one per line",
        )
        for row_number, (group, members) in enumerate(rows, start=3):
            worksheet.cell(row=row_number, column=1, value=group)
            worksheet.cell(row=row_number, column=2, value=members)

    def test_missing_sheet_dimensions_preserve_project_and_marking_data(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        for column, value in enumerate(
            ["DCA State", "Start Line Text", "DCA 1"], 1
        ):
            states_sheet.cell(row=4, column=column, value=value)
        for column, value in enumerate(["S1", "Opening", "TRIO"], 1):
            states_sheet.cell(row=5, column=column, value=value)
        self.add_character_list(
            workbook,
            [("Tom", "Captain"), ("Jerry", "Guide"), ("Apple", "Guest")],
        )
        groups = workbook.create_sheet("Shared Groups")
        groups.cell(row=8, column=1, value="Shared Group Name")
        groups.cell(row=8, column=2, value="DCA Members — one per line")
        groups.cell(row=9, column=1, value="TRIO")
        groups.cell(row=9, column=2, value="Tom\nJerry\nApple")
        path = self.save_workbook(workbook)
        expected_marking = marker.load_template(path)
        expected_project = marker.import_excel_project(path)

        for data_only, operation, expected in (
            (True, marker.load_template, expected_marking),
            (False, marker.import_excel_project, expected_project),
        ):
            reader = load_workbook(path, read_only=True, data_only=data_only)
            self.addCleanup(reader.close)
            for sheet in reader:
                # A valid XLSX may omit the optional dimension element.
                sheet.reset_dimensions()
                self.assertIsNone(sheet.max_row)
            with patch.object(marker, "load_workbook", return_value=reader):
                self.assertEqual(operation(path), expected)

    def test_one_performer_expands_to_roles_in_every_state(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Character",
            "Start Line Text",
            "State Start Position",
            "DCA 1",
            "DCA 2",
        ])
        states_sheet.append([
            "Scene 1",
            "Mary",
            "Opening one",
            "Before",
            "Mary",
            None,
        ])
        states_sheet.append([
            "Scene 2",
            "Michael",
            "Opening two",
            "After",
            None,
            "Mary",
        ])
        self.add_character_list(
            workbook,
            [("Mary", "Tom, Jerry\nMichael")],
        )

        states, assignments = marker.load_template(
            self.save_workbook(workbook)
        )

        expected_names = {"mary", "tom", "jerry", "michael"}
        self.assertEqual(set(assignments["scene 1"]), expected_names)
        self.assertEqual(set(assignments["scene 2"]), expected_names)
        for name in expected_names:
            self.assertEqual(assignments["scene 1"][name], ["1"])
            self.assertEqual(assignments["scene 2"][name], ["2"])

        self.assertIs(
            marker.get_matching_state(
                states,
                "Opening one",
                set(),
                speaker_names=["Jerry"],
            ),
            states[0],
        )
        self.assertIs(
            marker.get_matching_state(
                states,
                "Opening two",
                set(),
                speaker_names=["Mary"],
            ),
            states[1],
        )
        self.assertIsNone(
            marker.get_matching_state(
                states,
                "Opening one",
                set(),
                speaker_names=["Beth"],
            )
        )

        self.assertEqual(
            marker.build_legend_text(states[0], assignments),
            "Scene 1\n1: mary",
        )
        self.assertEqual(
            marker.build_legend_text(states[1], assignments),
            "Scene 2\n2: mary",
        )
        self.assertEqual(
            states[0]["performer_role_rows"],
            [{
                "dca": ["1"],
                "performer": "Mary",
                "roles": ["Tom", "Jerry", "Michael"],
            }],
        )
        self.assertEqual(
            states[1]["performer_role_rows"],
            [{
                "dca": ["2"],
                "performer": "Mary",
                "roles": ["Tom", "Jerry", "Michael"],
            }],
        )

    def test_old_workbook_headers_and_inline_aliases_remain_supported(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Cue Speaker",
            "Start Cue Text",
            "Start Position",
            "DCA 3",
        ])
        states_sheet.append([
            "Legacy Scene",
            "Mary",
            "Legacy opening",
            "Before",
            "Mary [Tom, Jerry]",
        ])

        old_character_list = workbook.create_sheet("Character List")
        old_character_list["A1"] = "Character List"
        old_character_list["A3"] = "Mary"

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "old-template.xlsx")
        )

        self.assertEqual(
            assignments,
            {
                "legacy scene": {
                    "mary": ["3"],
                    "tom": ["3"],
                    "jerry": ["3"],
                }
            },
        )
        self.assertIs(
            marker.get_matching_state(
                states,
                "Legacy opening",
                set(),
                speaker_names=["Mary"],
            ),
            states[0],
        )
        self.assertEqual(
            marker.build_legend_text(states[0], assignments),
            "Legacy Scene\n3: mary\n3: tom\n3: jerry",
        )
        self.assertEqual(states[0]["performer_role_rows"], [])

    def test_old_script_characters_played_header_remains_supported(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
        ])
        states_sheet.append(["Scene 1", "Opening", "Ben"])
        self.add_character_list(
            workbook,
            [("Ben", "Barber")],
            role_header="Script Characters Played",
        )

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "old-role-header.xlsx")
        )

        self.assertEqual(set(assignments["scene 1"]), {"ben", "barber"})
        self.assertEqual(states[0]["dca_reference_rows"][0]["roles"], [
            "Barber",
        ])

    def test_inspector_rows_include_dcas_without_other_characters(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
            "DCA 2",
        ])
        states_sheet.append(["Scene 1", "Opening", "Ben", "Alex"])
        self.add_character_list(
            workbook,
            [
                ("Ben", "Barber\nButcher"),
                ("Alex", None),
            ],
        )

        states, _assignments = marker.load_template(
            self.save_workbook(workbook, "complete-inspector.xlsx")
        )

        self.assertEqual(
            states[0]["performer_role_rows"],
            [{
                "dca": ["1"],
                "performer": "Ben",
                "roles": ["Barber", "Butcher"],
            }],
        )
        self.assertEqual(
            states[0]["dca_reference_rows"],
            [
                {
                    "dca": ["1"],
                    "performer": "Ben",
                    "roles": ["Barber", "Butcher"],
                },
                {
                    "dca": ["2"],
                    "performer": "Alex",
                    "roles": [],
                },
            ],
        )

    def test_mapping_card_rows_exclude_inline_aliases(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 4",
        ])
        states_sheet.append([
            "Scene 1",
            "Opening",
            "Ben [B.]",
        ])
        self.add_character_list(
            workbook,
            [("Ben", "Barber\nButcher\nCoach")],
        )

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "mapping-card-rows.xlsx")
        )

        self.assertIn("b.", assignments["scene 1"])
        self.assertEqual(
            states[0]["performer_role_rows"],
            [{
                "dca": ["4"],
                "performer": "Ben",
                "roles": ["Barber", "Butcher", "Coach"],
            }],
        )
        self.assertNotIn(
            "B.",
            marker.build_performer_role_mapping_text(states[0]),
        )

    def test_long_performer_header_remains_supported(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
        ])
        states_sheet.append(["Scene 1", "Opening", "Ben"])
        self.add_character_list(
            workbook,
            [("Ben", "Barber\nButcher\nCoach")],
            canonical_header="DCA Name / Performer",
        )

        _, assignments = marker.load_template(
            self.save_workbook(workbook, "long-header-template.xlsx")
        )

        self.assertEqual(
            set(assignments["scene 1"]),
            {"ben", "barber", "butcher", "coach"},
        )

    def test_legacy_assignments_sheet_uses_global_roles_and_aliases(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Character",
            "Start Line Text",
        ])
        states_sheet.append(["Scene V", "Tom", "Vertical opening"])
        self.add_character_list(
            workbook,
            [("Mary", "Tom; Jerry | Michael")],
        )

        assignments_sheet = workbook.create_sheet("Assignments")
        assignments_sheet.append([
            "DCA State",
            "Character",
            "DCA",
            "Aliases",
        ])
        assignments_sheet.append(["Scene V", "Mary", 7, "Madame"])

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "vertical-template.xlsx")
        )

        self.assertEqual(
            set(assignments["scene v"]),
            {"mary", "tom", "jerry", "michael", "madame"},
        )
        for name in assignments["scene v"]:
            self.assertEqual(assignments["scene v"][name], ["7"])
        self.assertTrue(marker.cue_speaker_matches(states[0], ["Michael"]))
        self.assertEqual(
            marker.build_legend_text(states[0], assignments),
            "Scene V\n7: mary\n7: madame",
        )

    def test_unique_mapped_role_in_dca_state_resolves_to_performer(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
        ])
        states_sheet.append(["Scene 1", "Opening", "Tom"])
        self.add_character_list(
            workbook,
            [("Mary", "Tom, Jerry")],
        )
        states, assignments = marker.load_template(
            self.save_workbook(workbook, "mapped-role-in-state.xlsx")
        )

        self.assertEqual(
            assignments["scene 1"],
            {
                "mary": ["1"],
                "tom": ["1"],
                "jerry": ["1"],
            },
        )
        self.assertEqual(
            marker.build_legend_text(states[0], assignments),
            "Scene 1\n1: mary",
        )
        self.assertEqual(
            states[0]["performer_role_rows"],
            [{
                "dca": ["1"],
                "performer": "Mary",
                "roles": ["Tom", "Jerry"],
            }],
        )

    def test_manual_name_on_multiple_faders_combines_only_selected_dcas(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DCA States"
        sheet.append(["DCA State", "Start Line Character", "Start Line Text", "DCA 1", "DCA 2", "DCA 3"])
        sheet.append(["Scene 1", "ALL THREE", "Opening one", "Tom\nALL THREE", "Jerry\nALL THREE", "Apple\nALL THREE"])
        sheet.append(["Scene 2", "ALL THREE", "Opening two", "Tom\nALL THREE", "", "Apple\nALL THREE"])
        self.add_character_list(workbook, [(name, None) for name in ("Tom", "Jerry", "Apple", "ALL THREE")])
        states, assignments = marker.load_template(self.save_workbook(workbook))
        self.assertEqual(assignments["scene 1"]["all three"], ["1", "2", "3"])
        self.assertEqual(assignments["scene 2"]["all three"], ["1", "3"])
        self.assertEqual(assignments["scene 1"]["tom"], ["1"])
        self.assertEqual(assignments["scene 1"]["jerry"], ["2"])
        self.assertEqual(assignments["scene 1"]["apple"], ["3"])
        self.assertTrue(marker.cue_speaker_matches(states[0], ["ALL THREE"]))
        self.assertFalse(marker.cue_speaker_matches(states[0], ["Tom"]))

    def test_one_fader_role_mapping_can_expand_additional_roles(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 4",
        ])
        states_sheet.append(["Scene 1", "Opening", "PERFORMER A"])
        self.add_character_list(
            workbook,
            [("PERFORMER A", "BARBER\nCOACH\nHOST")],
        )
        self.add_shared_groups(workbook, [])

        _, assignments = marker.load_template(
            self.save_workbook(workbook, "one-fader-role-mapping.xlsx")
        )

        self.assertEqual(assignments["scene 1"]["performer a"], ["4"])
        self.assertEqual(assignments["scene 1"]["barber"], ["4"])
        self.assertEqual(assignments["scene 1"]["coach"], ["4"])
        self.assertEqual(assignments["scene 1"]["host"], ["4"])

    def test_unused_legacy_definition_does_not_link_individual_cues(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 4",
            "DCA 5",
            "DCA 6",
        ])
        states_sheet.append([
            "Scene 1",
            "Opening",
            "TOM",
            "JERRY",
            "APPLE",
        ])
        self.add_character_list(
            workbook,
            [("TOM", None), ("JERRY", None), ("APPLE", None)],
        )
        self.add_shared_groups(
            workbook,
            [("ALL THREE", "TOM\nJERRY\nAPPLE")],
        )

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "explicit-shared-group.xlsx")
        )

        self.assertEqual(assignments["scene 1"]["tom"], ["4"])
        self.assertEqual(assignments["scene 1"]["jerry"], ["5"])
        self.assertEqual(assignments["scene 1"]["apple"], ["6"])
        self.assertNotIn("all three", assignments["scene 1"])
        self.assertEqual(
            [row["performer"] for row in states[0]["dca_reference_rows"]],
            ["TOM", "JERRY", "APPLE"],
        )

    def test_new_workbook_does_not_guess_repeated_roles_are_groups(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
            "DCA 2",
        ])
        states_sheet.append(["Scene 1", "Opening", "M1", "M2"])
        self.add_character_list(
            workbook,
            [("M1", "ALL"), ("M2", "ALL")],
        )
        self.add_shared_groups(workbook, [])

        with self.assertRaisesRegex(
            ValueError,
            r'Give each mapped role one DCA Name',
        ):
            marker.load_template(
                self.save_workbook(workbook, "explicit-groups-required.xlsx")
            )

    def test_ensemble_is_an_ordinary_name_without_membership(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "DCA States"
        sheet.append(["DCA State", "Start Line Text", "DCA 7"])
        sheet.append(["Scene 1", "Opening", "MALE ENSEMBLE"])
        self.add_character_list(workbook, [(f"M{i}", None) for i in range(1,7)] + [("MALE ENSEMBLE", None)])
        states, assignments = marker.load_template(self.save_workbook(workbook))
        self.assertEqual(assignments["scene 1"], {"male ensemble": ["7"]})
        self.assertEqual(states[0]["dca_reference_rows"],
                         [{"dca": ["7"], "performer": "MALE ENSEMBLE", "roles": []}])
        self.assertEqual(states[0]["legend_assignments"], {"male ensemble": ["7"]})

    def test_ordinary_ensemble_inspector_uses_character_list_case(self):
        for name in ("MEN", "Women", "ChOrUs", "全体"):
            with self.subTest(name=name):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "DCA States"
                sheet.append(["DCA State", "Start Line Text", "DCA 1", "DCA 2"])
                sheet.append(["Scene 1", "Opening", name.lower(), name.upper()])
                self.add_character_list(workbook, [(name, None), ("M1", None)])
                states, assignments = marker.load_template(self.save_workbook(workbook))
                self.assertEqual(states[0]["dca_reference_rows"],
                                 [{"dca": ["1", "2"], "performer": name, "roles": []}])
                self.assertEqual(assignments["scene 1"][marker.normalise(name)], ["1", "2"])
                self.assertNotIn("m1", assignments["scene 1"])

    def test_direct_cell_names_keep_display_case_without_character_list_entries(self):
        for has_character_list in (False, True):
            with self.subTest(has_character_list=has_character_list):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "DCA States"
                sheet.append(["DCA State", "Start Line Text", "DCA 1", "DCA 2", "DCA 3"])
                sheet.append([
                    "Scene 1", "Opening",
                    "ALL THREE MEN [TRIO]\nALL MEN\nMiXeD Name",
                    "all three men\nall men",
                    "ALL THREE MEN\nALL MEN",
                ])
                if has_character_list:
                    self.add_character_list(workbook, [("Other Person", "Student")])
                states, assignments = marker.load_template(self.save_workbook(workbook))
                rows = {row["performer"]: row for row in states[0]["dca_reference_rows"]}
                self.assertEqual(set(rows), {"ALL THREE MEN", "ALL MEN", "MiXeD Name"})
                self.assertEqual(rows["ALL THREE MEN"]["dca"], ["1", "2", "3"])
                self.assertEqual(rows["ALL MEN"]["dca"], ["1", "2", "3"])
                self.assertEqual(rows["MiXeD Name"]["dca"], ["1"])
                self.assertTrue(all(row["roles"] == [] for row in rows.values()))
                self.assertEqual(assignments["scene 1"], {
                    "all three men": ["1", "2", "3"], "all men": ["1", "2", "3"],
                    "trio": ["1"], "mixed name": ["1"],
                })

    def test_role_shortcut_labels_keep_all_roles_and_survive_excel_roundtrip(self):
        project = marker.blank_project("Role labels")
        project["characters"] = [{
            "id": "m6", "dca_name": "M6",
            "other_characters": "张警官\n兄弟3\n手下1",
        }]
        project["states"] = [{
            "id": "s1", "name": "Scene 119", "start_line_text": "Opening",
            "dca_assignments": ["", "", "", "M6"],
        }]
        original = json.dumps(project, ensure_ascii=False)
        baseline_book = marker.project_to_workbook(project)
        try:
            _, baseline_assignments = marker._load_template_workbook(baseline_book)
        finally:
            baseline_book.close()
        for role in ["张警官", "兄弟3", "手下1"]:
            with self.subTest(role=role):
                selected = json.loads(original)
                label = marker.excel_role_choice("M6", role)
                selected["states"][0]["dca_assignments"][3] = label
                project_path = self.temporary_path / "selected.dcamarker"
                project_path.write_text(json.dumps(selected), encoding="utf-8")
                before = project_path.read_bytes()
                _, assignments = marker.load_project(project_path)
                self.assertEqual(assignments, baseline_assignments)
                self.assertEqual(project_path.read_bytes(), before)
                workbook_path = self.save_workbook(marker.project_to_workbook(selected))
                reopened = marker.import_excel_project(workbook_path)
                self.assertEqual(reopened["states"][0]["dca_assignments"][3], label)
        self.assertEqual(json.dumps(project, ensure_ascii=False), original)

    def test_project_inspector_preserves_direct_names_and_canonical_role_owner(self):
        project = marker.blank_project("Display case")
        project["characters"] = [{"id": "jack", "dca_name": "Jack [John]", "other_characters": "Student\nTeacher"}]
        project["states"] = [{
            "id": "s1", "name": "S1", "start_line_text": "Opening",
            "dca_assignments": ["ALL THREE", "ALL THREE\nSTUDENT", "ALL THREE"],
        }]
        path = self.temporary_path / "case.dcamarker"
        path.write_text(json.dumps(project), encoding="utf-8")
        original = path.read_bytes()
        states, assignments = marker.load_project(path)
        rows = {row["performer"]: row for row in states[0]["dca_reference_rows"]}
        self.assertEqual(set(rows), {"ALL THREE", "Jack [John]"})
        self.assertEqual(rows["ALL THREE"]["dca"], ["1", "2", "3"])
        self.assertEqual(rows["Jack [John]"]["dca"], ["2"])
        self.assertEqual(assignments["s1"]["teacher"], ["2"])
        self.assertEqual(path.read_bytes(), original)

    def test_character_entry_display_option_does_not_change_matching_defaults(self):
        value = "  ALL THREE [Trio]\n\nMiXeD Name\n全体  "
        expected = [("all three", ["trio"]), ("mixed name", []), ("全体", [])]
        self.assertEqual(marker.split_character_entries(value), expected)
        self.assertEqual(marker.split_character_entries(value, preserve_display_names=True),
                         [("ALL THREE", ["trio"]), ("MiXeD Name", []), ("全体", [])])
        self.assertEqual(marker.split_character_entries(None, preserve_display_names=True), [])

    def test_legacy_converted_member_in_another_dca_warns_but_is_allowed(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
            "DCA 2",
        ])
        states_sheet.append(["Scene 1", "Opening", "ALL", "Adam"])
        self.add_character_list(
            workbook,
            [("Adam", "ALL"), ("Jerry", "ALL")],
        )
        self.add_shared_groups(workbook, [("ALL", "Adam\nJerry")])

        diagnostics = {}
        states, assignments = marker.load_template(
            self.save_workbook(workbook, "shared-group-duplicate.xlsx"),
            diagnostics=diagnostics,
        )

        self.assertEqual(assignments["scene 1"]["adam"], ["1", "2"])
        self.assertEqual(assignments["scene 1"]["jerry"], ["1"])
        self.assertEqual(assignments["scene 1"]["all"], ["1"])
        self.assertEqual(
            diagnostics["duplicate_dca_assignments"],
            [{
                "state_key": "scene 1",
                "state_name": "Scene 1",
                "dca_name": "Adam",
                "dcas": ["1", "2"],
            }],
        )
        notices = marker.build_review_notices(
            states,
            assignments,
            marked_count=1,
            activated_states={"scene 1"},
            diagnostics=diagnostics,
        )
        duplicate_notice = next(
            notice
            for notice in notices
            if notice["code"] == "DUPLICATE_DCA_ASSIGNMENTS"
        )
        self.assertEqual(duplicate_notice["severity"], "warning")
        self.assertIn("generation continued", duplicate_notice["message"])

    def test_internal_dca_assignment_gap_creates_review_warning(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
            "DCA 2",
            "DCA 3",
            "DCA 4",
        ])
        states_sheet.append([
            "Scene 102",
            "Opening",
            "Ben",
            None,
            "Sam",
            None,
        ])
        diagnostics = {}

        states, assignments = marker.load_template(
            self.save_workbook(workbook, "assignment-gap-template.xlsx"),
            diagnostics=diagnostics,
        )

        self.assertEqual(
            diagnostics["assignment_gaps"],
            [{
                "state_key": "scene 102",
                "state_name": "Scene 102",
                "missing_dcas": [2],
                "first_dca": 1,
                "last_dca": 3,
            }],
        )
        notices = marker.build_review_notices(
            states,
            assignments,
            1,
            {"scene 102"},
            diagnostics=diagnostics,
        )
        gap_notice = next(
            notice
            for notice in notices
            if notice["code"] == "DCA_ASSIGNMENT_GAPS"
        )

        self.assertEqual(gap_notice["severity"], "warning")
        self.assertIn("Scene 102", gap_notice["message"])
        self.assertIn("DCA 2", gap_notice["message"])
        self.assertNotIn("DCA 4", gap_notice["message"])

    def test_role_that_reuses_another_dca_name_is_rejected(self):
        workbook = Workbook()
        states_sheet = workbook.active
        states_sheet.title = "DCA States"
        states_sheet.append([
            "DCA State",
            "Start Line Text",
            "DCA 1",
        ])
        states_sheet.append(["Scene 1", "Opening", "Mary"])
        self.add_character_list(
            workbook,
            [
                ("Mary", "Beth"),
                ("Beth", None),
            ],
        )
        template_file = self.save_workbook(
            workbook,
            "conflicting-dca-name-template.xlsx",
        )

        with self.assertRaisesRegex(
            ValueError,
            r'(?i)conflicts with a DCA Name.*"beth"',
        ):
            marker.load_template(template_file)


if __name__ == "__main__":
    unittest.main()
